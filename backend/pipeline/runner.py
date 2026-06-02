"""
Pipeline 狀態機主引擎。

流程：
  START → 逐步執行 → LLM 驗證 → 通過則下一步
                                 → 失敗且有重試次數 → 自動重試
                                 → 失敗且重試耗盡  → 暫停 + Telegram inline keyboard
  用戶按 [重試 / 跳過 / 中止] → resume_pipeline() 繼續或結束

Telegram 通知時機：
  - 步驟失敗需人為決策 → 詢問訊息 + inline keyboard
  - Pipeline 全部完成 / 中止 → 結果摘要
"""
import asyncio
import logging
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

from telegram import Bot, InlineKeyboardButton, InlineKeyboardMarkup

from config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID
from .models import PipelineConfig
from .store import PipelineRun, StepResult, get_store
from .logger import create_run_logger, resume_run_logger
from .executor import execute_step, execute_step_with_skill, execute_step_with_outlook, execute_step_with_web_crawler
from .validator import validate_step, validate_step_with_skill, ValidationResult


# ── Abort flags（in-memory）────────────────────────────────────────────────────
_abort_flags: set[str] = set()

# ── Running task tracking（for immediate cancel）──────────────────────────────
_running_tasks: dict[str, asyncio.Task] = {}


def register_task(run_id: str, task: asyncio.Task):
    _running_tasks[run_id] = task


def unregister_task(run_id: str):
    _running_tasks.pop(run_id, None)


def request_abort(run_id: str):
    """前端/API 呼叫：標記此 run 需要中止"""
    _abort_flags.add(run_id)


async def force_abort(run_id: str):
    """立即中止：kill 子進程 + 標記 computer_use abort + cancel asyncio task + 更新狀態"""
    from .executor import kill_run_processes
    from .computer_use import request_abort as _cu_abort
    _abort_flags.add(run_id)
    # 1. 立即 kill 所有子進程
    kill_run_processes(run_id)
    # 1a. 通知 computer_use 引擎中止（它跑在 executor thread 裡，kill 不到）
    _cu_abort(run_id)
    # 2. Cancel asyncio task
    task = _running_tasks.pop(run_id, None)
    if task and not task.done():
        task.cancel()
        try:
            await task
        except (asyncio.CancelledError, Exception):
            pass
    # 3. 更新 run 狀態
    store = get_store()
    run = store.load(run_id)
    if run and run.status in ("running", "awaiting_human"):
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info("⛔ Pipeline 被立即中止（force abort）")
        try:
            config = PipelineConfig.from_dict(run.config_dict)
            await _notify_final(run, config)
        except Exception:
            pass
    clear_abort(run_id)


def is_abort_requested(run_id: str) -> bool:
    return run_id in _abort_flags


def clear_abort(run_id: str):
    _abort_flags.discard(run_id)


# ── Telegram helpers ─────────────────────────────────────────────────────────

def _decision_keyboard(run_id: str, has_prev: bool = True) -> InlineKeyboardMarkup:
    # 步驟失敗時的決策鍵盤。
    # 💡 截圖按鈕:失敗時使用者可能人不在電腦前,加截圖按鈕讓遠端也能先看畫面再決策
    #    (不分節點類型 — skill 腳本 / 桌面自動化 / shell 失敗都可能需要看現場)
    # has_prev=True 顯示「重做上一步」按鈕;current_step=0 時隱藏
    rows = [
        [
            InlineKeyboardButton("🔄 重試", callback_data=f"pipe_retry:{run_id}"),
            InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"),
        ],
    ]
    skip_row = [InlineKeyboardButton("⏩ 跳過此步", callback_data=f"pipe_skip:{run_id}")]
    if has_prev:
        skip_row.append(InlineKeyboardButton("↩ 重做上一步", callback_data=f"pipe_redo_prev:{run_id}"))
    rows.append(skip_row)
    rows.extend([
        [
            InlineKeyboardButton("📸 截圖", callback_data=f"pipe_screenshot:{run_id}"),
            InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}"),
        ],
        [
            InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
        ],
    ])
    return InlineKeyboardMarkup(rows)


def _missing_dep_keyboard(run_id: str, packages: list[str]) -> InlineKeyboardMarkup:
    """缺套件時的決策鍵盤。每個套件一行『允許安裝』按鈕；最下方拒絕 / 改任務 / 中止。
    callback_data: pipe_install_dep:{run_id}:{pkg_name}
    """
    rows = []
    for pkg in packages[:5]:  # 最多 5 個（callback_data 64 byte 限制）
        rows.append([InlineKeyboardButton(
            f"✅ 允許安裝 {pkg}",
            callback_data=f"pipe_install_dep:{run_id}:{pkg}",
        )])
    # 一次允許全部（多個套件時）
    if len(packages) > 1:
        rows.append([InlineKeyboardButton(
            f"✅ 全部安裝（{len(packages)} 個）",
            callback_data=f"pipe_install_all:{run_id}",
        )])
    rows.append([
        InlineKeyboardButton("💬 改任務", callback_data=f"pipe_hint:{run_id}"),
        InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
    ])
    rows.append([InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}")])
    return InlineKeyboardMarkup(rows)


def _confirm_keyboard(run_id: str, screenshot: bool = False, allow_hint: bool = False,
                      preview_enabled: bool = False) -> InlineKeyboardMarkup:
    # 人工確認節點的按鈕。allow_hint 只在「上一個可執行節點是 AI 技能（skill_mode）」時 True。
    # preview_enabled 只在 step.preview_prev_output=True 時 True（=自動預覽有啟用才給 HQ 選項）
    top = [InlineKeyboardButton("✅ 繼續執行", callback_data=f"pipe_continue:{run_id}")]
    if allow_hint:
        top.append(InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"))
    top.append(InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"))
    rows = [
        top,
        # 「📎 上一步輸出 / 📂 任一步輸出」永遠存在；點下去由 backend 判斷有沒有檔案、回應使用者
        # 跟 send_prev_output 自動傳送無關（自動傳是抵達節點當下推一份；按鈕是隨時要重抓用）
        [InlineKeyboardButton("📎 上一步輸出", callback_data=f"pipe_prev_output:{run_id}"),
         InlineKeyboardButton("📂 任一步輸出", callback_data=f"pipe_select_step:{run_id}")],
        [InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}")],
    ]
    if screenshot:
        rows[2].append(InlineKeyboardButton("📸 截圖", callback_data=f"pipe_screenshot:{run_id}"))
    # HQ 預覽：B1 自動預覽只抽文字（docx/pptx 品質 40%），點此按鈕改用 LibreOffice
    # 轉 PDF → render，版式 ~80-90% 還原。要 5-10s，所以不自動跑；使用者按了才跑。
    if preview_enabled:
        rows.append([InlineKeyboardButton("🎨 原版式預覽（LibreOffice）",
                                          callback_data=f"pipe_preview_hq:{run_id}")])
    return InlineKeyboardMarkup(rows)


def _ask_user_keyboard(run_id: str, options: list) -> InlineKeyboardMarkup:
    """
    ask_user 問題的 Telegram 鍵盤。
    - 有 options:
        - 任一選項長度 > 14 → 改一行一個按鈕(長文不被 TG 攔腰截斷)
        - 否則維持一行兩個按鈕(短選項更省螢幕)
    - 無 options → 只有「自由輸入」和「中止」
    """
    rows: list[list[InlineKeyboardButton]] = []
    if options:
        max_len = max(len(str(o)) for o in options)
        one_per_row = max_len > 14
        per_row = 1 if one_per_row else 2
        label_cap = 60 if one_per_row else 30
        row: list[InlineKeyboardButton] = []
        for i, opt in enumerate(options):
            label = str(opt)
            if len(label) > label_cap:
                label = label[: label_cap - 1] + "…"
            row.append(InlineKeyboardButton(label, callback_data=f"pipe_answer:{run_id}:{i}"))
            if len(row) == per_row:
                rows.append(row)
                row = []
        if row:
            rows.append(row)
    rows.append([
        InlineKeyboardButton("✍ 自由輸入", callback_data=f"pipe_answer_free:{run_id}"),
        InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
    ])
    return InlineKeyboardMarkup(rows)


async def _send_ask_user_notification(run, question: str, options: list, context: str, step_name: str):
    """Skill agent 呼叫 ask_user 時發送 Telegram 通知。"""
    import html as _html
    total = len(run.config_dict.get("steps", [])) if run.config_dict else 0
    step_num = run.current_step + 1
    lines = [
        "❓ <b>AI 技能請求人工協助</b>",
        "",
        f"📋 {run.pipeline_name}",
        f"📍 步驟 {step_num}/{total}：<b>{_html.escape(step_name)}</b>",
        "",
        f"<b>問題</b>：{_html.escape(question)}",
    ]
    if context:
        lines.append(f"\n<b>背景</b>：{_html.escape(context)}")
    if options:
        lines.append("\n請從下方選項選擇，或點「自由輸入」回答。")
    else:
        lines.append("\n請點「自由輸入」並傳送文字回答。")
    await _tg_send(run.telegram_chat_id, "\n".join(lines),
                   _ask_user_keyboard(run.run_id, options))


def _command_approval_keyboard(run_id: str) -> InlineKeyboardMarkup:
    """ask_mode 敏感命令授權的 TG 按鈕。"""
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ 允許執行", callback_data=f"pipe_approve_cmd:{run_id}"),
        ],
        [
            InlineKeyboardButton("❌ 拒絕並中止", callback_data=f"pipe_deny_cmd:{run_id}"),
            InlineKeyboardButton("💬 改任務", callback_data=f"pipe_hint_cmd:{run_id}"),
        ],
    ])


async def _send_command_approval_notification(
    run, category: str, label: str, preview: str, step_name: str,
):
    """ask_mode 敏感命令攔截時送 TG 訊息 + inline keyboard。"""
    import html as _html
    total = len(run.config_dict.get("steps", [])) if run.config_dict else 0
    step_num = run.current_step + 1
    cat_emoji = {
        "destructive":      "🗑",
        "privileged":       "🔐",
        "remote-exec":      "🌐",
        "install":          "📦",
        "subprocess":       "⚙",
        "outside-workflow": "📁",
    }.get(category, "🛡")
    lines = [
        f"{cat_emoji} <b>敏感命令需授權（ask 模式）</b>",
        "",
        f"📋 {run.pipeline_name}",
        f"📍 步驟 {step_num}/{total}:<b>{_html.escape(step_name)}</b>",
        "",
        f"<b>類型</b>:{_html.escape(label)}（<code>{category}</code>）",
        f"<b>命令</b>:<code>{_html.escape(preview)}</code>",
        "",
        "請選擇授權方式:",
    ]
    await _tg_send(run.telegram_chat_id, "\n".join(lines),
                   _command_approval_keyboard(run.run_id))


def _is_valid_tg_token(token: str) -> bool:
    """檢查 Telegram Bot Token 格式是否正確（數字:字母混合）"""
    if not token or ":" not in token:
        return False
    parts = token.split(":", 1)
    return parts[0].isdigit() and len(parts[1]) > 10


def _get_tg_token() -> str:
    """取得 Telegram Bot Token（優先用 settings UI 設定，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        token = get_settings().get("telegram_bot_token", "")
        if token and _is_valid_tg_token(token):
            return token
        elif token:
            logger.warning(f"[Telegram] settings 中的 token 格式不正確（'{token[:15]}...'），改用 .env")
    except Exception:
        pass
    if TELEGRAM_BOT_TOKEN:
        logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_BOT_TOKEN")
    return TELEGRAM_BOT_TOKEN


def _get_tg_chat_id() -> int:
    """取得 Telegram Chat ID（優先 settings UI，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        cid = get_settings().get("telegram_chat_id", "")
        if cid:
            return int(cid)
    except Exception:
        pass
    # fallback 到 .env
    if TELEGRAM_CHAT_ID:
        try:
            logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_CHAT_ID")
            return int(TELEGRAM_CHAT_ID)
        except ValueError:
            logger.warning(f"[Telegram] .env TELEGRAM_CHAT_ID 格式不正確：{TELEGRAM_CHAT_ID}")
    return 0


async def _tg_send(chat_id: int, text: str, reply_markup=None):
    """發送 Telegram 訊息（錯誤靜默記錄，不拋出）"""
    logger = logging.getLogger("pipeline")
    token = _get_tg_token()
    # 如果沒傳 chat_id，嘗試從 settings 取得
    if not chat_id:
        chat_id = _get_tg_chat_id()
    if not chat_id or not token:
        logger.warning(f"[Telegram] 跳過發送：chat_id={chat_id}, token={'有' if token else '無'}")
        return
    logger.info(f"[Telegram] 發送訊息到 chat_id={chat_id}（token={token[:15]}...）")
    try:
        # 用 async with 取代手動 bot.close() — 後者實際是 TG API 的 `close` method
        # （TG 文件警告：前 10 分鐘必回 429、嚴格 rate-limit、不該在 bot code 呼叫）
        # async with 走 shutdown() 路徑、只關 httpx 連線、不打 TG API
        async with Bot(token=token) as bot:
            await bot.send_message(
                chat_id=chat_id,
                text=text,
                parse_mode="HTML",
                reply_markup=reply_markup,
            )
        logger.info(f"[Telegram] ✅ 發送成功")
    except Exception as e:
        logger.error(f"[Telegram] ❌ 發送失敗：{e}")


# 送 TG photo 的壓縮參數：一律壓縮（不看原檔大小），讓每張 traffic 一致、傳輸時間接近
# → 避免「大的壓了變小、小的沒壓還是大」的不對稱上傳時間造成誤判 timeout + 重複訊息
# 長寬上限：1920（TG 本來就會壓到 ~1280 顯示，1920 已經足夠清楚，肉眼看不出差）
_TG_PHOTO_MAX_DIM = 1920
_TG_PHOTO_JPEG_Q  = 85


def _compress_for_tg(src_path: str) -> str:
    """一律轉 JPEG + 縮邊到 _TG_PHOTO_MAX_DIM。回傳新產生的 _compressed.jpg 路徑。
    Pillow 缺席 / 讀圖失敗 → 回原路徑當 fallback。
    為什麼不看門檻：上次 bug 就是 mon1 沒壓（大）+ mon2 壓了（小）→ mon1 上傳慢 120s 超時誤判。
    統一都壓就沒這問題，而且 TG 顯示時本來就壓到 ~1280，我們先壓 1920 剛剛好。
    """
    logger = logging.getLogger("pipeline")
    try:
        from pathlib import Path as _P
        src = _P(src_path)
        if not src.exists():
            return src_path
        orig_size = src.stat().st_size
        try:
            from PIL import Image
        except Exception:
            logger.warning(f"[Telegram] Pillow 未安裝，照原圖送 {src.name}（{orig_size/1024/1024:.1f}MB）")
            return src_path
        im = Image.open(src)
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        w, h = im.size
        if max(w, h) > _TG_PHOTO_MAX_DIM:
            scale = _TG_PHOTO_MAX_DIM / max(w, h)
            im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        out = src.with_suffix("")
        out = out.with_name(out.name + "_compressed.jpg")
        im.save(out, "JPEG", quality=_TG_PHOTO_JPEG_Q, optimize=True)
        new_size = out.stat().st_size
        logger.info(
            f"[Telegram] 壓縮 {src.name}：{orig_size/1024/1024:.2f}MB ({w}×{h}) "
            f"→ {out.name}：{new_size/1024/1024:.2f}MB {im.size}"
        )
        return str(out)
    except Exception as e:
        logger.warning(f"[Telegram] 壓縮失敗（照原圖送）：{e}")
        return src_path


async def _tg_send_photos(chat_id: int, paths: list[str], caption_prefix: str = ""):
    """批次送截圖。每張 caption「螢幕 k/n」。
    流程：超過 4.5MB → Pillow 壓到 JPEG+1920 邊再送；send_photo 還是失敗 → 退 send_document。
    每張分開 try/except，一張壞不會拖垮整批。
    """
    logger = logging.getLogger("pipeline")
    if not paths:
        return
    token = _get_tg_token()
    if not chat_id:
        chat_id = _get_tg_chat_id()
    if not chat_id or not token:
        logger.warning(f"[Telegram] 批次截圖跳過：chat_id={chat_id} token={'有' if token else '無'}")
        return
    logger.info(f"[Telegram] 批次送 {len(paths)} 張截圖 → chat_id={chat_id}")

    # send 每張的硬超時：避免 send_photo hang 把 poll loop 整個卡死
    _PHOTO_TIMEOUT_S = 120

    # 重複訊息 root cause：timeout / network error 時 Python 以為失敗，但 TG 其實已收到，
    # 我們又送了一次 document → 使用者收到 2 份同內容。
    # 修法：只在「確定 TG 拒收這個檔」（BadRequest，例如格式/尺寸錯）時才 fallback document；
    #       timeout / NetworkError / TimedOut 都視為「很可能已送達」不重送。
    from telegram.error import BadRequest as _TgBadRequest  # noqa: WPS433 (局部 import 沒關係)

    async def _send_one(bot, send_path: str, cap: str, i: int, total: int) -> bool:
        """送單張。回傳 True=已送達（或近乎送達），False=徹底失敗。"""
        # 1) 先試 send_photo
        try:
            with open(send_path, "rb") as fh:
                await asyncio.wait_for(
                    bot.send_photo(chat_id=chat_id, photo=fh, caption=cap or None),
                    timeout=_PHOTO_TIMEOUT_S,
                )
            logger.info(f"[Telegram]   ✓ 送出截圖 {i}/{total}")
            return True
        except _TgBadRequest as e:
            # 檔案格式 / 尺寸被 TG 拒收 — 真的壞了，退 document 才有意義
            logger.warning(f"[Telegram]   photo {i}/{total} TG 拒收（BadRequest），退 document：{e}")
        except asyncio.TimeoutError:
            logger.warning(
                f"[Telegram]   photo {i}/{total} 超過 {_PHOTO_TIMEOUT_S}s 沒回 ack，"
                f"TG 可能已收到（不重送 document 避免重複）"
            )
            return True
        except Exception as e:
            # network / httpx timeout / retry-after 等 — TG 是否收到不確定；
            # 為避免重複訊息，一律視為「可能已送達」不 fallback（之前 case 就是這裡誤判）
            logger.warning(
                f"[Telegram]   photo {i}/{total} 送出時出例外（{type(e).__name__}: {e}）— "
                f"TG 可能已收到，不重送 document 以避免重複"
            )
            return True
        # 2) Fallback：send_document（只在真的 BadRequest 時才走）
        try:
            with open(send_path, "rb") as fh:
                await asyncio.wait_for(
                    bot.send_document(chat_id=chat_id, document=fh, caption=cap or None),
                    timeout=_PHOTO_TIMEOUT_S,
                )
            logger.info(f"[Telegram]   ✓ 以 document 形式送出截圖 {i}/{total}")
            return True
        except asyncio.TimeoutError:
            logger.warning(f"[Telegram]   document {i}/{total} 超時但 TG 可能已收到")
            return True
        except Exception as e2:
            logger.error(f"[Telegram]   ✗ 截圖 {i}/{total} 送出徹底失敗：{type(e2).__name__}: {e2}")
        return False

    try:
        # 用 async with：避免手動 bot.close()（那是 TG API、嚴格 rate-limit）
        async with Bot(token=token) as bot:
            total = len(paths)
            for i, p in enumerate(paths, start=1):
                cap = caption_prefix + (f"（螢幕 {i}/{total}）" if total > 1 else "")
                send_path = _compress_for_tg(p)
                try:
                    sz = os.path.getsize(send_path)
                except Exception as e:
                    logger.error(f"[Telegram]   截圖 {i}/{total} 檔案讀取失敗（{e}）→ 跳過")
                    continue
                if sz <= 0:
                    logger.error(f"[Telegram]   截圖 {i}/{total} 檔案 0 bytes → 跳過")
                    continue
                ok = await _send_one(bot, send_path, cap, i, total)
                if ok:
                    for cleanup in {p, send_path}:
                        try:
                            if os.path.exists(cleanup):
                                os.unlink(cleanup)
                        except Exception as _e:
                            logger.warning(f"[Telegram]   清理截圖 {cleanup} 失敗：{_e}")
    except Exception as e:
        logger.error(f"[Telegram] 截圖批次送出異常：{e}")


# Telegram 文件大小上限（一般 bot；自架 local server 可放寬到 2GB）
_TG_DOC_MAX_BYTES = 50 * 1024 * 1024


def _workflow_output_dir(workflow_name: str):
    """回傳 <OUTPUT_BASE_PATH>/<workflow_name>/ 的絕對路徑(不存在也回、由呼叫端處理)。

    重要:統一用 config.OUTPUT_BASE_PATH、不再從 __file__ 推 proj_root + 'ai_output',
    避免 chat_tools / send_file_to_tg 用 OUTPUT_BASE_PATH 找檔、workflow 卻寫到別處的 bug。
    """
    if not workflow_name:
        return None
    from config import OUTPUT_BASE_PATH
    return OUTPUT_BASE_PATH / workflow_name


# 副檔名猜測:batch 含哪個關鍵字 → 推測該副檔名(沒填 step.output.path 時用)
_EXT_KEYWORDS: list[tuple[str, str]] = [
    (".pptx", "pptx"), (".pptx", "投影片"), (".pptx", "簡報"), (".pptx", "PPT"), (".pptx", "ppt"),
    (".docx", "docx"), (".docx", "Word"), (".docx", "word"),
    (".xlsx", "xlsx"), (".xlsx", "excel"), (".xlsx", "Excel"), (".xlsx", "表格"),
    (".csv", "csv"), (".csv", "CSV"),
    (".png", ".png"), (".png", "圖表"), (".png", "折線圖"), (".png", "長條圖"), (".png", "散佈"), (".png", "繪圖"),
    (".pdf", ".pdf"), (".pdf", "PDF"),
    (".json", "JSON"), (".json", ".json"),
    (".html", ".html"), (".html", "HTML"),
]


def _safe_step_filename(step_name: str) -> str:
    """把中文 / 空白 / 特殊字元的 step_name 轉成安全檔名 base(無副檔名)。"""
    import re as _re
    safe = _re.sub(r"[^\w一-鿿-]+", "_", (step_name or "").strip())
    safe = safe.strip("_") or "step_output"
    return safe[:60]   # 太長截掉


def _derive_default_output_path(step, workflow_dir: str) -> tuple[str, str]:
    """step.output.path 沒設時、自動推一個合理路徑給 LLM 知道存哪。

    Returns: (resolved_abs_path, ext_used)
    """
    from pathlib import Path as _P
    # 猜副檔名:掃 batch 找關鍵字、找不到預設 .md(skill 寫產物最常用 markdown)
    batch_lower = (step.batch or "").lower()
    ext = ".md"
    for _ext, kw in _EXT_KEYWORDS:
        if kw.lower() in batch_lower:
            ext = _ext
            break
    fname = _safe_step_filename(step.name) + ext
    return str(_P(workflow_dir).absolute() / fname), ext


# 用來判斷哪些檔案是「真正的步驟產出」、哪些是雜訊（log / preview / 內部 DB 檔）
_OUTPUT_SKIP_PREFIXES = ("screenshot_",)
_OUTPUT_SKIP_SUFFIXES = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
_OUTPUT_SKIP_EXTS = {".log"}
_OUTPUT_SKIP_NAMES = ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal")


def _is_output_candidate(path) -> bool:
    """檔名是否該被視為「步驟可能產出的檔案」。過濾掉系統雜訊。"""
    n = path.name
    if n.startswith(_OUTPUT_SKIP_PREFIXES):
        return False
    if any(n.endswith(suf) for suf in _OUTPUT_SKIP_SUFFIXES):
        return False
    if path.suffix.lower() in _OUTPUT_SKIP_EXTS:
        return False
    if n in _OUTPUT_SKIP_NAMES:
        return False
    return True


def _snapshot_workflow_dir(workflow_name: str) -> dict:
    """掃 ai_output/<workflow_name>/ 取每個檔的 mtime（給步驟前後比對用）。
    回 {str(absolute_path): mtime}。失敗回空 dict。"""
    out: dict = {}
    wf = _workflow_output_dir(workflow_name)
    if not wf or not wf.exists() or not wf.is_dir():
        return out
    try:
        for f in wf.rglob("*"):
            if f.is_file() and _is_output_candidate(f):
                try:
                    out[str(f.absolute())] = f.stat().st_mtime
                except OSError:
                    pass
    except Exception:
        pass
    return out


def _diff_snapshot_pick_main(before: dict, workflow_name: str):
    """比對 before（_snapshot_workflow_dir 結果）跟現在的狀態、找出本步驟新增/修改的檔，
    挑「主要產出」回傳（絕對路徑字串），沒有變化回 None。

    挑選邏輯：
      1. 先看新增檔（before 沒有的） — 比修改現有檔更可能是「最終產出」
      2. 都沒新增、看修改的（mtime 變新）
      3. 多個候選時，偏好「報告類副檔名」（.docx/.pdf/.xlsx/.md/.csv/.html/.pptx/.json/.txt）
         排在前面、再看 mtime 最新的
    """
    wf = _workflow_output_dir(workflow_name)
    if not wf or not wf.exists():
        return None
    # 取現在快照
    after = _snapshot_workflow_dir(workflow_name)

    new_files = [p for p in after.keys() if p not in before]
    modified = [p for p in after.keys() if p in before and after[p] > before[p]]
    candidates = new_files if new_files else modified
    if not candidates:
        return None

    from pathlib import Path as _P
    # 報告類副檔名優先（排前面）
    report_exts = {".docx", ".pdf", ".xlsx", ".xls", ".pptx", ".md", ".csv",
                   ".html", ".htm", ".json", ".txt", ".png", ".jpg", ".jpeg"}

    def sort_key(p_str: str):
        p = _P(p_str)
        is_report = 0 if p.suffix.lower() in report_exts else 1
        # 用負 mtime 讓「最新」排前面
        return (is_report, -after.get(p_str, 0))

    candidates.sort(key=sort_key)
    return candidates[0]


def _crawl_looks_failed(output_path, logger):
    """掃爬蟲輸出檔(crawler 寫的 .md 帶 frontmatter status_code / word_count / 子頁數)。
    回失敗原因字串(明顯抓失敗)或 None(看起來 OK)。**保守**:只在明確壞掉時回原因、
    避免誤殺正常爬蟲。給 web_crawler 沒填 expect 時的確定性把關用。"""
    from pathlib import Path as _P
    import re as _re
    try:
        p = _P(output_path) if output_path else None
        if not p or not p.exists():
            return None  # 沒檔由其他邏輯處理、這裡不擅自判失敗
        texts = []
        if p.is_dir():
            for f in list(p.rglob("*.md"))[:30]:
                try:
                    texts.append(f.read_text(encoding="utf-8", errors="replace"))
                except Exception:
                    pass
        else:
            texts.append(p.read_text(encoding="utf-8", errors="replace"))
        blob = "\n".join(texts)
        if not blob.strip():
            return "爬蟲輸出為空"
        # 子頁全失敗(多 URL 列表頁爬蟲會印「子頁數: N/M 成功」)
        m = _re.search(r"子頁數[:：]\s*0\s*/\s*\d+", blob)
        if m:
            return "所有子頁抓取失敗(0 成功)"
        # 所有抓取的 status_code 都是 4xx/5xx → 全失敗(部分成功就放行)
        codes = _re.findall(r"status_code[:：]\s*(\d{3})", blob)
        if codes and all(c[0] in ("4", "5") for c in codes):
            return f"所有抓取都失敗(status_code={', '.join(sorted(set(codes)))})"
        return None
    except Exception as e:
        try:
            logger.warning(f"爬蟲輸出檢查失敗(略過、不擋):{e}")
        except Exception:
            pass
        return None


def _latest_workflow_output_file(workflow_name: str):
    """掃 ai_output/<workflow_name>/ 拿最新一個非雜訊檔（圖檔 / log / preview / 內部 db 檔過濾掉）。
    給「skill 節點 / 沒明確 output.path」這種「實際有產檔但 step 沒記錄」的場景兜底。
    """
    from pathlib import Path as _P
    if not workflow_name:
        return None
    wf_dir = _workflow_output_dir(workflow_name)
    if not wf_dir.exists() or not wf_dir.is_dir():
        return None
    skip_prefixes = ("screenshot_",)
    skip_suffixes = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
    skip_exts = {".log"}
    skip_names = ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal")
    candidates = []
    for f in wf_dir.iterdir():
        if not f.is_file():
            continue
        n = f.name
        if n.startswith(skip_prefixes):
            continue
        if any(n.endswith(suf) for suf in skip_suffixes):
            continue
        if f.suffix.lower() in skip_exts:
            continue
        if n in skip_names:
            continue
        candidates.append((f.stat().st_mtime, f))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def _step_default_output_path(step, workflow_name: str) -> Optional[str]:
    """如果 step 沒明確設 output.path、回傳 runner 會自動 default 的路徑。
    跟 runner.py 內 outlook_automation / web_crawler 的 default 邏輯保持一致。

    沒 default 規則的節點類型（skill / script / human_confirm 等）回 None。
    """
    if not workflow_name:
        return None
    # 哪些節點類型有 default rule
    has_default = bool(
        getattr(step, "outlook_automation", False)
        or getattr(step, "web_crawler", False)
    )
    if not has_default:
        return None
    import re as _re
    safe_step = _re.sub(r"[^\w一-鿿_-]", "_", step.name).strip("_") or "result"
    return f"ai_output/{workflow_name}/{safe_step}_result.md"


def _resolve_step_output_for_tg(
    step, *, workflow_name: str = "", logger=None, step_result=None,
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """檢查指定 step 的輸出檔、回傳能傳給 Telegram 的檔案資料。

    解析優先順序：
      1. step.output.path 明確設定 → 用它
      2. 沒設、但節點類型有 default rule → 算出 default 路徑
      3. default 路徑不存在但 parent 是 workflow 輸出資料夾、裡面有東西
         → 多 URL 爬蟲常見、實際輸出在資料夾不是單檔；改用 parent dir
      4. 都沒有 → None（呼叫端組「無輸出檔」訊息）

    回傳 (file_path, display_name, error_msg)：
      - 三者皆 None：完全沒輸出可傳
      - file_path / display_name 給值、error_msg=None：可以傳
      - error_msg 有值：能解析但有狀況（不存在 / 太大），傳給使用者看

    資料夾：自動 zip 後給；zip 寫到系統 temp、呼叫端用完要 unlink。
    """
    from pathlib import Path as _P
    import logging as _log

    log_fn = (logger or _log.getLogger("pipeline")).warning

    if not step:
        return None, None, None

    # ── Step 1：抓明確或 default 的目標路徑 ──────────────────────
    target_str: Optional[str] = None
    is_explicit = False
    # 最高優先：StepResult.actual_output_path（runner 在執行時 snapshot diff 算出來的）
    # 這個對「沒設 output.path 的 skill 節點」特別重要 — 不會跟其他步驟搶 latest 檔
    actual_path = getattr(step_result, "actual_output_path", "") if step_result else ""
    if actual_path:
        target_str = actual_path
        is_explicit = True  # 視為「指定路徑」、後面 multi-URL fallback 不要動它
    elif getattr(step, "output", None) and step.output.path:
        target_str = str(step.output.path)
        is_explicit = True
    else:
        target_str = _step_default_output_path(step, workflow_name)

    # Skill / Script 沒 actual_output_path、沒明確 output.path 也沒 default rule，
    # 通常表示：(a) 該步沒實際寫檔（如純 stdout）；或 (b) 是舊 run（升級前沒記錄 actual_output_path）
    # 對 (b) 兜底掃 workflow dir 最新檔（不完美但聊勝於無）
    # human_confirm / visual_validation / computer_use 不寫檔、不適用。
    if not target_str and workflow_name:
        could_produce = bool(
            getattr(step, "skill_mode", False)
            or (not getattr(step, "human_confirm", False)
                and not getattr(step, "visual_validation", False)
                and not getattr(step, "computer_use", False)
                and getattr(step, "batch", ""))  # script 節點：有 batch 表示會跑
        )
        if could_produce:
            latest = _latest_workflow_output_file(workflow_name)
            if latest:
                target_str = str(latest)
                if logger:
                    logger.info(f"[_resolve_step_output_for_tg] step={step.name} 沒設 output、"
                                f"fallback 到 workflow dir 最新檔：{latest.name}")

    if not target_str:
        return None, None, None

    p = _P(target_str).expanduser()
    if not p.is_absolute():
        p = _P(__file__).parent.parent.parent.absolute() / p

    if not p.exists():
        # ── Step 2：default 路徑檔案不存在、看 parent dir 有沒有東西 ──
        # 多 URL 爬蟲：default 路徑是 .md、實際輸出寫到 parent 資料夾（多檔 + index.json）
        # 多 URL 模式判斷：step.web_crawler + wc_urls 有效項數 > 1
        is_multi_wc = False
        if getattr(step, "web_crawler", False):
            wc_urls = getattr(step, "wc_urls", None) or []
            valid = [u for u in wc_urls if u and u.strip() and not u.strip().startswith("#")]
            is_multi_wc = len(valid) > 1
        # 沒明確設 output.path、且是多 URL 爬蟲、且 parent 有東西 → 用 parent
        if not is_explicit and is_multi_wc and p.parent.exists() and p.parent.is_dir():
            non_chrome = [
                f for f in p.parent.iterdir()
                if f.is_file() and not f.name.startswith("screenshot_") and f.suffix.lower() not in {".log"}
            ]
            if non_chrome:
                p = p.parent  # 用 workflow 輸出資料夾
            else:
                return None, None, f"輸出檔案不存在：{p}"
        else:
            return None, None, f"輸出檔案不存在：{p}"

    if p.is_file():
        size = p.stat().st_size
        if size > _TG_DOC_MAX_BYTES:
            return None, None, (f"檔案太大（{size/1024/1024:.1f} MB > 50 MB Telegram 上限）。"
                                f"請去 host 取：{p}")
        return str(p), p.name, None

    if p.is_dir():
        # 整個資料夾打包成 zip 送
        import tempfile, zipfile
        try:
            tmp_zip = tempfile.NamedTemporaryFile(suffix=".zip", delete=False).name
            with zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in p.rglob("*"):
                    if f.is_file():
                        # 用 dir 自身為 root；解壓會看到原本資料夾結構
                        zf.write(f, f.relative_to(p.parent))
            zsize = _P(tmp_zip).stat().st_size
            if zsize > _TG_DOC_MAX_BYTES:
                try: os.unlink(tmp_zip)
                except Exception: pass
                return None, None, (f"資料夾打包後 {zsize/1024/1024:.1f} MB、超過 Telegram 上限。"
                                    f"請去 host 取：{p}")
            return tmp_zip, p.name + ".zip", None
        except Exception as e:
            log_fn(f"[_resolve_step_output_for_tg] zip 失敗：{e}")
            return None, None, f"資料夾打包失敗：{e}"

    return None, None, f"不認識的路徑類型：{p}"


#
# UTF-8 BOM injection 共用 helper(給 runner + chat_tools 兩處 send_document 用)
#
# 為什麼要這層:iOS Telegram 文件預覽器若拿到沒 BOM 的純 UTF-8 文字檔,
# 在繁中環境會被自動判定成 Big5 / CP950 → 顯示成 mojibake。
# 桌面端、Web 端、Android 不受影響(都會猜 UTF-8)。為求一致、在送 .md / .txt /
# .csv / .json / .yaml / .html / .log 時注入 BOM。
# 寫到 system temp、原檔不污染。caller 負責 cleanup return 的 temp 路徑。
#
_TG_TEXT_EXTS_FOR_BOM = {".md", ".txt", ".csv", ".json", ".yaml", ".yml", ".html", ".htm", ".log"}


def _prepare_tg_file_with_bom(file_path: str, log: logging.Logger | None = None,
                               display: str | None = None) -> tuple[str, str | None]:
    """檢查 file_path 副檔名是否為文字檔、是的話寫一個含 UTF-8 BOM 的 temp 給 TG send。

    Args:
        file_path: 來源檔絕對路徑
        log: optional logger
        display: 顯示給 log / caption 用的檔名(可選、預設 file_path 的 basename)

    Returns:
        (send_path, temp_to_cleanup)
        - send_path:呼叫端 open() + send 用的路徑(可能是 file_path 本身、也可能是 temp)
        - temp_to_cleanup:如果有產 temp 就回 path、caller send 完後 os.unlink();
          沒產 temp 回 None
    """
    import tempfile as _tf
    _ext = Path(file_path).suffix.lower()
    if _ext not in _TG_TEXT_EXTS_FOR_BOM:
        return file_path, None
    _label = display or Path(file_path).name
    try:
        with open(file_path, "rb") as _f:
            _bytes = _f.read()
        # 已有 BOM 不重複加
        if _bytes.startswith(b"\xef\xbb\xbf"):
            return file_path, None
        _fd, _tmp = _tf.mkstemp(suffix=_ext)
        os.close(_fd)
        with open(_tmp, "wb") as _wf:
            _wf.write(b"\xef\xbb\xbf" + _bytes)
        if log:
            log.info(f"[Telegram] 文字檔加 UTF-8 BOM 避免 iOS 解碼成 Big5 亂碼:{_label}")
        return _tmp, _tmp
    except Exception as _e:
        if log:
            log.warning(f"[Telegram] 加 BOM 失敗、用原檔送:{_e}")
        return file_path, None


async def _send_step_output_to_tg(
    chat_id: int, step, step_label: str = "", *,
    workflow_name: str = "", logger=None, step_result=None,
) -> tuple[bool, str]:
    """把 step 的 output 檔案（或 zip 化的資料夾）送到指定 chat_id。
    workflow_name 用來算 default output path（沒填 output.path 但節點類型有 default 時）。
    step_result：StepResult 物件（含 actual_output_path）— 多步 skill 共用 workflow dir 時不會搶錯檔。
    回傳 (ok, msg)：ok=True 時 msg 是 status 摘要；False 時 msg 是錯誤訊息（可直接給使用者）。
    """
    import logging as _log
    import tempfile as _tf
    log = logger or _log.getLogger("pipeline")
    file_path, display, err = _resolve_step_output_for_tg(
        step, workflow_name=workflow_name, logger=log, step_result=step_result,
    )

    if err:
        return False, err
    if not file_path:
        return False, "上一步沒設輸出檔（output.path），且該節點類型沒有 default fallback"

    # chat_id fallback：跟 _tg_send / _tg_send_photos 同邏輯。
    # 之前 auto-send 會失敗的就是這裡 — chat_id 0 時沒退到 settings 拿
    if not chat_id:
        chat_id = _get_tg_chat_id()
    token = _get_tg_token()
    if not chat_id or not token:
        return False, f"Telegram 設定不完整：chat_id={chat_id or '無'}, token={'有' if token else '無'}"

    try:
        # 文字檔送 TG 前加 UTF-8 BOM 避免 iOS Telegram 解成 Big5 亂碼(共用 helper)
        _send_path, _temp_to_cleanup = _prepare_tg_file_with_bom(file_path, log, display)
        # 用 async with:避免手動 bot.close()(那是 TG `close` API method、
        # 文件寫前 10 分鐘必回 429、嚴格 rate-limit、不該在 bot code 呼叫)。
        # 之前 user 報「每次手動點按鈕必出現速率限制警告」就是這個 bug。
        async with Bot(token=token) as bot:
            with open(_send_path, "rb") as fp:
                await bot.send_document(
                    chat_id=chat_id,
                    document=fp,
                    filename=display,
                    caption=f"📎 {step_label or '上一步輸出'}:{display}" if step_label else None,
                )
        if _temp_to_cleanup:
            try: os.unlink(_temp_to_cleanup)
            except Exception: pass
        log.info(f"[Telegram] ✅ 已傳送 {display} 到 chat {chat_id}")
        return True, f"已傳送:{display}"
    except Exception as e:
        # Telegram rate limit (flood control) → 翻譯成易懂訊息
        es = str(e)
        if "Flood control" in es or "RetryAfter" in es or "Too Many Requests" in es:
            import re as _re
            m = _re.search(r"(\d+)\s*seconds", es)
            wait_s = int(m.group(1)) if m else 0
            log.warning(f"[Telegram] FloodWait：{es}")
            return False, (
                f"Telegram 速率限制（短時間內傳太多訊息了）。"
                f"{f'請等 {wait_s} 秒（約 {wait_s//60} 分 {wait_s%60} 秒）後再試。' if wait_s else '稍候幾分鐘再試。'}"
            )
        log.error(f"[Telegram] send_document 失敗：{e}")
        return False, f"Telegram 傳送失敗：{e}"
    finally:
        # _resolve_step_output_for_tg 對「資料夾」會 zip 到 system temp 下，這邊清掉
        try:
            if file_path.startswith(_tf.gettempdir()) and file_path.endswith(".zip"):
                os.unlink(file_path)
        except Exception:
            pass


def _find_prev_output_file(run, config) -> Optional[str]:
    """找上一個非 human_confirm 步驟的輸出檔案。人工確認節點「附檔案預覽」用。
    策略：
      1. 往前找第一個 step.output.path 有設、且檔案存在 → 回傳它
      2. 若都沒設 output.path（或設了但檔案不存在）→ 退到預設資料夾 ai_output/<workflow>/
         抓最近修改時間最新的檔案（排除目錄、截圖、log、.json 設定等雜訊）
      3. 都找不到 → None
    """
    try:
        from pathlib import Path as _P
        import time as _t
        from config import OUTPUT_BASE_PATH as _OUT_BASE

        # 策略 1:看 step.output.path(同 _resolve_path / _deterministic_validate 規則)
        idx = run.current_step - 1
        while idx >= 0:
            st = config.steps[idx]
            if st.human_confirm:
                idx -= 1
                continue
            if st.output and st.output.path:
                p = _P(st.output.path).expanduser()
                if not p.is_absolute():
                    parts = p.parts
                    if parts and parts[0] == "ai_output":
                        # 相容舊 YAML 寫法:ai_output/xxx 視為相對 OUTPUT_BASE_PATH 的父
                        p = _OUT_BASE.parent / p
                    else:
                        p = _OUT_BASE / run.pipeline_name / p
                if p.exists() and p.is_file():
                    return str(p)
            idx -= 1

        # 策略 2:預設目錄最新檔
        wf_dir = _OUT_BASE / run.pipeline_name
        if not wf_dir.exists() or not wf_dir.is_dir():
            return None
        # 過濾規則：
        #   排除資料夾
        #   排除我們自己產的截圖（screenshot_*.png / _preview.png / _compressed.jpg）
        #   排除 log / 內部 JSON
        skip_prefixes = ("screenshot_",)
        skip_suffixes = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
        skip_exts = {".log"}
        candidates = []
        for f in wf_dir.iterdir():
            if not f.is_file():
                continue
            name = f.name
            if name.startswith(skip_prefixes):
                continue
            if any(name.endswith(suf) for suf in skip_suffixes):
                continue
            if f.suffix.lower() in skip_exts:
                continue
            # 也排除 pipeline_settings / recipes / runs 等已知內部檔（以防掃到 OUTPUT_BASE）
            if name in ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal"):
                continue
            candidates.append((f.stat().st_mtime, f))
        if not candidates:
            return None
        candidates.sort(key=lambda x: x[0], reverse=True)
        latest = candidates[0][1]
        age = _t.time() - candidates[0][0]
        # 太舊的（>7 天）可能只是殘留，警告一下但還是回傳
        if age > 7 * 86400:
            import logging as _lg
            _lg.getLogger("pipeline").info(
                f"[preview] 退回找預設目錄最新檔，但最新檔已 {age/86400:.1f} 天沒更新，可能是殘留：{latest}"
            )
        return str(latest)
    except Exception:
        pass
    return None


def take_screenshots(pipeline_name: str, step_name: str) -> list[str]:
    """逐螢幕截圖（1 螢幕 → 1 張、2 螢幕 → 2 張）。回傳檔案路徑 list。
    mss 的 sct.monitors[0] 是「所有螢幕拼成的虛擬桌面」、monitors[1..N] 是每台實體螢幕。
    Telegram 看起來更直覺（不會因為多螢幕被壓成超寬一張），所以直接逐螢幕抓。
    """
    import time as _t
    from pathlib import Path as _P
    logger = logging.getLogger("pipeline")
    results: list[str] = []
    try:
        import mss as _mss
        _PROJ_ROOT = _P(__file__).parent.parent.parent.absolute()
        ss_dir = _PROJ_ROOT / "ai_output" / pipeline_name
        ss_dir.mkdir(parents=True, exist_ok=True)
        ts = _t.strftime('%Y%m%d_%H%M%S')
        with _mss.mss() as sct:
            monitors = sct.monitors[1:] or sct.monitors  # 單螢幕系統 monitors[1:] 可能空，退回全部
            for idx, mon in enumerate(monitors, start=1):
                tag = f"mon{idx}" if len(monitors) > 1 else "full"
                ss_path = ss_dir / f"screenshot_{step_name}_{ts}_{tag}.png"
                try:
                    img = sct.grab(mon)
                    from mss.tools import to_png as _to_png
                    _to_png(img.rgb, img.size, output=str(ss_path))
                except Exception as e:
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖失敗（略過）：{e}")
                    continue
                # 確認檔案真的產生且不是 0 bytes（to_png 偶爾會沉默失敗）
                if not ss_path.exists():
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖檔未產生：{ss_path}")
                    continue
                fsize = ss_path.stat().st_size
                if fsize <= 0:
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖 0 bytes，刪除並略過：{ss_path}")
                    try:
                        ss_path.unlink()
                    except Exception:
                        pass
                    continue
                logger.info(f"[{step_name}]   ✓ 螢幕 {idx} 截圖 {fsize/1024:.0f} KB → {ss_path.name}")
                results.append(str(ss_path))
        if results:
            logger.info(f"[{step_name}] 📸 截圖 {len(results)}/{len(monitors)} 張已儲存")
        else:
            logger.warning(f"[{step_name}] 截圖失敗：沒有檔案產生")
    except Exception as e:
        logger.warning(f"[{step_name}] 截圖失敗：{e}")
    return results


def take_screenshot(pipeline_name: str, step_name: str) -> Optional[str]:
    """舊接口保留：回傳第一張截圖（向後相容，如有其他呼叫者）"""
    paths = take_screenshots(pipeline_name, step_name)
    return paths[0] if paths else None


async def _notify_failure(run: PipelineRun, val: ValidationResult, step_name: str):
    """詢問用戶如何處理失敗步驟"""
    step_num = run.current_step + 1
    total = len(PipelineConfig.from_dict(run.config_dict).steps)

    # ── 缺套件特化通知 ──
    # awaiting_type=missing_dependency 時用專屬訊息 + 安裝確認按鈕（不走 generic 失敗）
    if run.awaiting_type == "missing_dependency":
        import json as _json
        try:
            meta = _json.loads(run.awaiting_suggestion or "{}")
        except Exception:
            meta = {}
        pkgs = meta.get("packages") or []
        stderr_tail = (meta.get("stderr_tail") or "")[-200:]

        # 比對 skill_packages.txt — 在清單內代表「應該裝、但 venv 損壞 / import 失敗」
        try:
            import sys as _sys
            from pathlib import Path as _PI
            _backend = str(_PI(__file__).parent.parent.absolute())
            if _backend not in _sys.path:
                _sys.path.insert(0, _backend)
            from skill_pkg_manager import _read_packages, _read_sandbox_packages
            from settings import get_settings as _gs
            sandbox_mode = _gs().get("skill_sandbox_mode", "host") == "wsl_docker"
            existing = set(_read_sandbox_packages() if sandbox_mode else _read_packages())
        except Exception:
            existing = set()

        in_list = [p for p in pkgs if p in existing]
        not_in_list = [p for p in pkgs if p not in existing]

        text = (
            f"📦 <b>需要安裝套件</b>\n\n"
            f"📋 {run.pipeline_name}\n"
            f"📍 步驟 {step_num}/{total}：<b>{step_name}</b>\n\n"
            f"程式碼用到的套件還沒安裝：\n"
        )
        for p in pkgs:
            tag = "（清單內、可能 venv 損壞）" if p in existing else "（不在清單）"
            text += f"  • <code>{p}</code> {tag}\n"
        if stderr_tail:
            # html.escape 一下避免 stderr 含 < > 解析錯
            import html as _html
            text += f"\n<i>stderr：{_html.escape(stderr_tail)}</i>\n"
        text += "\n按下方按鈕授權安裝、或改任務避開這個套件。"

        await _tg_send(run.telegram_chat_id, text, _missing_dep_keyboard(run.run_id, pkgs))
        return

    # 一般失敗
    text = (
        f"⚠️ <b>Pipeline 需要決策</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"📍 步驟 {step_num}/{total}：<b>{step_name}</b>\n\n"
        f"🔴 {val.reason}\n"
    )
    if val.suggestion:
        text += f"💡 建議：{val.suggestion}\n"
    text += "\n請選擇處理方式："
    has_prev = run.current_step > 0
    await _tg_send(run.telegram_chat_id, text, _decision_keyboard(run.run_id, has_prev=has_prev))


async def _notify_final(run: PipelineRun, config: PipelineConfig):
    """發送 pipeline 最終結果摘要"""
    total = len(config.steps)
    ok_count = sum(1 for r in run.step_results if r.validation_status == "ok")

    status_map = {
        "completed": ("✅", "Pipeline 完成"),
        "aborted":   ("🛑", "Pipeline 已中止"),
    }
    emoji, title = status_map.get(run.status, ("❌", "Pipeline 失敗"))

    duration = ""
    if run.ended_at and run.started_at:
        try:
            secs = int((
                datetime.fromisoformat(run.ended_at) -
                datetime.fromisoformat(run.started_at)
            ).total_seconds())
            duration = f"⏱ 耗時：{secs // 60}m {secs % 60}s\n"
        except Exception:
            pass

    # Step 摘要
    step_lines = []
    for i, step in enumerate(config.steps):
        if i < len(run.step_results):
            r = run.step_results[i]
            icon = {"ok": "✅", "warning": "⚠️", "failed": "❌"}.get(r.validation_status, "❓")
            step_lines.append(f"  {icon} {step.name}")
        else:
            step_lines.append(f"  ⬜ {step.name}（未執行）")

    text = (
        f"{emoji} <b>{title}</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"🔢 {ok_count}/{total} 步驟成功\n"
        f"{duration}"
        f"\n<b>步驟概覽：</b>\n" + "\n".join(step_lines) +
        f"\n\n📁 <code>{run.log_path}</code>"
    )
    await _tg_send(run.telegram_chat_id, text)


# ── Deterministic validation (fast recipe mode) ──────────────────────────────

def _deterministic_validate(step, exec_result, logger, workflow_name: str = "") -> ValidationResult:
    """Recipe 快速模式：不叫 LLM，只做確定性檢查。
    workflow_name：用來解析相對路徑（同 _resolve_path 規則 — 純檔名落到 workflow dir、
    `ai_output/...` 開頭視為相對於專案根）。沒傳就走 backend cwd（向後相容、但會踩坑）。
    """
    from pathlib import Path as _Path

    # 1. exit code
    if exec_result.exit_code != 0:
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exec_result.exit_code}",
            suggestion="Recipe 執行失敗，建議改用完整模式重跑",
        )

    # 2. 輸出檔存在 + 大小(路徑用跟 run_pipeline _resolve_path 一致的規則)
    if step.output and step.output.path:
        p = _Path(step.output.path).expanduser()
        if not p.is_absolute():
            from config import OUTPUT_BASE_PATH as _OUT_BASE
            parts = p.parts
            if parts and parts[0] == "ai_output":
                p = _OUT_BASE.parent / p
            elif workflow_name:
                p = _OUT_BASE / workflow_name / p
        if not p.exists():
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 不存在",
                suggestion="Recipe 未產生預期檔案，建議改用完整模式",
            )
        size = p.stat().st_size
        if size == 0:
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 為空檔案（0 bytes）",
                suggestion="Recipe 產生了空檔案，建議改用完整模式",
            )
        # CSV: 檢查有 header
        if p.suffix.lower() == ".csv":
            try:
                with open(p, "r", encoding="utf-8") as f:
                    lines = sum(1 for _ in f)
                if lines < 2:
                    return ValidationResult(
                        status="failed",
                        reason=f"CSV 檔案只有 {lines} 行（預期至少有 header + 資料）",
                        suggestion="",
                    )
            except Exception:
                pass
        # Excel: 檢查有 sheet
        if p.suffix.lower() in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(p, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                if sheet_count == 0:
                    return ValidationResult(
                        status="failed",
                        reason="Excel 檔案沒有任何工作表",
                        suggestion="",
                    )
            except Exception:
                pass

    logger.info(f"[{step.name}] ⚡ 確定性檢查通過")
    return ValidationResult(
        status="ok",
        reason="exit code=0、輸出檔案存在且非空",
        suggestion="",
    )


def _resolve_legacy_log_path(stored_path: str):
    """支援 #142 OUTPUT_BASE_PATH 統一前的舊 log_path。

    舊版 .env 設 OUTPUT_BASE_PATH=./ai_output 的 user、DB 內 log_path 仍是
    `D:\\...\\backend\\ai_output\\pipeline_logs\\xxx.log`。新版 ai_output 在 repo_root、
    .log 檔已被 auto-migrate 搬到新位置、但 DB 內 path 沒改 → 找不到。

    這層 fallback:原 path 不存在 → 嘗試把 backend/ai_output 換成 ai_output 再找一次。
    """
    from pathlib import Path as _P
    if not stored_path:
        return None
    p = _P(stored_path)
    if p.exists():
        return p
    s = str(p)
    # 兩種斜線都試
    for old, new in (("backend\\ai_output", "ai_output"), ("backend/ai_output", "ai_output")):
        if old in s:
            alt = _P(s.replace(old, new))
            if alt.exists():
                return alt
    return None


def get_run_log_tail(run_id: str, lines: int = 30) -> str:
    """取得 pipeline 執行 log 的最後 N 行（供 Telegram 查看）"""
    store = get_store()
    run = store.load(run_id)
    if not run or not run.log_path:
        return "（找不到 log）"
    log_file = _resolve_legacy_log_path(run.log_path)
    if not log_file:
        return "（log 檔案不存在）"
    try:
        all_lines = log_file.read_text(encoding="utf-8").splitlines()
        tail = all_lines[-lines:] if len(all_lines) > lines else all_lines
        return "\n".join(tail)
    except Exception as e:
        return f"（讀取失敗：{e}）"


# ── Main pipeline engine ──────────────────────────────────────────────────────

async def run_pipeline(
    config_dict: dict,
    chat_id: int,
    run_id: Optional[str] = None,
    start_from_step: int = 0,
) -> str:
    """執行(或恢復)一個 pipeline 的對外入口。

    包薄殼 wrapper、處理「computer_use workflow 啟動時自動縮小前景視窗」
    的 setup/teardown(setting 開啟才生效),實際邏輯在 _run_pipeline_inner。

    Wrapper 用 reference counting 處理並發:多個 workflow 同時跑時、
    第一個 minimize、最後一個 restore;中間呼叫只是 +1/-1 ref count。
    """
    from . import window_helper
    try:
        from settings import get_settings
        _s = get_settings()
        _auto_min = bool(_s.get("auto_minimize_for_computer_use", False))
    except Exception:
        _auto_min = False
    _do_minimize = _auto_min and window_helper.config_has_computer_use(config_dict)
    if _do_minimize:
        window_helper.request_minimize()
    try:
        return await _run_pipeline_inner(
            config_dict=config_dict,
            chat_id=chat_id,
            run_id=run_id,
            start_from_step=start_from_step,
        )
    finally:
        if _do_minimize:
            window_helper.request_restore()


async def _run_pipeline_inner(
    config_dict: dict,
    chat_id: int,
    run_id: Optional[str] = None,
    start_from_step: int = 0,
) -> str:
    """執行(或恢復)一個 pipeline 的實作本體。對外請呼叫 run_pipeline。"""
    store = get_store()

    # 建立或恢復 run
    if run_id:
        run = store.load(run_id)
        if not run:
            raise ValueError(f"找不到 pipeline run: {run_id}")
        
        # 確保 run 物件同步使用傳入的最新配置（包含 hint）
        run.config_dict = config_dict
        run.status = "running"
        run.current_step = start_from_step
        # 附加到原始 log 檔（不建新檔），前端讀到的 log_path 保持不變
        logger = resume_run_logger(run.run_id, run.log_path)
        logger.info(f"恢復執行，從步驟 {start_from_step + 1} 繼續")
    else:
        # 新建 run
        config = PipelineConfig.from_dict(config_dict)
        run_id = str(uuid.uuid4())[:12]
        logger, log_path = create_run_logger(run_id, config.name)
        run = PipelineRun(
            run_id=run_id,
            pipeline_name=config.name,
            config_dict=config_dict,
            telegram_chat_id=chat_id,
            log_path=log_path,
        )
        logger.info(f"Pipeline 開始：{config.name}，共 {len(config.steps)} 步驟")

    config = PipelineConfig.from_dict(run.config_dict)
    use_recipe = run.config_dict.get("_use_recipe", False)
    workflow_id = run.config_dict.get("_workflow_id") or run.workflow_id
    # 輸出目錄 / 日誌一律用「使用者在側邊欄看到的工作流名稱」(DB),不要用 AI 在 YAML 自取的
    # name → 否則檔案落到 AI 自取的資料夾(如 demo_sales_analysis)、使用者依工作流名找不到檔。
    # config.name 只用於輸出資料夾 / 日誌 / pipeline_id fallback,不影響步驟變數參照(那用 step name),
    # 所以單點覆寫即可全覆蓋 _workflow_output_dir / default_wd / _resolve_path。
    if workflow_id:
        try:
            from db import get_workflow as _get_wf
            _wf_name = (((_get_wf(workflow_id) or {}).get("name")) or "").strip()
            if _wf_name and _wf_name != config.name:
                logger.info(f"輸出目錄改用工作流名稱「{_wf_name}」(原 YAML name「{config.name}」)")
                config.name = _wf_name
                run.pipeline_name = _wf_name
        except Exception as _e:
            logger.warning(f"查工作流名稱失敗、沿用 YAML name「{config.name}」:{_e}")
    store.save(run)

    # ── Step loop ────────────────────────────────────────────
    completed_outputs: list[dict] = []  # 收集前步驟的輸出資訊
    # ── 控制流 dispatcher(Ticket 2)──────────────────────────
    # condition 節點可跳到任意 step name → 用 name_to_index 索引
    # visit_count 防無限迴圈(同一 step 訪問超過 MAX_VISITS = 1000 → abort)
    name_to_index: dict[str, int] = {s.name: i for i, s in enumerate(config.steps)}
    visit_count: dict[int, int] = {}
    MAX_VISITS_PER_STEP = 1000

    # 恢復執行時，重建已完成步驟的輸出資訊（供後續步驟參考）
    if start_from_step > 0:
        from pathlib import Path as _Path
        # 預先用 step_index 索引 step_results、給沒設 output.path 的步驟回填 actual_output_path
        _sr_idx = {sr.step_index: sr for sr in run.step_results}
        for i in range(start_from_step):
            prev_step = config.steps[i] if i < len(config.steps) else None
            if not prev_step or prev_step.human_confirm:
                continue
            # 優先：明確 output.path > 該步 StepResult.actual_output_path
            _eff_path = ""
            if prev_step.output and prev_step.output.path:
                _eff_path = prev_step.output.path
            else:
                _sr = _sr_idx.get(i)
                if _sr and getattr(_sr, "actual_output_path", ""):
                    _eff_path = _sr.actual_output_path
            if _eff_path:
                p = _Path(_eff_path).expanduser()
                out_info = {"path": str(p), "schema": ""}
                try:
                    if p.suffix == ".csv" and p.exists():
                        with open(p, "r") as f:
                            out_info["schema"] = f.readline().strip()
                    elif p.suffix in (".xlsx", ".xls") and p.exists():
                        out_info["schema"] = "Excel 工作簿"
                    elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                        out_info["schema"] = "圖片檔案"
                except Exception:
                    pass
                completed_outputs.append(out_info)
        if completed_outputs:
            logger.info(f"已重建 {len(completed_outputs)} 個前步驟的輸出資訊：{[o['path'] for o in completed_outputs]}")

    no_save_recipe = run.config_dict.get("_no_save_recipe", False)
    silent_recipe = run.config_dict.get("_silent_recipe", False)

    if silent_recipe:
        # 無人值守模式（TG / 排程觸發）:
        # - 有 recipe → 跳過、不覆寫（保護桌面用戶調好的版本）
        # - 無 recipe → 直接建立（首次跑可以 seed）
        # - 永不延遲累積 pending_recipes（桌面不會卡 dialog）
        logger.info("無人值守模式（silent_recipe）:有 recipe 跳過不覆寫、無 recipe 才建立、不彈 dialog")
    else:
        # ── 自動偵測：工作流含 web_crawler 節點時，自動關閉 recipe 儲存 ──
        # 理由：爬蟲輸出每次不同（網頁內容 / 抓取時間都會變），下游 skill 步驟
        # 的 input_fingerprint 永遠不會跟既有 recipe 相符 → 命中率 0、純占空間。
        # 還會在 UI 顯示「已有 Recipe 快取」假訊號誤導使用者。
        # 使用者明確要求「跑就跑、不管命不命中」的話可手動 force `_use_recipe=true`。
        if not no_save_recipe and any(getattr(s, "web_crawler", False) for s in config.steps):
            no_save_recipe = True
            logger.info("偵測到 web_crawler 節點 → 自動關閉 recipe 儲存"
                        "（爬蟲輸出每次不同、recipe 永遠 miss、存了也不會命中）")

    # ── 跨 step 沿用 working_dir(避免「step 1 設了自訂子資料夾、step 2 沒設 → 兩 step 落不同 dir」) ──
    # 規則:後續 step 沒明確 working_dir / output.path 時、預設沿用前一步算出來的 wd
    # 不沿用 default_wd(workflow name 那條)、保證輸出統一在同一個資料夾
    _prev_step_wd: Optional[str] = None

    while run.current_step < len(config.steps):
        # ── 每步開始前檢查中止旗標 ──
        if is_abort_requested(run.run_id):
            clear_abort(run.run_id)
            unregister_task(run.run_id)
            run.status = "aborted"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info("用戶透過 UI 中止 Pipeline")
            await _notify_final(run, config)
            return run.run_id

        step = config.steps[run.current_step]
        step_num = run.current_step + 1
        total = len(config.steps)
        logger.info(f"══ 步驟 {step_num}/{total}：{step.name} ══")

        # Trace / token tracking — 每步重置、subagent / skill 分支會填入。
        # 其他 step type（script / human_confirm / web_crawler / 等）保持空 dict / 空 list、
        # 表示這步沒有 LLM 多輪推理可記。step_result 結尾再寫進 StepResult.token_usage / tool_calls。
        step_token_usage: dict = {}
        step_tool_calls: list = []
        # 該步驟匯出的變數(UIA/Computer Use 的 save_as 累積)。執行完寫進 StepResult.step_vars
        # 供下游 step 用 `{{ steps.<name>.output.<key> }}` 引用
        step_step_vars: dict = {}
        step_started_at = datetime.now().isoformat()

        # ── 變數 / 表達式系統:render 該步所有 {{ }} 欄位 ──
        # 沒寫 {{ }} 的欄位完全不動(舊 workflow 行為不變)
        # render 用的 context = run.step_results(已完成) + run.input_params + os.environ
        try:
            from .expression import (
                build_context as _build_var_context,
                render_step as _render_var_step,
                ExpressionError as _VarExpressionError,
            )
            _var_ctx = _build_var_context(
                step_results=run.step_results,
                input_params=getattr(run, "input_params", None) or {},
            )
            _render_var_step(step, _var_ctx)
        except _VarExpressionError as _var_exc:
            logger.error(f"[{step.name}] 變數展開失敗:{_var_exc}")
            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=1,
                stdout_tail="",
                stderr_tail=str(_var_exc),
                validation_status="failed",
                validation_reason=f"變數展開失敗:{_var_exc}",
                validation_suggestion="檢查 {{ }} 內變數是否拼錯、或啟動 workflow 時是否漏帶 input_params",
                retries_used=0,
                started_at=step_started_at,
                ended_at=datetime.now().isoformat(),
            )
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            run.status = "failed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            await _notify_final(run, config)
            unregister_task(run.run_id)
            return run.run_id

        # ── 控制流節點:condition(Ticket 2)─────────────────────
        # 不執行任何命令、純求值 + 跳轉。寫個 step_result(stdout 記錄決策)、
        # 跳到目標 step name 對應的 index、continue outer loop(不走後面 retry 邏輯)
        if step.condition:
            from .expression import (
                eval_condition as _eval_cond,
                eval_value as _eval_val,
                ExpressionError as _CondError,
            )
            # 訪問次數上限保護
            visit_count[run.current_step] = visit_count.get(run.current_step, 0) + 1
            if visit_count[run.current_step] > MAX_VISITS_PER_STEP:
                logger.error(f"[{step.name}] condition 訪問超過 {MAX_VISITS_PER_STEP} 次、判定無限迴圈、中止")
                step_result = StepResult(
                    step_index=run.current_step, step_name=step.name,
                    exit_code=1, stdout_tail="",
                    stderr_tail=f"訪問超過 {MAX_VISITS_PER_STEP} 次、可能無限迴圈",
                    validation_status="failed",
                    validation_reason=f"condition 節點被訪問 > {MAX_VISITS_PER_STEP} 次",
                    validation_suggestion="檢查 on_true/on_false 是否導致循環",
                    retries_used=0,
                    started_at=step_started_at,
                    ended_at=datetime.now().isoformat(),
                )
                if len(run.step_results) > run.current_step:
                    run.step_results[run.current_step] = step_result
                else:
                    run.step_results.append(step_result)
                run.status = "failed"
                run.ended_at = datetime.now().isoformat()
                store.save(run)
                await _notify_final(run, config)
                unregister_task(run.run_id)
                return run.run_id

            decision_msg = ""
            target_name = ""
            try:
                if step.switch:
                    # Switch 模式
                    val = _eval_val(step.switch, _var_ctx)
                    target_name = step.cases.get(val, "") or step.default
                    decision_msg = f"switch 求值={val!r} → 跳到 {target_name or '(end)'}"
                elif step.expression:
                    # IF 模式
                    cond = _eval_cond(step.expression, _var_ctx)
                    target_name = step.on_true if cond else step.on_false
                    decision_msg = f"IF 求值={cond} → 跳到 {target_name or '(end)'}"
                else:
                    raise _CondError("condition 節點需填 expression(IF)或 switch(Switch)")
            except _CondError as _ce:
                logger.error(f"[{step.name}] condition 求值失敗:{_ce}")
                _cond_sugg = (
                    "檢查 expression / switch 表達式語法、引用變數是否存在。"
                    "Jinja2 判斷包含用 \"'關鍵字' in 變數\"、不是 .contains();"
                    "字串相等用 ==;list/dict 取值用 []。"
                )
                step_result = StepResult(
                    step_index=run.current_step, step_name=step.name,
                    exit_code=1, stdout_tail="",
                    stderr_tail=str(_ce),
                    validation_status="failed",
                    validation_reason=f"condition 求值失敗:{_ce}",
                    validation_suggestion=_cond_sugg,
                    retries_used=0,
                    started_at=step_started_at,
                    ended_at=datetime.now().isoformat(),
                )
                if len(run.step_results) > run.current_step:
                    run.step_results[run.current_step] = step_result
                else:
                    run.step_results.append(step_result)
                # 走 awaiting_human=failure(讓人工 / 自我修復能改表達式重跑),不直接判 failed —
                # condition 求值失敗多半是表達式語法錯(如 .contains)、是 self_heal 改 YAML 能修的
                run.status = "awaiting_human"
                run.awaiting_type = "failure"
                run.awaiting_message = f"condition 求值失敗:{_ce}"
                run.awaiting_suggestion = _cond_sugg
                store.save(run)
                await _notify_failure(
                    run,
                    ValidationResult(status="failed",
                                     reason=f"condition 求值失敗:{_ce}",
                                     suggestion=_cond_sugg),
                    step.name,
                )
                unregister_task(run.run_id)
                return run.run_id

            logger.info(f"[{step.name}] 🔀 {decision_msg}")
            # 寫 step_result 記錄決策
            step_result = StepResult(
                step_index=run.current_step, step_name=step.name,
                exit_code=0, stdout_tail=decision_msg, stderr_tail="",
                validation_status="ok", validation_reason="condition 求值成功",
                validation_suggestion="", retries_used=0,
                started_at=step_started_at,
                ended_at=datetime.now().isoformat(),
            )
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)

            # 決定下一步:目標名稱 → index;留空或找不到 → 流程結束
            if not target_name:
                logger.info(f"[{step.name}] 跳轉目標為空、結束流程")
                run.current_step = len(config.steps)
            elif target_name not in name_to_index:
                logger.error(f"[{step.name}] 跳轉目標 step '{target_name}' 不存在")
                step_result.validation_status = "failed"
                step_result.validation_reason = f"跳轉目標 '{target_name}' 不存在於 workflow"
                step_result.stderr_tail = step_result.validation_reason
                run.status = "failed"
                run.ended_at = datetime.now().isoformat()
                store.save(run)
                await _notify_final(run, config)
                unregister_task(run.run_id)
                return run.run_id
            else:
                run.current_step = name_to_index[target_name]
            store.save(run)
            continue  # 跳過後面所有 step 執行邏輯、回 while 頂端

        # 步驟開始前 snapshot workflow 輸出資料夾（mtime 比對用）
        # 步驟結束後 _diff_snapshot_pick_main 找新增/修改的檔，存進 StepResult.actual_output_path
        # → TG「取任一步輸出」就能對應到該步真正寫的檔（不再讓多個 skill 步驟搶到「最新檔」）
        _step_dir_snapshot_before = _snapshot_workflow_dir(config.name)

        # ── 人工確認節點：暫停等待確認 ──
        if step.human_confirm:
            logger.info(f"[{step.name}] ✋ 人工確認節點，暫停等待確認")

            # 收集前一步結果摘要
            prev_summary = ""
            if run.step_results:
                prev = run.step_results[-1]
                status_icon = {"ok": "✅", "failed": "❌"}.get(prev.validation_status, "⚠️")
                prev_summary = (
                    f"前一步驟：{prev.step_name}\n"
                    f"狀態：{status_icon} {prev.validation_status}\n"
                    f"原因：{prev.validation_reason or '（無）'}\n"
                )
                if prev.stdout_tail:
                    prev_summary += f"輸出摘要：{prev.stdout_tail[-300:]}\n"

            confirm_msg = step.message or "請確認上一步結果是否正確，再繼續執行"
            full_message = f"{prev_summary}\n📋 {confirm_msg}"

            run.status = "awaiting_human"
            run.awaiting_type = "human_confirm"
            run.awaiting_message = confirm_msg

            # step_result 跟 status 一起一次寫完，後面 await TG / preview 期間就不再 save。
            # 否則用戶按通過時 resume_pipeline 把狀態改成 running，本協程的 stale run 物件
            # 在後面 store.save 又把狀態蓋回 awaiting_human → 用戶第二次按通過會再過 gate
            # → 同一步驟被啟動兩次（race condition，造成工作流跑慢／重複執行）。
            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=0,
                stdout_tail="等待人工確認",
                stderr_tail="",
                validation_status="ok",
                validation_reason="人工確認節點 — 等待中",
                validation_suggestion="",
                retries_used=0,
            )
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            # 判斷「補充指示」按鈕要不要給：只有上一個可執行節點是 skill_mode 才顯示
            # （往回跳過其他連續的 human_confirm 節點，找真正要被重做的 step）
            _prev = run.current_step - 1
            while _prev >= 0 and config.steps[_prev].human_confirm:
                _prev -= 1
            allow_hint = _prev >= 0 and bool(config.steps[_prev].skill_mode)

            # Telegram 通知
            if step.notify_telegram:
                tg_text = (
                    f"✋ <b>Pipeline 等待確認</b>\n\n"
                    f"📋 {run.pipeline_name}\n"
                    f"📍 步驟 {step_num}/{total}：<b>{step.name}</b>\n\n"
                )
                if prev_summary:
                    tg_text += f"{prev_summary}\n"
                tg_text += f"💬 {confirm_msg}\n\n請選擇："
                await _tg_send(run.telegram_chat_id, tg_text,
                               _confirm_keyboard(run.run_id, screenshot=step.screenshot,
                                                 allow_hint=allow_hint,
                                                 preview_enabled=step.preview_prev_output))
                # 自動傳上一步輸出檔：step.send_prev_output=True 時，立刻把上一步 output.path
                # 的檔案（或資料夾打包成 zip）當 document 傳到 TG。手機上可直接點開 / 下載。
                # 失敗（沒設輸出 / 檔案不存在 / 太大）也只 log warning、不擋人工確認流程。
                if step.send_prev_output:
                    try:
                        # 找上一個非 human_confirm 步驟（連續多個 human_confirm 時往前跳過）
                        _po_prev = run.current_step - 1
                        while _po_prev >= 0 and config.steps[_po_prev].human_confirm:
                            _po_prev -= 1
                        if _po_prev >= 0:
                            _po_step = config.steps[_po_prev]
                            _po_result = next((sr for sr in run.step_results if sr.step_index == _po_prev), None)
                            ok, msg = await _send_step_output_to_tg(
                                run.telegram_chat_id, _po_step,
                                step_label=f"步驟 {_po_prev+1}：{_po_step.name}",
                                workflow_name=config.name,
                                logger=logger,
                                step_result=_po_result,
                            )
                            if ok:
                                logger.info(f"[{step.name}] ✓ 自動傳上一步輸出：{msg}")
                            else:
                                # 不再廣播警告到 TG（noise）：按鈕「📎 上一步輸出」永遠存在，
                                # 使用者要時自己點即可。失敗只 log 到 backend、debug 用。
                                logger.warning(
                                    f"[{step.name}] 自動傳上一步輸出未成功（不廣播到 TG）：{msg}"
                                )
                    except Exception as _e:
                        logger.warning(f"[{step.name}] send_prev_output 例外：{_e}")

                # 自動截圖：step.screenshot=True 時，發完決策訊息立刻截全螢幕附過去，
                # 逐螢幕送（雙螢幕 → 2 張，方便 TG 上直接看到上一步結果不用再按按鈕）
                if step.screenshot:
                    try:
                        ss_paths = take_screenshots(run.pipeline_name, step.name)
                        if ss_paths:
                            await _tg_send_photos(
                                run.telegram_chat_id,
                                ss_paths,
                                caption_prefix=f"📸 {step.name}",
                            )
                    except Exception as _e:
                        logger.warning(f"[{step.name}] 自動截圖傳送失敗：{_e}")

                # 檔案預覽：preview_prev_output=True 時，把上一步的 output.path 檔案
                # render 成 PNG 一併傳，讓使用者在手機上直接看到內容（不用 SSH 回電腦）
                # B1 路線：pandas/python-docx/python-pptx/pypdfium2/PIL 純 headless；
                # 後備：LibreOffice 無頭轉 PDF（需使用者自己裝 soffice）
                if step.preview_prev_output:
                    prev_file = _find_prev_output_file(run, config)
                    if not prev_file:
                        logger.info(f"[{step.name}] preview_prev_output 開啟但找不到上一步輸出檔，跳過")
                    else:
                        try:
                            # render 同步跑（pandas/PIL 不算慢；LibreOffice 若用到才可能 >5s）
                            # 放 executor 裡別 block event loop
                            from pipeline.file_preview import render_file_preview
                            preview_paths = await asyncio.get_event_loop().run_in_executor(
                                None,
                                lambda fp=prev_file: render_file_preview(fp, out_dir=str(Path(fp).parent)),
                            )
                            if preview_paths:
                                logger.info(f"[{step.name}] 📄 預覽產生 {len(preview_paths)} 張 → 傳 TG")
                                await _tg_send_photos(
                                    run.telegram_chat_id,
                                    preview_paths,
                                    caption_prefix=f"📄 上一步驟輸出預覽：{Path(prev_file).name}",
                                )
                            else:
                                logger.warning(f"[{step.name}] 預覽 render 回傳空清單：{prev_file}")
                        except Exception as _e:
                            logger.warning(f"[{step.name}] 檔案預覽失敗：{_e}")

            # ── hc_timeout:超過 step.timeout 秒沒人按、自動取設定的行動 ─────
            # 預設 hc_on_timeout='wait' = 永遠等(忽略 step.timeout、保留現有行為)
            # 'pass' / 'reject' / 'abort' 才會啟動 watcher
            _hc_to = int(getattr(step, "timeout", 0) or 0)
            _hc_act = (getattr(step, "hc_on_timeout", "wait") or "wait").lower()
            if _hc_to > 0 and _hc_act in ("pass", "reject", "abort"):
                async def _hc_timeout_watcher(rid: str, secs: int, action: str, step_name: str):
                    try:
                        await asyncio.sleep(secs)
                        # 檢查還在 awaiting_human、是同一個 step、才動作
                        _cur = store.load(rid)
                        if not _cur or _cur.status != "awaiting_human":
                            return  # 已被使用者處理或別的原因離開
                        if _cur.awaiting_type != "human_confirm":
                            return  # 是別種等待(e.g. missing_dependency),不要打架
                        logger.warning(f"[{step_name}] ⏰ hc_timeout 觸發({secs}s 沒回應)→ 自動取行動 '{action}'")
                        # resume_pipeline 接受 decision: "retry"/"skip"/"abort"/"continue"/"retry_with_hint"
                        if action == "pass":
                            await resume_pipeline(rid, decision="continue", hint="(自動超時通過)")
                        elif action == "reject":
                            await resume_pipeline(rid, decision="retry", hint="(自動超時駁回、上一步重做)")
                        elif action == "abort":
                            _cur.status = "aborted"
                            _cur.ended_at = datetime.now().isoformat()
                            store.save(_cur)
                            try:
                                await _notify_final(_cur, config)
                            except Exception:
                                pass
                    except Exception as _e:
                        logger.warning(f"[{step_name}] hc_timeout watcher 例外:{_e}")
                asyncio.create_task(_hc_timeout_watcher(run.run_id, _hc_to, _hc_act, step.name))
                logger.info(f"[{step.name}] ⏰ 已設 hc_timeout={_hc_to}s、超時動作={_hc_act}")

            # step_result 已於 await 前寫入；這裡只純粹釋放 task 並退出協程
            unregister_task(run.run_id)
            return run.run_id  # 暫停，等 resume_pipeline 被呼叫

        logger.debug(f"[{step.name}] batch 全文（{len(step.batch)} 字元）：{step.batch[:500]}")

        retries_used = 0
        step_failures: list[dict] = []  # 累積此步驟的失敗歷史，傳給下次重試

        # 計算當前步驟的工作目錄 (Working Directory)
        from pathlib import Path as _Path
        # 定義專案根目錄 (backend/pipeline/runner.py 的上三層)
        _PROJ_ROOT = _Path(__file__).parent.parent.parent.absolute()

        def _resolve_path(p: str) -> _Path:
            """把 output.path 解析成絕對路徑：
            - `~/xxx` 展開到使用者家目錄
            - 絕對路徑直接用
            - 相對路徑（一般情況）→ 以**本 workflow 的輸出資料夾**為基準
              （ai_output/<workflow_name>/）。所以 `path: posts_list.md` 會
              落到 ai_output/<workflow>/posts_list.md，不會跑到專案根目錄。
            - 相對路徑（已含 `ai_output/` 開頭）→ 以**專案根目錄**為基準。
              這條保留是因為 runner 內部 auto-default 路徑都是
              `ai_output/<workflow>/<step>_result.md` 的完整格式，必須以專案
              根當基準才解析得到正確位置。
            """
            pp = _Path(p).expanduser()
            if pp.is_absolute():
                return pp
            parts = pp.parts
            if parts and parts[0] == "ai_output":
                # 已經是「相對於專案根」的完整 workflow 路徑（auto-default 走這條）
                return _PROJ_ROOT / pp
            # 一般使用者寫的相對路徑 → 落到本 workflow 的輸出資料夾
            return _PROJ_ROOT / "ai_output" / config.name / pp

        # 預設：專案根目錄/ai_output/{pipeline_name}/
        default_wd = str(_PROJ_ROOT / "ai_output" / config.name)
        wd = step.working_dir
        if not wd and step.output and step.output.path:
            wd = str(_resolve_path(step.output.path).parent)
        if not wd and _prev_step_wd:
            # 沿用前一步的 working_dir(此 step 沒設 working_dir、也沒設 output.path)
            # 避免「step 1 自訂子資料夾、step 2 沒設」造成輸出散落
            wd = _prev_step_wd
            logger.info(f"[{step.name}] working_dir 沿用前一步:{wd}")
        if not wd:
            wd = default_wd
        _prev_step_wd = wd  # 給下一步沿用
        _Path(wd).mkdir(parents=True, exist_ok=True)

        # Retry loop for this step
        while True:
            if step.visual_validation:
                # ── 視覺驗證節點：純 VLM 判斷，不執行命令 ──
                from .visual_validator import run_visual_validation
                from .executor import ExecResult as _ExecResult
                prev_file = _find_prev_output_file(run, config) if step.vv_source != "current_screen" else None
                # search_region 解析：4 整數 [l,t,w,h]，否則 None
                _vsr = step.vv_search_region or []
                vv_region = None
                if isinstance(_vsr, (list, tuple)) and len(_vsr) == 4:
                    try:
                        vv_region = (int(_vsr[0]), int(_vsr[1]), int(_vsr[2]), int(_vsr[3]))
                        if vv_region[2] <= 0 or vv_region[3] <= 0:
                            vv_region = None
                    except (TypeError, ValueError):
                        vv_region = None
                vv_pass, vv_reason = await run_visual_validation(
                    source=step.vv_source,
                    prompt=step.vv_prompt,
                    prev_output_file=prev_file,
                    out_dir=wd,
                    search_region=vv_region,
                    logger=logger,
                    llm_role=getattr(step, "llm_role", "primary"),
                )
                exec_result = _ExecResult(
                    exit_code=0 if vv_pass else 1,
                    stdout=f"[visual_validation] {vv_reason}",
                    stderr="" if vv_pass else f"VLM 判斷未通過：{vv_reason}",
                )
            elif step.computer_use:
                # ── 桌面自動化節點：純 pyautogui + opencv，不走 LLM / recipe ──
                from .computer_use import execute_computer_use_step
                from .executor import ExecResult as _ExecResult
                # assets_dir 相對路徑解析：若為空，預設 ai_output/<pipeline>/<step_name>_assets
                if step.assets_dir:
                    assets_abs = str(_resolve_path(step.assets_dir))
                else:
                    assets_abs = str(_resolve_path(f"ai_output/{config.name}/{step.name}_assets"))
                # actions 是 ComputerUseAction pydantic model，轉成 dict list 傳進引擎
                # by_alias=True 確保 else_ 這類為了閃 Python 保留字取的別名，
                # 在轉 dict 時還原為 YAML 原生的 "else" key（讓 execute_action 用 .get("else") 讀得到）
                actions_dicts = [a.model_dump(by_alias=True) if hasattr(a, "model_dump") else dict(a) for a in (step.actions or [])]
                _cu_result = await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: execute_computer_use_step(
                        actions=actions_dicts,
                        assets_dir=assets_abs,
                        logger=logger,
                        run_id=run.run_id,
                        fail_fast=step.fail_fast,
                        cv_threshold=step.cv_threshold,
                        cv_search_only_near=step.cv_search_only_near,
                        cv_search_radius=step.cv_search_radius,
                        cv_trigger_hover=step.cv_trigger_hover,
                        cv_hover_wait_ms=step.cv_hover_wait_ms,
                        cv_coord_fallback=step.cv_coord_fallback,
                        ocr_threshold=step.ocr_threshold,
                        ocr_cv_fallback=step.ocr_cv_fallback,
                        cu_vlm_check_strategy=step.cu_vlm_check_strategy,
                        cu_on_mismatch=step.cu_on_mismatch,
                        cu_vlm_max_retries=step.cu_vlm_max_retries,
                        uia_window=step.uia_window,
                        llm_role=getattr(step, "llm_role", "primary"),
                    ),
                )
                # 映射回 ExecResult 讓後續驗證/重試邏輯通用
                exec_result = _ExecResult(
                    exit_code=_cu_result.exit_code,
                    stdout=_cu_result.stdout,
                    stderr=_cu_result.stderr,
                )
                # 把 UIA / Computer Use 透過 save_as 累積的變數帶到 step_result.step_vars,
                # 後續 step 可用 `{{ steps.<name>.output.<key> }}` 引用。
                if getattr(_cu_result, "step_variables", None):
                    step_step_vars.update(_cu_result.step_variables)
            elif step.outlook_automation:
                # Outlook 自動化節點：永遠跑 host（pywin32），跳過 sandbox / recipe 路徑
                if step.output and step.output.path:
                    _resolved_out = str(_resolve_path(step.output.path))
                else:
                    # 使用者沒設 outputPath → 自動 default 到 ai_output/<workflow>/<step>_result.md
                    # 這樣自由輸入需求（例如「整理今天的信件做摘要」）不會只 print 到 stdout
                    # 後消失，整理成果一定會落地成檔，使用者開資料夾就看得到。
                    import re as _re
                    _safe_step = _re.sub(r"[^\w一-鿿_-]", "_", step.name).strip("_") or "result"
                    _resolved_out = str(_resolve_path(f"ai_output/{config.name}/{_safe_step}_result.md"))
                    logger.info(f"[{step.name}] 自動 default outputPath → {_resolved_out}")
                exec_result = await execute_step_with_outlook(
                    template=step.outlook_template,
                    template_params=step.outlook_params or {},
                    free_text=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    output_path=_resolved_out,
                    working_dir=wd,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    run_id=run.run_id,
                    ask_mode=step.ask_mode,
                    llm_role=getattr(step, "llm_role", "primary"),
                )
            elif step.web_crawler:
                # 網頁爬蟲節點：永遠跑沙盒（Crawl4AI 在 pipeline-sandbox-v5 容器內）；
                # CF fallback 走 host 端 FlareSolverr。不進 LLM、不進 recipe。
                if step.output and step.output.path:
                    _resolved_out = str(_resolve_path(step.output.path))
                else:
                    # 沒設 outputPath → default 到 ai_output/<workflow>/<step>_result.md
                    import re as _re
                    _safe_step = _re.sub(r"[^\w一-鿿_-]", "_", step.name).strip("_") or "result"
                    _resolved_out = str(_resolve_path(f"ai_output/{config.name}/{_safe_step}_result.md"))
                    logger.info(f"[{step.name}] 自動 default outputPath → {_resolved_out}")
                # interactions 是 list[dict]，pydantic 帶 default_factory 會給空 list
                exec_result = await execute_step_with_web_crawler(
                    mode=step.wc_mode or "web",
                    url=step.wc_url,
                    urls=step.wc_urls or [],
                    js_render=step.wc_js_render,
                    wait_for_selector=step.wc_wait_for_selector,
                    cloudflare_fallback=step.wc_cloudflare_fallback,
                    cookies=step.wc_cookies,
                    interactions=step.wc_interactions or [],
                    download_assets=step.wc_download_assets,
                    scroll_count=getattr(step, "wc_scroll_count", 0) or 0,
                    target_post_count=getattr(step, "wc_target_post_count", 0) or 0,
                    with_children=getattr(step, "wc_with_children", False),
                    child_link_pattern=getattr(step, "wc_child_link_pattern", "") or "",
                    max_children=getattr(step, "wc_max_children", 10) or 10,
                    video_url=step.wc_video_url,
                    video_quality=step.wc_video_quality,
                    video_max_filesize_mb=step.wc_video_max_filesize_mb,
                    video_max_duration_min=step.wc_video_max_duration_min,
                    video_subs=step.wc_video_subs,
                    video_subs_langs=step.wc_video_subs_langs,
                    video_save_info_json=step.wc_video_save_info_json,
                    output_path=_resolved_out,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                )
            elif step.subagent:
                # Subagent 節點：多輪 LLM agent loop、role-based 系統提示、tool 白名單過濾
                # 跳過 recipe cache（多輪結果非確定性）、跳過 validator（loop 內已自我驗證）
                from .subagent_runner import run_subagent
                from .executor import ExecResult as _ExecResult
                if step.output and step.output.path:
                    _resolved_out = str(_resolve_path(step.output.path))
                else:
                    # 沒設 output.path → 自動 derive 到 workflow_dir 下、給 LLM 明確存哪
                    _resolved_out, _ext = _derive_default_output_path(step, wd)
                    logger.info(f"[{step.name}] step.output.path 未設,自動 derive → {_resolved_out} (副檔名 {_ext} 由 batch 關鍵字推測)")
                sub_result = await run_subagent(
                    role_name=step.subagent_role or "data_analyst",
                    task=step.batch,
                    max_iter=step.subagent_max_iter or 5,
                    workflow_dir=wd,
                    run_id=run.run_id,
                    step_name=step.name,
                    output_path=_resolved_out,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    timeout=step.timeout,
                    step_logger=logger,
                    llm_role=getattr(step, "llm_role", "primary"),
                )
                _tools_used = [tc.get("name", "?") for tc in sub_result.tool_calls_made]
                # 把 subagent 的 token usage / tool 呼叫時間軸記到該 step、StepResult 結尾寫入
                step_token_usage = sub_result.token_usage or {}
                step_tool_calls = list(sub_result.tool_calls_made or [])

                # ⛔ Hallucination 防護:subagent done(success=true) 但 output 檔不存在 → 強制 failed
                #
                # 真實案例:LLM 在 boss role 寫了 4000-6000 字計畫 reply、直接呼 done(success=true)、
                # 從沒呼 run_python 寫檔。runner 看 sub_result.success=True 標 ✅ 通過、
                # 但下個 step 找不到 final_report.md 才爆。修法:這層強制 check 檔在不在、
                # 不在就 override 成 failed、訊息列「LLM 宣稱完成但 output 不存在」。
                _hallucinated = False
                if sub_result.success and _resolved_out:
                    if not Path(_resolved_out).exists():
                        _hallucinated = True
                        logger.error(
                            f"[{step.name}] ⛔ Subagent 宣稱 done(success=true) 但 output 檔不存在:{_resolved_out}\n"
                            f"   LLM 可能寫了 reply 但跳過 run_python 寫檔的步驟、或寫到別處去了。"
                        )
                        sub_result.success = False
                        sub_result.error = (
                            f"hallucinated_done:LLM 主動 done(success=true) 但 output 檔 {_resolved_out} 不存在。"
                            f"請改進 role 系統提示要求 done 前先 Path(output).exists() 自驗。"
                        )

                exec_result = _ExecResult(
                    exit_code=0 if sub_result.success else 1,
                    stdout=(
                        f"[subagent/{step.subagent_role}] {sub_result.final_message}\n"
                        f"\n(iters={sub_result.iterations}, tools={_tools_used})"
                        + (f"\n⚠ 偵測到 hallucinated done、已 override 為失敗" if _hallucinated else "")
                    ),
                    stderr="" if sub_result.success else (sub_result.error or "subagent 執行失敗"),
                )
            elif step.skill_mode:
                # recipe key 使用「索引:名稱」避免同名步驟互相覆蓋
                recipe_step_key = f"{step_num}:{step.name}"
                # 把 output_path 解析成絕對路徑傳給 LLM、避免 LLM 搞不清楚相對於哪個 cwd
                if step.output and step.output.path:
                    _resolved_out = str(_resolve_path(step.output.path))
                else:
                    # 沒設 output.path → 自動 derive 到 workflow_dir 下、給 LLM 明確存哪
                    # 對應「不指定輸出路徑 = 落在工作流資料夾」的核心設計精神
                    _resolved_out, _ext = _derive_default_output_path(step, wd)
                    logger.info(f"[{step.name}] step.output.path 未設、自動 derive → {_resolved_out} (副檔名 {_ext} 由 batch 關鍵字推測)")
                # 判斷此 step 之後有沒有外部 AI validator 會跑、用於 skill loop 內 output-driven hint
                # validator 跑的條件(對齊 validator 分派邏輯 line 1974/1993):
                #   pipeline.validate=True AND (有 expect 描述 OR 此為 skill 節點)
                # skill 節點即使沒 expect、也會跑淺 LLM 驗證防 silent fail
                _has_validator = (
                    config.validate
                    and step.output is not None
                    and (bool(step.output.get_expect()) or step.skill_mode)
                )
                exec_result = await execute_step_with_skill(
                    task_description=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    output_path=_resolved_out,
                    working_dir=wd,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    pipeline_id=workflow_id or config.name,
                    use_recipe=use_recipe,
                    no_save_recipe=no_save_recipe,
                    readonly=step.readonly,
                    run_id=run.run_id,
                    previous_failures=step_failures if step_failures else None,
                    recipe_step_key=recipe_step_key,
                    skill_name=step.skill,
                    ask_mode=step.ask_mode,
                    silent_recipe=silent_recipe,
                    llm_role=getattr(step, "llm_role", "primary"),
                    has_external_validator=_has_validator,
                )
            else:
                exec_result = await execute_step(
                    command=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    run_id=run.run_id,
                    working_dir=wd,
                    background=getattr(step, "background", False),
                    ready_after_seconds=getattr(step, "ready_after_seconds", 0),
                )

            # 快速模式：Recipe 命中 + 執行成功 → 確定性驗證（不叫 LLM）
            recipe_hit = (exec_result.stderr == "__RECIPE_HIT__")
            if recipe_hit:
                exec_result.stderr = ""  # 清掉標記

            # Trace：把 ExecResult 帶回的 token_usage / tool_calls 接到該 step（skill / outlook
            # / 一般 script 都走這條；subagent 分支已自己 set 過、這裡不會覆寫到、因為
            # ExecResult.token_usage 預設 empty dict、覆寫只會把空值塞回去 — 用 truthy 判斷保護）
            _tu = getattr(exec_result, 'token_usage', None) or {}
            if _tu and _tu.get('total_tokens', 0) > 0:
                step_token_usage = _tu
            _tc = getattr(exec_result, 'tool_calls', None) or []
            if _tc:
                step_tool_calls = list(_tc)

            has_expect = step.output and step.output.get_expect()

            # ── 預先算 snapshot diff(在 validate 之前)─────────────────
            # 目的:就算使用者沒設 output.path,系統也能告訴 validator「這個 step 實際寫了什麼新檔」
            # → validator 拿到的 effective_output_path 才能做 mtime 檢查、擋住 LLM 用舊檔幻覺成功
            # 不算 multi-URL web_crawler(輸出是資料夾、單檔 diff 會挑錯)
            _eff_output_path: Optional[str] = None
            try:
                if step.output and step.output.path:
                    _p = _resolve_path(step.output.path)
                    if _p.exists() and _p.is_file():
                        _eff_output_path = str(_p.absolute())
                if not _eff_output_path:
                    _is_multi_wc = bool(getattr(step, "web_crawler", False)) and len(
                        [u for u in (getattr(step, "wc_urls", None) or []) if u and u.strip()]
                    ) > 1
                    if not _is_multi_wc:
                        _eff_output_path = _diff_snapshot_pick_main(_step_dir_snapshot_before, config.name)
            except Exception as _e:
                logger.debug(f"[{step.name}] 預算 snapshot diff 失敗(略過):{_e}")

            # exit_code -429 = LLM 配額用盡（executor 標記），直接走 rate_limited 路徑、不再叫 validator（會再 429 一次）
            if exec_result.exit_code == -429:
                val = ValidationResult(
                    status="rate_limited",
                    reason=(exec_result.stderr or "LLM 配額用盡或速率受限（429）"),
                    suggestion="等配額重置或在 Settings 切換 provider（Groq / OpenAI / Anthropic / Ollama 本地）",
                )
            # outlook_automation 節點：agent 自己回 done(success) 就決定成敗了，不需 LLM 驗證
            elif step.outlook_automation:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                # 從 stdout 抽 [Outlook 完成] 那行給使用者看
                _summary = ""
                for _ln in (exec_result.stdout or "").splitlines():
                    if "[Outlook 完成]" in _ln:
                        _summary = _ln.split("[Outlook 完成]", 1)[1].strip()
                        break
                val = ValidationResult(
                    status=_status,
                    reason=_summary or ("Outlook 任務成功" if _status == "ok" else (exec_result.stderr or "Outlook 任務失敗")),
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            # visual_validation 節點：節點自己就是 VLM 判斷，不需要再跑一次 LLM 驗證
            elif step.visual_validation:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                _vv_suggestion = exec_result.stderr if _status == "failed" else ""
                # 若 pptx/docx 失敗且 host 沒裝 LibreOffice → 強烈建議使用者裝
                # 因為 file_preview 走 B1 純文字 PNG、VLM 看到的不是真實版面、永遠評不過
                if _status == "failed":
                    try:
                        _prev_file = _find_prev_output_file(run, config) or ""
                        _ext = _prev_file.lower().rsplit(".", 1)[-1] if "." in _prev_file else ""
                        if _ext in ("pptx", "docx"):
                            from .host_tools import get_libreoffice_status
                            _lo_ok, _ = get_libreoffice_status()
                            if not _lo_ok:
                                _vv_suggestion = (
                                    "⚠ 偵測到主機未裝 LibreOffice、VLM 看到的不是真實 PPT/DOCX 版面"
                                    "(走純文字 PNG 退化路徑、VLM 永遠評不過)。"
                                    "強烈建議:winget install -e --id TheDocumentFoundation.LibreOffice "
                                    "(macOS:brew install --cask libreoffice / Linux:sudo apt install libreoffice)。"
                                    "裝完重啟 backend 再重試。"
                                    + ("\n\n原 VLM 訊息:" + _vv_suggestion if _vv_suggestion else "")
                                )
                    except Exception:
                        pass
                val = ValidationResult(
                    status=_status,
                    reason=exec_result.stdout.replace("[visual_validation] ", "") or "視覺驗證",
                    suggestion=_vv_suggestion,
                )
            # subagent 節點：loop 內已自我驗證、不需要再叫 LLM 驗證一次
            elif step.subagent:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=_status,
                    reason=(exec_result.stdout or "").splitlines()[0] if exec_result.stdout else (
                        "subagent 完成" if _status == "ok" else (exec_result.stderr or "subagent 失敗")
                    ),
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            # web_crawler 節點:
            #  「抓到頁面」≠「抓到真實目標資料」。爬蟲可能成功 fetch 一個 404 頁 / 反爬錯頁 /
            #  空 SPA,exit_code 仍=0。所以:
            #   (a) 有填 expect(AI 驗證節點/規則要求)→ 跑 AI 內容驗證、讀爬蟲輸出判斷是否
            #       真的抓到目標資料(非 404/空頁/錯頁)→ 不真就 fail、別讓下游用 LLM 知識硬補。
            #   (b) 沒填 expect → 至少掃輸出的 status_code / 子頁數,擋掉「全 404 / 全空」的假成功。
            elif step.web_crawler:
                if config.validate and has_expect:
                    logger.info(f"[{step.name}] 🔍 爬蟲節點有 AI 驗證需求 → 驗證抓回的內容是否真實目標資料")
                    val = await validate_step(
                        step_name=step.name,
                        command=step.batch,
                        exit_code=exec_result.exit_code,
                        stdout=exec_result.stdout,
                        stderr=exec_result.stderr,
                        output_path=_eff_output_path,
                        output_expect=step.output.get_expect() if step.output else None,
                        logger=logger,
                        llm_role=getattr(step, "llm_role", "primary"),
                        step_start_time=step_started_at,
                    )
                else:
                    _crawl_fail = _crawl_looks_failed(_eff_output_path, logger)
                    if exec_result.exit_code == 0 and _crawl_fail:
                        val = ValidationResult(
                            status="failed",
                            reason=f"爬蟲表面成功、但內容無效:{_crawl_fail}",
                            suggestion="目標 URL 可能錯誤 / 404 / 被反爬 / SPA 動態渲染抓不到。"
                                       "請確認 URL 正確、或改用其他來源。",
                        )
                    else:
                        _status = "ok" if exec_result.exit_code == 0 else "failed"
                        val = ValidationResult(
                            status=_status,
                            reason=exec_result.stdout.replace("[爬蟲完成] ", "") or "網頁爬取",
                            suggestion=exec_result.stderr if _status == "failed" else "",
                        )
            # computer_use 節點：成敗已由 action 執行結果決定，不需 LLM 驗證
            elif step.computer_use:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=_status,
                    reason=f"桌面自動化 {exec_result.stdout.count('OK')} 個動作成功"
                           + (f"，{exec_result.exit_code} 個失敗" if exec_result.exit_code != 0 else ""),
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and not has_expect:
                # 確定性檢查：exit code=0、輸出檔存在、檔案大小合理（無 AI 驗證節點）
                val = _deterministic_validate(step, exec_result, logger, workflow_name=config.name)
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and has_expect:
                # Recipe 命中但有 AI 驗證節點 → 快速 LLM 驗證（不走 Skill 深度驗證）
                logger.info(f"[{step.name}] 🔍 Recipe 命中 + 有 AI 驗證需求，走快速 LLM 驗證")
                val = await validate_step(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=_eff_output_path,
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                    llm_role=getattr(step, "llm_role", "primary"),
                    step_start_time=step_started_at,
                )
            elif config.validate and has_expect:
                # 使用者填了「預期輸出描述」→ 跑 LLM 驗證
                # 走 deep（agent 自己跑工具驗）的時機：
                #   1. skill 節點 + 有 expect（前端 UI 沒提供 output.skill_mode 開關，純看 expect）
                #   2. script 節點 + AI 驗證節點勾「Skill 模式」→ output.skill_mode=true
                # 兩者以外（script 純 AI 驗證淺、或 skill 沒填 expect）走 shallow validate_step。
                use_skill = step.skill_mode or bool(step.output and step.output.skill_mode)
                validate_fn = validate_step_with_skill if use_skill else validate_step
                val = await validate_fn(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=_eff_output_path,
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                    llm_role=getattr(step, "llm_role", "primary"),
                    step_start_time=step_started_at,
                )
            elif config.validate and not has_expect and step.skill_mode:
                # Skill 節點沒填 expect → 仍走 LLM 淺驗證
                # 理由：LLM 寫的程式碼容易 silent fail（exit_code=0 但結果語意錯，例如
                # 「抓 10 篇」實際只抓到 3 篇），確定性檢查抓不到、需要外層 LLM 看內容把關。
                # 跑 validate_step（淺、單次 LLM call ~5-15s），不走 skill 深度模式。
                logger.info(f"[{step.name}] 🔍 Skill 節點預設驗證（沒填 expect、走淺 LLM 把關 silent fail）")
                val = await validate_step(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=_eff_output_path,
                    output_expect=None,
                    logger=logger,
                    llm_role=getattr(step, "llm_role", "primary"),
                    step_start_time=step_started_at,
                )
            elif config.validate and not has_expect:
                # Script / 其他無 skill_mode 節點 + 沒填 expect → 只做確定性檢查
                # 理由：script 是使用者自己寫的程式、自己負責正確性，外層 LLM 沒足夠上下文判斷
                val = _deterministic_validate(step, exec_result, logger, workflow_name=config.name)
                logger.info(f"[{step.name}] ⚡ Script 節點沒填預期輸出，只看 exit code + 檔案存在")
            else:
                status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=status,
                    reason=f"Exit code {exec_result.exit_code}（LLM 驗證已停用）",
                    suggestion="" if status == "ok" else "請查看 log 取得詳細錯誤",
                )
                logger.info(f"[{step.name}] 驗證（僅 exit code）：{val.status}")

            # ── 算這步真正寫到 workflow dir 的主要檔案 ─────────────────
            # 上面 validate 前已經算過 _eff_output_path、這邊直接重用、不要重複呼 snapshot diff
            actual_out = _eff_output_path or ""

            # ── skill 顯式 export:讀 workflow dir 的 _step_export.json 併進 step_vars ──
            # 讓 skill / script 節點能把算好的「具名乾淨值」傳給下游(尤其 condition 節點 —
            # condition 的 expression 只吃乾淨值、skill 的 stdout 太雜)。
            # 約定:節點在 workflow dir 寫一個扁平 JSON dict 到 _step_export.json,
            # runner 讀進該步的 step_vars(→ {{ steps.<name>.output.<key> }})後刪檔、
            # 避免洩漏到下一步。
            try:
                _export_f = _workflow_output_dir(config.name) / "_step_export.json"
                if _export_f.is_file():
                    import json as _json
                    _exported = _json.loads(_export_f.read_text(encoding="utf-8"))
                    if isinstance(_exported, dict):
                        for _ek, _ev in _exported.items():
                            step_step_vars[str(_ek)] = _ev
                        logger.info(f"[{step.name}] 收到節點 export 變數:{list(_exported.keys())}")
                    _export_f.unlink()
            except Exception as _exp_e:
                logger.warning(f"[{step.name}] 讀 _step_export.json 失敗(忽略):{_exp_e}")

            # ── 自動開放 skill 的 JSON 輸出欄位 → step_vars ──
            # 若這步的輸出檔是個 JSON 物件(扁平 dict),把它的「純量欄位」收進
            # step_vars,讓下游(尤其 condition)能直接 {{ steps.X.output.<欄位> }}
            # 引用 —— skill 只要正常寫它的 JSON 輸出、不用學任何工具或約定。
            try:
                if actual_out and str(actual_out).lower().endswith(".json"):
                    _oj = Path(actual_out)
                    if _oj.is_file():
                        import json as _json2
                        _ojd = _json2.loads(_oj.read_text(encoding="utf-8"))
                        if isinstance(_ojd, dict):
                            _promoted = []
                            for _jk, _jv in _ojd.items():
                                if (isinstance(_jv, (str, int, float, bool))
                                        and str(_jk) not in step_step_vars):
                                    step_step_vars[str(_jk)] = _jv
                                    _promoted.append(str(_jk))
                            if _promoted:
                                logger.info(f"[{step.name}] 自動開放 JSON 輸出欄位:{_promoted}")
            except Exception as _oj_e:
                logger.warning(f"[{step.name}] 讀 JSON 輸出欄位失敗(忽略):{_oj_e}")

            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=exec_result.exit_code,
                stdout_tail=exec_result.stdout[-500:],
                stderr_tail=exec_result.stderr[-200:],
                validation_status=val.status,
                validation_reason=val.reason,
                validation_suggestion=val.suggestion,
                retries_used=retries_used,
                actual_output_path=actual_out,
                token_usage=step_token_usage,
                tool_calls=step_tool_calls,
                started_at=step_started_at,
                ended_at=datetime.now().isoformat(),
                step_vars=dict(step_step_vars),
            )

            # 更新或追加步驟結果
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            # rate_limited（LLM provider 429）：立即暫停，不重試（重試只會再 429 燒配額）
            # 跳到 awaiting_human，讓使用者決定等多久 / 切 provider / 中止
            if val.status == "rate_limited":
                logger.warning(f"步驟 {step_num} ⏸ LLM 配額用盡（429）— 暫停等使用者決策，不重試")
                run.status = "awaiting_human"
                run.awaiting_type = "rate_limited"
                run.awaiting_message = (
                    f"⚠ LLM provider 配額用盡或速率受限（429）\n\n"
                    f"原因：{val.reason}\n\n"
                    f"建議：{val.suggestion}\n\n"
                    f"請選擇：等待重試 / 切換 provider 後重試 / 中止"
                )
                store.save(run)
                return run.run_id

            # ── Phase B: ask_mode 命令授權拒絕/改任務 → 不 retry、直接 awaiting ──
            # executor 攔截敏感命令、用戶按拒絕或改任務後 stderr 會帶這些 prefix
            # 不 retry 因為 LLM 還是會寫同樣的命令 → 反覆觸發授權 = 死結
            if val.status != "ok":
                _se = (exec_result.stderr or "").strip()
                if _se.startswith("使用者拒絕執行敏感命令") or _se.startswith("使用者選擇改任務"):
                    logger.warning(f"步驟 {step_num} 命令授權:{_se[:80]} → 不 retry、等用戶在 failure awaiting 決策")
                    run.status = "awaiting_human"
                    run.awaiting_type = "failure"
                    run.awaiting_message = _se
                    run.awaiting_suggestion = val.suggestion or ""
                    store.save(run)
                    await _notify_failure(run, val, step.name)
                    unregister_task(run.run_id)
                    return run.run_id

            # 缺套件早期攔截：executor 已偵測到 ModuleNotFoundError 且回 missing_packages,
            # 立刻轉 missing_dependency awaiting_human，**不進 step retry**
            # （否則 retry 會讓 LLM 改用「不裝套件、用 API 繞」的策略繞過確認對話框）
            if val.status != "ok":
                _missing = getattr(exec_result, 'missing_packages', None) or []
                if _missing:
                    logger.warning(
                        f"步驟 {step_num} 缺套件 {_missing} → 立即轉 awaiting_human "
                        f"(missing_dependency)、跳過 retry"
                    )
                    import json as _json
                    run.status = "awaiting_human"
                    run.awaiting_type = "missing_dependency"
                    run.awaiting_message = f"缺少套件:{', '.join(_missing)}"
                    run.awaiting_suggestion = _json.dumps({
                        "packages": _missing,
                        "stderr_tail": (exec_result.stderr or "")[-500:],
                        "step_name": step.name,
                        "ai_suggestion": val.suggestion or "",
                    }, ensure_ascii=False)
                    store.save(run)
                    await _notify_failure(run, val, step.name)
                    unregister_task(run.run_id)
                    return run.run_id

            # ── skill agent 主動 done(success=false) → 不 retry、直接 awaiting ──
            # 這是 agent 給的「明確結論」(例:沙盒跑不動、請切 host),不是 crash。
            # 重試只會重跑一輪得出同樣結論、浪費 LLM 額度、也延後使用者看到結論。
            if val.status != "ok" and getattr(exec_result, "agent_concluded_fail", False):
                logger.warning(
                    f"步驟 {step_num} skill agent 主動 done(success=false) → 不 retry、直接 awaiting_human"
                )
                _agent_summary = (exec_result.stderr or "").strip()
                run.status = "awaiting_human"
                run.awaiting_type = "failure"
                run.awaiting_message = _agent_summary or val.reason or ""
                run.awaiting_suggestion = ""
                # TG / 前端卡片是用 val.reason / val.suggestion 組的;agent 的 summary 才是
                # 有意義的失敗結論(含「請切 host」之類引導)、覆蓋掉通用的「Exit code 1」。
                if _agent_summary:
                    val.reason = _agent_summary
                    val.suggestion = ""
                store.save(run)
                await _notify_failure(run, val, step.name)
                unregister_task(run.run_id)
                return run.run_id

            if val.status == "ok":
                logger.info(f"步驟 {step_num} ✅ 通過")
                # 收集延遲儲存的 recipe
                if hasattr(exec_result, 'pending_recipe') and exec_result.pending_recipe:
                    run.pending_recipes.append(exec_result.pending_recipe)
                # 收集此步驟的輸出資訊供後續步驟參考
                # 優先：明確 output.path > snapshot 算出來的 actual_output_path
                # 後者讓沒設 output.path 的 skill 步驟也能被後續 outlook send_with_attachment 自動抓到正確檔
                _eff_path = ""
                if step.output and step.output.path:
                    # 一律存絕對路徑進 completed_outputs，下一步的 LLM agent 拿到
                    # 純檔名也找不到、必須給它絕對路徑（_resolve_path 把純檔名接到
                    # workflow dir、ai_output/... 接專案根、絕對路徑直接用）
                    _eff_path = str(_resolve_path(step.output.path))
                elif actual_out:  # 上面 snapshot diff 算出來的（已是絕對）
                    _eff_path = actual_out
                if _eff_path:
                    out_info = {"path": _eff_path, "schema": ""}
                    try:
                        from pathlib import Path as _Path
                        p = _Path(_eff_path)
                        if p.suffix == ".csv" and p.exists():
                            with open(p, "r") as f:
                                header = f.readline().strip()
                            out_info["schema"] = header
                        elif p.suffix in (".xlsx", ".xls") and p.exists():
                            out_info["schema"] = "Excel 工作簿"
                        elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                            out_info["schema"] = "圖片檔案"
                    except Exception:
                        pass
                    completed_outputs.append(out_info)
                # 一般情況線性前進;若 step.next 有設就改跳到指定 step
                # ("end" / "__end__" → 結束流程;step name → 跳到該 step)
                _next_target = (getattr(step, "next", "") or "").strip()
                if _next_target in ("end", "__end__"):
                    run.current_step = len(config.steps)
                elif _next_target and _next_target in name_to_index:
                    run.current_step = name_to_index[_next_target]
                elif _next_target:
                    logger.warning(f"[{step.name}] next='{_next_target}' 不存在於 workflow、改線性前進")
                    run.current_step += 1
                else:
                    run.current_step += 1
                store.save(run)
                break  # 進入下一步

            elif retries_used < step.retry:
                retries_used += 1
                # 記錄此次失敗的原因與建議，供下次重試時傳給 LLM
                step_failures.append({
                    "attempt": retries_used,
                    "reason": val.reason,
                    "suggestion": val.suggestion,
                    "stdout_tail": exec_result.stdout[-800:] if exec_result.stdout else "",
                    "stderr_tail": exec_result.stderr[-400:] if exec_result.stderr else "",
                })
                logger.warning(
                    f"步驟 {step_num} 驗證失敗，自動重試 {retries_used}/{step.retry}：{val.reason}"
                )
                continue  # 重試

            else:
                # 重試耗盡，暫停等待人為決策
                # 優先使用 LLM 回報的 missing_packages 建立具體安裝建議
                missing_pkgs = getattr(exec_result, 'missing_packages', None) or []
                # 也嘗試從 stderr 偵測 ModuleNotFoundError
                if not missing_pkgs and exec_result.stderr:
                    import re as _re
                    found = _re.findall(r"ModuleNotFoundError: No module named '([^']+)'", exec_result.stderr)
                    if found:
                        missing_pkgs = [p.split(".")[0] for p in found]  # 取頂層套件名
                        missing_pkgs = list(dict.fromkeys(missing_pkgs))  # 去重保序

                run.status = "awaiting_human"
                # ── 缺套件 → 走專屬 awaiting_type=missing_dependency ──
                # TG / 前端會看到特別的「允許安裝 X / 拒絕 / 改任務」按鈕,而不是
                # 一般 failure 模板的「重試 / 跳過 / 中止 / 補充指示」
                if missing_pkgs:
                    run.awaiting_type = "missing_dependency"
                    run.awaiting_message = f"缺少套件：{', '.join(missing_pkgs)}"
                    # awaiting_suggestion 用 JSON 帶結構化資料給 TG / 前端解析
                    import json as _json
                    run.awaiting_suggestion = _json.dumps({
                        "packages": missing_pkgs,
                        "stderr_tail": (exec_result.stderr or "")[-500:],
                        "step_name": step.name,
                        "ai_suggestion": val.suggestion or "",
                    }, ensure_ascii=False)
                    logger.warning(
                        f"步驟 {step_num} 缺套件 {missing_pkgs} → 等待用戶確認安裝"
                    )
                else:
                    # ── 自我修復攔截:開關開 + 次數未滿 + step 可修 → 背景 AI 修復、不轉人工 ──
                    # (缺套件 missing_dependency / rate_limited 在前面的分支已 return、不會走到這、
                    #  那些不該自動改 YAML。只有「一般 failure」才進自我修復。)
                    if _should_self_heal(run, step, val, exec_result):
                        await _enter_self_heal(run, val, step, step_num, exec_result, logger)
                        unregister_task(run.run_id)
                        return run.run_id  # 背景修復中、runner 先退出
                    run.awaiting_type = "failure"
                    run.awaiting_message = val.reason or ""
                    run.awaiting_suggestion = val.suggestion or ""
                    logger.warning(f"步驟 {step_num} 失敗且重試次數耗盡，等待人為決策")

                store.save(run)
                await _notify_failure(run, val, step.name)
                unregister_task(run.run_id)
                return run.run_id  # 暫停

    # ── 全部步驟完成 ─────────────────────────────────────────
    clear_abort(run.run_id)
    unregister_task(run.run_id)
    # 清掉殘留的背景進程(background=true 的 step 留著的 GUI / daemon)
    # 不清的話 workflow 結束後 GUI 還掛著、占記憶體
    try:
        from .executor import kill_run_processes
        kill_run_processes(run.run_id)
    except Exception as _e:
        logger.warning(f"清理背景進程失敗(忽略):{_e}")
    run.status = "completed"
    run.ended_at = datetime.now().isoformat()
    store.save(run)
    logger.info(f"Pipeline {config.name} 全部完成！")
    await _notify_final(run, config)
    return run.run_id


# ── 自我修復(Self-Healing)────────────────────────────────────────────────────
# 某步驟失敗(重試耗盡)且使用者開啟自我修復:讓 AI 助手讀 log + 比對自己寫的 YAML、
# 找 root cause、改 run.config_dict(這次 run 的暫存 YAML、**不碰存檔 workflow**)、從失敗步重跑。
# 到 self_heal_max_attempts / AI 認輸 / 不收斂 → fallback 回 awaiting_type=failure 人工決策。
# 設計與 retry_with_hint 同源(deepcopy/改/存回 run/重跑),差別是改整份 YAML 而非單步 batch。
_SELF_HEAL_HARD_CAP = 5
_SELF_HEAL_LOG_LINES = 160

_SELF_HEAL_SYSTEM_HINT = (
    "【自動修復模式】你正在背景自動修復一個失敗的工作流,沒有使用者在線回答問題。\n"
    "鐵則:\n"
    "1. 仔細讀 log 與目前 YAML、判斷失敗 root cause(YAML 規劃錯 / 路徑錯 / 步驟順序錯 / "
    "缺前置步驟 / 參數錯 / batch 描述不清 / max_iter 太低 等)。\n"
    "2. 只改必要處、輸出**完整**修正後 workflow YAML(用 ```yaml 區塊),不要省略任何步驟。\n"
    "3. 用一兩句說明你改了什麼、為什麼。\n"
    "4. 若判斷這**不是改 YAML 能修**的(外部服務 503 / 需使用者授權 / 缺套件)→ 明講"
    "「無法自動修復」+ 原因,**不要硬改、不要輸出 YAML**。\n"
    "5. 輸出格式跟你平常生成工作流給桌面 web 完全一樣:**先單獨寫一行 `YAML_READY`、"
    "緊接著 ```yaml 區塊**(系統靠 YAML_READY 標記偵測並自動套用)。少了 YAML_READY 系統會抓不到、修復就失敗。\n"
    "6. **絕對不要呼叫 save_workflow_yaml / start_workflow 或任何工具** —— 只要 emit YAML_READY + ```yaml``` 即可。"
)


def _step_is_healable(step) -> bool:
    """只修 AI 靠改 YAML 能救的節點。human_confirm(人為)/ computer_use / visual_validation /
    outlook(環境相關、改 YAML 沒用)/ condition(分支邏輯、極少是失敗源)不修。"""
    for attr in ("human_confirm", "computer_use", "visual_validation",
                 "outlook_automation", "condition"):
        if getattr(step, attr, False):
            return False
    return True


def _self_heal_max() -> int:
    try:
        from settings import get_settings
        return min(_SELF_HEAL_HARD_CAP, int(get_settings().get("self_heal_max_attempts", 2) or 2))
    except Exception:
        return 2


def _is_stuck_failure(val, exec_result) -> bool:
    """卡死 / 不收斂型失敗(步驟跑很久才失敗:subagent 撞 max_iter、timeout、連續無 tool)。
    這類重跑會再卡一次、改 YAML 也救不了根本問題 → 不修、直接轉人工,避免把時間放大數倍
    (實測 H3 coder 撞 max_iter、self_heal 重跑 → 卡 50 分鐘)。對齊 [[non-convergence-guard]]。"""
    blob = (
        (getattr(val, "reason", "") or "") + " "
        + (getattr(exec_result, "error", "") or "") + " "
        + (getattr(exec_result, "stderr", "") or "")
    ).lower()
    for kw in ("reached_max_iter", "max_iter", "timeout", "timed out", "逾時",
               "不收斂", "非收斂", "consecutive_no_tool"):
        if kw in blob:
            return True
    return False


def _should_self_heal(run, step, val=None, exec_result=None) -> bool:
    try:
        from settings import get_settings
        s = get_settings()
    except Exception:
        return False
    if not s.get("self_heal_enabled", False):
        return False
    if getattr(run, "self_heal_count", 0) >= _self_heal_max():
        return False
    if not _step_is_healable(step):
        return False
    # 卡死型失敗不重跑(重跑會再卡、放大時間)→ 交人工
    if _is_stuck_failure(val, exec_result):
        return False
    return True


def _ai_gave_up(reply: str) -> bool:
    if not reply:
        return False
    low = reply.replace(" ", "")
    for kw in ("無法自動修復", "無法修復", "不是改YAML", "非YAML能修", "無法靠YAML",
               "需要使用者", "需要授權", "外部服務", "cannotfix", "cannotbefixed"):
        if kw.replace(" ", "") in low:
            return True
    return False


def _yaml_near_identical(a: str, b: str) -> bool:
    """新舊 YAML 是否「沒有實質修改」→ 視為不收斂。
    比 steps 的語義內容(name/batch/各節點旗標),不用字串相似度 —— 字串 ratio 對小工作流
    太鈍(改一個關鍵字如 prnt→print 的 ratio 仍 >0.985 會被誤判沒改)。只要任何 step 的
    name/batch 變了就算有實質改動、放行。"""
    import yaml as _y

    def _sig(s):
        try:
            d = _y.safe_load(s) or {}
            steps = d.get("steps", []) if isinstance(d, dict) else []
            return [
                (st.get("name"), st.get("batch"), st.get("skill_mode"),
                 st.get("subagent_role"), st.get("subagent_max_iter"))
                for st in steps if isinstance(st, dict)
            ]
        except Exception:
            return None

    sa, sb = _sig(a), _sig(b)
    if sa is None or sb is None:
        # parse 失敗 → 退回字串完全相等比較(只有一字不差才算沒改)
        return (a or "").strip() == (b or "").strip()
    return sa == sb


def _format_prior_attempts(prior: list) -> str:
    if not prior:
        return "(這是第一次修復、沒有先前嘗試)"
    lines = []
    for p in prior:
        lines.append(f"- 第 {p.get('attempt', '?')} 次:{(p.get('diagnosis') or '')[:300]}")
    return "\n".join(lines)


async def _notify_self_heal(run: PipelineRun, text: str):
    """自我修復進度通知(TG)。desktop run 沒 chat_id 時 _tg_send 自行略過。"""
    try:
        await _tg_send(run.telegram_chat_id, f"🔧 <b>{run.pipeline_name}</b>\n{text}")
    except Exception:
        pass


async def _enter_self_heal(run, val, step, step_num, exec_result, logger) -> None:
    """進入修復過渡狀態 + 背景啟動修復(不阻塞 runner)。"""
    max_n = _self_heal_max()
    run.self_heal_count = getattr(run, "self_heal_count", 0) + 1
    run.status = "awaiting_human"
    run.awaiting_type = "self_heal"
    run.awaiting_message = f"AI 自我修復中(第 {run.self_heal_count}/{max_n} 次)…失敗步驟:{step.name}"
    run.awaiting_suggestion = ""
    get_store().save(run)
    logger.warning(
        f"步驟 {step_num}「{step.name}」失敗 → 啟動自我修復(第 {run.self_heal_count}/{max_n} 次)"
    )
    await _notify_self_heal(
        run, f"步驟「{step.name}」失敗、AI 正在自動修復(第 {run.self_heal_count}/{max_n} 次)…"
    )
    _stderr_tail = ""
    if exec_result is not None and getattr(exec_result, "stderr", ""):
        _stderr_tail = (exec_result.stderr or "")[-1500:]
    asyncio.create_task(_run_self_heal_then_resume(
        run.run_id, step.name, run.current_step,
        val.reason or "", val.suggestion or "", _stderr_tail,
    ))


async def _run_self_heal_then_resume(run_id, failed_step_name, failed_step_index,
                                     fail_reason, fail_suggestion, stderr_tail):
    """背景:讀 log → AI 改 YAML → 套用 → 從失敗步重跑。失敗則 fallback 人工。"""
    store_ = get_store()
    run = store_.load(run_id)
    if not run:
        return
    logger = resume_run_logger(run.run_id, run.log_path)

    def _fallback(msg):
        run.status = "awaiting_human"
        run.awaiting_type = "failure"
        run.awaiting_message = msg
        run.awaiting_suggestion = fail_suggestion
        store_.save(run)

    try:
        import yaml as _yaml
        log_tail = get_run_log_tail(run_id, lines=_SELF_HEAL_LOG_LINES)
        current_yaml = _yaml.safe_dump(run.config_dict, allow_unicode=True, sort_keys=False)
        prior = getattr(run, "self_heal_history", []) or []

        heal_instruction = (
            "工作流「" + run.pipeline_name + "」執行到步驟「" + failed_step_name
            + "」失敗,需要你修復。\n\n"
            "=== 驗證判定 ===\n" + (fail_reason or "(無)") + "\n\n"
            "=== 原建議 ===\n" + (fail_suggestion or "(無)") + "\n\n"
            "=== 失敗步驟 stderr(末段)===\n" + (stderr_tail or "(無)") + "\n\n"
            "=== 執行 log(末段、錯誤通常在這)===\n" + (log_tail or "(無)") + "\n\n"
            "=== 目前完整 YAML ===\n```yaml\n" + current_yaml + "\n```\n\n"
            "=== 之前的修復嘗試(別重蹈覆轍)===\n" + _format_prior_attempts(prior) + "\n\n"
            "請判斷 root cause、輸出修正後的**完整** YAML(```yaml 區塊)、並用一兩句說明改了什麼。"
            "若判斷非 YAML 能修(外部服務掛 / 需授權 / 缺套件)→ 明說「無法自動修復」+ 原因、不要輸出 YAML。"
        )

        from main import _chat_agent_loop, PipelineChatRequest
        # 不帶 workflow_id:heal_instruction 已含完整 current YAML,再注入「現有工作流」context
        # 會讓 AI 進入「改現有工作流」模式、傾向呼叫 save_workflow_yaml 工具(要兩步確認)
        # 而非直接在回覆吐 ```yaml``` → 我們抓不到 yaml_content。實測這是修復失敗主因。
        result = await _chat_agent_loop(PipelineChatRequest(
            messages=[{"role": "user", "content": heal_instruction}],
            workflow_id=None,
            extra_system=_SELF_HEAL_SYSTEM_HINT,
        )) or {}

        new_yaml = result.get("yaml_content")
        yaml_err = result.get("yaml_error")
        reply = result.get("reply") or ""

        if not new_yaml or yaml_err or _ai_gave_up(reply):
            reason = yaml_err or (("AI 判定無法修復:" + reply[:200]) if _ai_gave_up(reply)
                                  else "AI 未產出修正 YAML")
            logger.warning("自我修復未產出可用 YAML → 轉人工:" + str(reason)[:200])
            _fallback("AI 自我修復未解決(" + str(reason)[:200] + ")。原失敗:" + (fail_reason or ""))
            await _notify_self_heal(run, "⚠️ AI 自我修復未能解決、轉交人工決策。")
            return

        if _yaml_near_identical(new_yaml, current_yaml):
            logger.warning("自我修復新舊 YAML 幾乎相同(疑似不收斂)→ 轉人工")
            _fallback("AI 自我修復未產生實質變更(疑似不收斂)。原失敗:" + (fail_reason or ""))
            await _notify_self_heal(run, "⚠️ AI 修復沒有實質變更、轉交人工決策。")
            return

        try:
            new_dict = _yaml.safe_load(new_yaml)
            if isinstance(new_dict, dict) and "pipeline" in new_dict:
                new_dict = new_dict["pipeline"]
            if not isinstance(new_dict, dict):
                raise ValueError("YAML 頂層不是 mapping")
            PipelineConfig.from_dict(new_dict)  # schema 驗、不合會 raise
        except Exception as _e:
            logger.warning("自我修復 YAML 不合 schema → 轉人工:" + str(_e)[:200])
            _fallback("AI 修復的 YAML 不合格式:" + str(_e)[:200] + "。原失敗:" + (fail_reason or ""))
            await _notify_self_heal(run, "⚠️ AI 修復的 YAML 格式錯誤、轉人工。")
            return

        # 保留內部旗標(_workflow_id 等),新 YAML 不會帶 → 從舊 config_dict 補回
        for k, v in (run.config_dict or {}).items():
            if k.startswith("_") and k not in new_dict:
                new_dict[k] = v

        import hashlib as _hl
        run.self_heal_history = list(prior) + [{
            "attempt": run.self_heal_count,
            "diagnosis": reply[:600],
            "old_yaml_hash": _hl.md5(current_yaml.encode("utf-8", "replace")).hexdigest()[:12],
        }]
        run.config_dict = new_dict

        # 重定位失敗步(名稱對得到 → 回該步;對不到 → 從 0 重跑整條)
        restart_idx = 0
        for i, st in enumerate(new_dict.get("steps", [])):
            if isinstance(st, dict) and st.get("name") == failed_step_name:
                restart_idx = i
                break
        run.current_step = restart_idx
        run.step_results = [sr for i, sr in enumerate(run.step_results) if i < restart_idx]
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store_.save(run)
        logger.info(
            "自我修復套用新 YAML、從步驟 " + str(restart_idx + 1) + " 重跑。診斷:" + reply[:150]
        )
        await _notify_self_heal(
            run, "✅ AI 已修復、從步驟 " + str(restart_idx + 1) + " 重跑。\n修了什麼:" + reply[:300]
        )

        async def _delayed_heal_start():
            await asyncio.sleep(0.3)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=restart_idx,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_heal_start())

    except Exception as e:
        logger.error("自我修復過程出錯:" + type(e).__name__ + ": " + str(e))
        try:
            _fallback("AI 自我修復過程出錯:" + str(e)[:200] + "。原失敗:" + (fail_reason or ""))
        except Exception:
            pass


# ── Human-in-the-loop resume ─────────────────────────────────────────────────

async def resume_pipeline(run_id: str, decision: str, hint: str = "") -> str:
    """
    用戶透過 Telegram inline keyboard 做出決策後，呼叫此函式繼續執行。

    Args:
        run_id:   pipeline run id
        decision: "retry" | "skip" | "abort" | "continue" | "retry_with_hint"
        hint:     補充指示（retry_with_hint 時使用）

    Returns:
        str 回應訊息（回覆給用戶）
    """
    store = get_store()
    run = store.load(run_id)

    if not run:
        return f"❌ 找不到 Pipeline run：{run_id}"
    if run.status != "awaiting_human":
        return f"⚠️ Pipeline {run_id} 目前狀態為 {run.status}，無需決策"

    config = PipelineConfig.from_dict(run.config_dict)
    step_num = run.current_step + 1
    total = len(config.steps)
    # 附加到原始 log 檔，確保前端讀到的 log_path 始終指向同一個檔案
    logger = resume_run_logger(run.run_id, run.log_path)

    # ── ask_user 回答：skill agent 仍在 in-memory 等待 event ──
    if run.awaiting_type == "ask_user":
        from pipeline.executor import deliver_ask_user_answer
        if decision == "answer":
            ok = deliver_ask_user_answer(run_id, hint)
            if not ok:
                # agent 可能已 timeout 或後端已重啟
                return "⚠️ 答案送達失敗：skill agent 已不在等待狀態（可能逾時或後端重啟）"
            logger.info(f"[ask_user] 使用者答案已送達：{hint[:100]}")
            return f"✅ 答案已送出"
        elif decision == "abort":
            # ⚠️ 修正:ask_user 等待時 skill agent 仍是「活著的 asyncio task、卡在
            # _wait_for_ask_user 的 await」。舊版只 deliver 空答案 → agent 拿到 "" 會
            # 繼續迴圈(再問一次),根本沒停;只改 store.status 也不會中斷正在跑的 task。
            # 正解:走跟右上角停止鈕同一套 force_abort —— kill 子進程 + cancel task
            # (CancelledError 會在 _wait_for_ask_user 的 await 拋出、agent 真的中斷)+ 標 aborted。
            logger.info("[ask_user] 使用者選擇中止 → force_abort(取消執行中的 task)")
            await force_abort(run_id)
            return f"🛑 Pipeline 已中止"
        else:
            return f"⚠️ ask_user 只接受 answer 或 abort，收到 {decision}"

    if decision == "abort":
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger.info("用戶選擇中止 Pipeline")
        await _notify_final(run, config)
        return f"🛑 Pipeline 已中止（步驟 {step_num}/{total}）"

    elif decision == "skip":
        logger.info(f"用戶選擇跳過步驟 {step_num}")
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            await _notify_final(run, config)
            return f"⏩ 跳過最後一步，Pipeline 完成"

        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_skip():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_skip())
        return f"⏩ 跳過步驟 {step_num}，繼續執行步驟 {step_num + 1}/{total}"

    elif decision == "retry":
        logger.info(f"用戶選擇重試步驟 {step_num}")
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_retry():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=run.current_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_retry())
        return f"🔄 重試步驟 {step_num}/{total}"

    elif decision == "redo_prev":
        # 重做上一個 step:對應 user 看到當前 step 失敗、判斷是因為上一步沒做好、想回頭重來
        # 例:step 5 VLM 驗 PPT 排版失敗 → 重做上一步(step 4 產 PPT)、再次推進到 step 5 驗
        if run.current_step <= 0:
            return "⚠️ 已經是第一步、無法重做上一步"
        prev_step = run.current_step - 1
        logger.info(f"用戶選擇重做上一步 {prev_step + 1}(原失敗在 {step_num})")
        # 清掉當前失敗 step 的 result + 上一步 result(讓兩步都重跑)
        run.step_results = [sr for sr in run.step_results if sr.step_index < prev_step]
        run.current_step = prev_step
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_redo_prev():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=prev_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_redo_prev())
        return f"↩ 重做上一步({prev_step + 1}/{total}),完成後再次推進到原步驟"

    elif decision == "install_dep":
        # 用戶從 missing_dependency awaiting_human 按「允許安裝」
        # hint 帶要安裝的套件名（單一名 or 逗號分隔多個）。
        # 安裝完同步驟 retry。
        logger.info(f"用戶允許安裝套件：{hint}")
        if not hint or not hint.strip():
            return "⚠️ install_dep 需要 hint 帶套件名"
        pkgs_to_install = [p.strip() for p in hint.split(",") if p.strip()]
        if not pkgs_to_install:
            return "⚠️ 沒有有效套件名"

        # 判斷裝到哪：sandbox 模式裝容器、否則裝 host venv
        try:
            import sys as _sys
            from pathlib import Path as _PI
            _backend = str(_PI(__file__).parent.parent.absolute())
            if _backend not in _sys.path:
                _sys.path.insert(0, _backend)
            from settings import get_settings as _gs
            sandbox_mode = _gs().get("skill_sandbox_mode", "host") == "wsl_docker"
        except Exception:
            sandbox_mode = False

        from skill_pkg_manager import add_package, add_package_sandbox
        installer = add_package_sandbox if sandbox_mode else add_package
        target = "sandbox" if sandbox_mode else "host"
        results = []
        for pkg in pkgs_to_install:
            ok, msg = installer(pkg)
            results.append((pkg, ok, msg))
            logger.info(f"[install_dep] {target} pip install {pkg}: ok={ok}, msg={msg}")

        all_ok = all(ok for _, ok, _ in results)
        if not all_ok:
            # 安裝失敗 → 回到 awaiting_human、訊息更新（讓用戶看到失敗原因）
            failed = [(p, m) for p, ok, m in results if not ok]
            run.awaiting_message = f"安裝失敗：{', '.join(p for p, _ in failed)}"
            run.awaiting_suggestion = "\n".join(f"• {p}: {m[:200]}" for p, m in failed)
            # awaiting_type 維持 missing_dependency（讓用戶可以「改任務」或重試）
            store.save(run)
            return f"❌ 安裝失敗：{', '.join(p for p, _ in failed)}"

        # 全部 ok → retry 該步驟
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_install_retry():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=run.current_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_install_retry())
        installed = ", ".join(p for p, _, _ in results)
        return f"✅ 已在 {target} 安裝：{installed}\n🔄 重試步驟 {step_num}/{total}"

    elif decision in ("approve_command", "deny_command", "hint_command"):
        # Phase B: ask_mode 命令授權回應。送 event 喚醒 skill agent loop
        # agent loop 收到後依 decision 決定執行 / 退出 / 改任務
        from pipeline.executor import deliver_command_approval
        # decision name → executor 內部 mapping
        _map = {
            "approve_command": "allow",
            "deny_command":    "deny",
            "hint_command":    "hint",
        }
        outcome = _map[decision]
        ok = deliver_command_approval(run.run_id, outcome)
        if not ok:
            return f"⚠️ 沒找到等待中的命令授權（可能已逾時或被其他人處理）"
        logger.info(f"用戶命令授權決定:{outcome}")
        return {
            "approve_command": "✅ 已允許執行命令、繼續...",
            "deny_command":    "❌ 已拒絕、終止此步驟",
            "hint_command":    "💬 已退出 step、稍後可在 failure awaiting 按「補充指示」",
        }[decision]

    elif decision == "retry_with_hint":
        import copy
        # 1. 使用深拷貝，確保 config 修改是獨立且完整的
        config_d = copy.deepcopy(run.config_dict)
        steps = config_d.get("steps", [])

        is_confirm = run.awaiting_type == "human_confirm"
        target = run.current_step

        if is_confirm:
            prev_step = run.current_step - 1
            while prev_step >= 0 and steps[prev_step].get("human_confirm"):
                prev_step -= 1
            if prev_step < 0:
                return "⚠️ 確認節點前沒有可重做的步驟"
            # 防呆：只有 skill_mode 節點能消化 hint。shell / computer_use 重跑 hint 無意義或會壞掉，
            # 正常 UI 不會給這個按鈕，但舊訊息或外部 API 呼叫還是可能打進來 → 拒絕
            if not steps[prev_step].get("skill_mode"):
                return (
                    "⚠️ 上一步不是 AI 技能節點，無法使用補充指示。"
                    "補充指示會附加給 LLM 重新生成程式碼；shell / 桌面自動化節點沒有 LLM 可消化。"
                )
            target = prev_step

        if target < len(steps):
            original_batch = steps[target].get("batch", "")
            # 清理舊的提示詞標籤，避免重複疊加
            clean_batch = original_batch.split("【用戶補充指示】")[0].strip()
            steps[target]["batch"] = f"{clean_batch}\n\n【用戶補充指示】{hint}"
            config_d["steps"] = steps

        # 2. 更新 run 狀態並「立即」同步回資料庫
        run.config_dict = config_d
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        # 3. 關鍵修正：給 Windows 一點點時間釋放資料庫鎖定
        async def _delayed_start():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=config_d,  # 傳入已修改的配置
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=target,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_start())
        
        if is_confirm:
            return f"💬 已附加指示，重做步驟 {target + 1}/{total}"
        else:
            return f"💬 已附加指示，重試步驟 {step_num}/{total}"

    elif decision == "continue":
        # 人工確認通過 → 繼續下一步
        logger.info(f"用戶確認通過步驟 {step_num}，繼續執行")

        # 更新確認步驟結果
        if run.current_step < len(run.step_results):
            run.step_results[run.current_step].validation_reason = "人工確認 — 已通過"
            run.step_results[run.current_step].stdout_tail = "已確認通過"

        run.awaiting_type = ""
        run.awaiting_message = ""
        next_step = run.current_step + 1

        if next_step >= total:
            # 清掉殘留的背景進程(同 main loop 結束時邏輯)
            try:
                from .executor import kill_run_processes
                kill_run_processes(run.run_id)
            except Exception as _e:
                logger.warning(f"清理背景進程失敗(忽略):{_e}")
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info(f"Pipeline {run.pipeline_name} 全部完成！")
            await _notify_final(run, config)
            return f"✅ 確認通過,Pipeline 全部完成"

        run.status = "running"
        store.save(run)

        async def _delayed_continue():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_continue())
        return f"✅ 確認通過，繼續執行步驟 {next_step + 1}/{total}"

    elif decision == "self_heal_now":
        # 使用者在 failure 卡片手動點「讓 AI 試修」— 觸發一次自我修復(不看開關、仍受硬上限)
        if run.awaiting_type not in ("failure", "self_heal"):
            return "⚠️ 目前狀態無法觸發自我修復"
        if getattr(run, "self_heal_count", 0) >= _SELF_HEAL_HARD_CAP:
            return f"⚠️ 已達自我修復硬上限({_SELF_HEAL_HARD_CAP} 次)、請改用其他決策"
        failed_idx = run.current_step
        steps_ = run.config_dict.get("steps", []) if isinstance(run.config_dict, dict) else []
        failed_name = steps_[failed_idx].get("name", "") if failed_idx < len(steps_) else ""
        fr = run.awaiting_message or ""
        fs = run.awaiting_suggestion or ""
        stderr_tail = ""
        if failed_idx < len(run.step_results):
            stderr_tail = (getattr(run.step_results[failed_idx], "stderr_tail", "") or "")[-1500:]
        run.self_heal_count = getattr(run, "self_heal_count", 0) + 1
        run.status = "awaiting_human"
        run.awaiting_type = "self_heal"
        run.awaiting_message = f"AI 自我修復中(手動觸發、第 {run.self_heal_count} 次)…失敗步驟:{failed_name}"
        run.awaiting_suggestion = ""
        store.save(run)
        asyncio.create_task(_run_self_heal_then_resume(
            run.run_id, failed_name, failed_idx, fr, fs, stderr_tail,
        ))
        return f"🔧 已啟動 AI 自我修復(第 {run.self_heal_count} 次)、修好會自動重跑"

    return "❓ 未知決策"
