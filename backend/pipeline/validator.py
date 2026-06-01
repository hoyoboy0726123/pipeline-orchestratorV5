"""
LLM 語意驗證器。

不靠關鍵字比對，讓 LLM 理解整體 log 內容，
判斷步驟是否真正成功——能區分「Python WARNING 不代表失敗」
與「真正的 Exception / 資料異常」。

支援：
- 文字檔讀取前 N 行供 LLM 判斷
- CSV / JSON / Excel 結構化摘要（欄位、行數、樣本）
- 圖片檔以 base64 傳給 Vision model 做視覺驗證
- Skill 模式：LLM 主動執行 Python / Shell 驗證程式碼（ReAct agent）
"""
import asyncio
import base64
import csv
import io
import json
import logging
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage, SystemMessage

from config import GROQ_API_KEY, GROQ_MODEL_MAIN

# Skill 工具執行的 timeout（秒）
SKILL_TOOL_TIMEOUT = 60
# Skill agent 最大迭代次數（防止無限迴圈）
SKILL_MAX_ITERATIONS = 15
# Groq Free tier: 30 RPM → 每次 LLM 呼叫間隔至少 2 秒
SKILL_REQUEST_INTERVAL = 2.0
# 每 N 次 LLM 呼叫後強制冷卻
SKILL_COOLDOWN_EVERY = 14
SKILL_COOLDOWN_SECONDS = 60


@dataclass
class ValidationResult:
    status: str      # "ok" | "warning" | "failed" | "rate_limited"
    reason: str      # 中文說明
    suggestion: str  # LLM 建議的修復方向（failed 時才有意義）


def _is_rate_limit_error(e: Exception) -> bool:
    """偵測 LLM provider 的配額/速率錯誤。429 / RESOURCE_EXHAUSTED 都算。
    用於避免「驗證失敗 → fallback 再叫 LLM → 又 429」的連環燒配額。"""
    s = str(e)
    return ("429" in s or "RESOURCE_EXHAUSTED" in s or "quota" in s.lower()
            or "rate limit" in s.lower() or "rate_limit" in s.lower())


# 雙模型 cache:key = (role, settings_signature)
_llm_cache: dict[tuple[str, str], Any] = {}


def _get_llm(role: str = "primary"):
    from settings import settings_signature
    from llm_factory import build_llm
    sig = settings_signature()
    key = (role, sig)
    if key not in _llm_cache:
        _llm_cache[key] = build_llm(temperature=0, role=role)
    return _llm_cache[key]


# ── 檔案內容讀取 ──────────────────────────────────────────────────────────────

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}
STRUCTURED_EXTS = {'.csv', '.json', '.jsonl', '.xlsx', '.xls'}
TEXT_EXTS = {'.txt', '.log', '.md', '.html', '.xml', '.yaml', '.yml', '.py', '.sh', '.js', '.ts'}
MAX_TEXT_LINES = 50
MAX_CSV_ROWS = 10


def _read_file_content(path: Optional[str]) -> dict:
    """
    讀取輸出檔案，回傳結構化資訊供 LLM 分析。

    Returns:
        {
            "summary": str,       # 給 prompt 的文字摘要
            "image_b64": str|None # base64 圖片（僅圖檔）
            "image_mime": str|None
        }
    """
    result = {"summary": "", "image_b64": None, "image_mime": None}
    if not path:
        return result

    p = _resolve_user_path(path)
    if not p.exists():
        return result

    # 目錄：列出檔案清單
    if p.is_dir():
        files = sorted(p.iterdir())[:20]
        listing = "\n".join(f"  {'📁' if f.is_dir() else '📄'} {f.name} ({f.stat().st_size:,} bytes)" for f in files)
        result["summary"] = f"目錄內容（前 20 項）：\n{listing}"
        return result

    ext = p.suffix.lower()

    # 圖片檔 → base64
    if ext in IMAGE_EXTS:
        try:
            data = p.read_bytes()
            if len(data) <= 20 * 1024 * 1024:  # ≤ 20MB
                result["image_b64"] = base64.b64encode(data).decode()
                mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                           '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
                result["image_mime"] = mime_map.get(ext, 'image/png')
                result["summary"] = f"圖片檔 {p.name}（{len(data):,} bytes），已附圖供視覺分析"
        except Exception as e:
            result["summary"] = f"圖片讀取失敗：{e}"
        return result

    # CSV
    if ext == '.csv':
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                result["summary"] = "CSV 檔案為空"
                return result
            header = rows[0]
            data_rows = rows[1:]
            sample = data_rows[:MAX_CSV_ROWS]
            sample_str = "\n".join([",".join(r) for r in sample])
            result["summary"] = (
                f"CSV 檔案：{p.name}\n"
                f"欄位（{len(header)} 個）：{', '.join(header)}\n"
                f"資料行數：{len(data_rows)}\n"
                f"前 {min(len(sample), MAX_CSV_ROWS)} 行樣本：\n{sample_str}"
            )
        except Exception as e:
            result["summary"] = f"CSV 讀取失敗：{e}"
        return result

    # JSON / JSONL
    if ext in ('.json', '.jsonl'):
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            if ext == '.jsonl':
                lines = [l for l in text.strip().split('\n') if l.strip()]
                result["summary"] = (
                    f"JSONL 檔案：{p.name}，共 {len(lines)} 行\n"
                    f"前 {min(5, len(lines))} 行樣本：\n" +
                    "\n".join(lines[:5])
                )
            else:
                data = json.loads(text)
                if isinstance(data, list):
                    sample = json.dumps(data[:5], ensure_ascii=False, indent=2)
                    result["summary"] = f"JSON 陣列：{p.name}，共 {len(data)} 筆\n前 5 筆樣本：\n{sample}"
                elif isinstance(data, dict):
                    keys = list(data.keys())[:20]
                    result["summary"] = f"JSON 物件：{p.name}\n鍵（前 20 個）：{', '.join(keys)}\n內容預覽：\n{json.dumps(data, ensure_ascii=False, indent=2)[:1000]}"
                else:
                    result["summary"] = f"JSON 檔案：{p.name}\n內容：{text[:500]}"
        except Exception as e:
            result["summary"] = f"JSON 讀取失敗：{e}"
        return result

    # Excel
    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=MAX_CSV_ROWS + 1, values_only=True))
                if not rows:
                    sheets_info.append(f"  Sheet「{sheet_name}」：空")
                    continue
                header = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]
                total_rows = ws.max_row - 1 if ws.max_row else 0
                sample_lines = []
                for r in data_rows[:MAX_CSV_ROWS]:
                    sample_lines.append(",".join(str(c) if c is not None else "" for c in r))
                sheets_info.append(
                    f"  Sheet「{sheet_name}」：{total_rows} 行，{len(header)} 欄\n"
                    f"    欄位：{', '.join(header)}\n"
                    f"    前 {len(sample_lines)} 行：\n    " + "\n    ".join(sample_lines)
                )
            wb.close()
            result["summary"] = f"Excel 檔案：{p.name}，共 {len(wb.sheetnames)} 個 Sheet\n" + "\n".join(sheets_info)
        except ImportError:
            result["summary"] = f"Excel 檔案：{p.name}（需安裝 openpyxl 才能讀取內容）"
        except Exception as e:
            result["summary"] = f"Excel 讀取失敗：{e}"
        return result

    # 一般文字檔（含未知副檔名）
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= MAX_TEXT_LINES:
                    break
                lines.append(line.rstrip())
        total_size = p.stat().st_size
        result["summary"] = (
            f"文字檔：{p.name}（{total_size:,} bytes）\n"
            f"前 {len(lines)} 行：\n" + "\n".join(lines)
        )
    except Exception:
        result["summary"] = f"檔案 {p.name} 無法以文字方式讀取"

    return result


# ── 主驗證函式 ─────────────────────────────────────────────────────────────────

async def validate_step(
    step_name: str,
    command: str,
    exit_code: int,
    stdout: str,
    stderr: str,
    output_path: Optional[str],
    output_expect: Optional[str],
    logger: logging.Logger,
    llm_role: str = "primary",
    step_start_time: Optional[str] = None,
) -> ValidationResult:
    """
    使用 LLM 語意分析執行結果，回傳結構化驗證結論。

    LLM 會考量：
    - exit code 與其含意
    - stdout/stderr 的語意（區分警告與錯誤）
    - 輸出檔案是否存在、大小是否合理
    - 輸出檔案內容（文字前 50 行 / CSV 結構 / Excel 摘要）
    - 圖片檔案以視覺方式驗證
    - 是否符合 expect 描述的期望
    """
    # 收集輸出檔案資訊
    file_info = _check_output_file(output_path, step_start_time=step_start_time)
    file_content = _read_file_content(output_path)

    # 截取重要片段（節省 token）
    stdout_tail = stdout[-1000:] if len(stdout) > 1000 else stdout
    stderr_tail = stderr[-500:] if len(stderr) > 500 else stderr

    prompt_text = f"""你是一個精確的 pipeline 步驟驗證器。
分析以下執行結果，判斷步驟是否成功。

【步驟資訊】
名稱：{step_name}
命令：{command}
Exit Code：{exit_code}
預期輸出描述：{output_expect or "無特定要求"}
輸出路徑：{output_path or "無"}
檔案狀態：{file_info}

【stdout（最後部分）】
```
{stdout_tail or "（無輸出）"}
```

【stderr（最後部分）】
```
{stderr_tail or "（無輸出）"}
```"""

    # 加入檔案內容摘要
    if file_content["summary"]:
        prompt_text += f"""

【輸出檔案內容】
{file_content["summary"]}"""

    # 如果是圖片，加入視覺分析提示
    if file_content["image_b64"]:
        prompt_text += """

【圖片分析】
已附上輸出的圖片檔案，請以視覺方式分析圖片內容是否符合預期描述。
檢查圖片是否正常渲染、內容是否完整、是否符合期望。"""

    prompt_text += """

請只回傳以下 JSON，不要加任何其他文字：
{
  "status": "ok",
  "reason": "一句話說明判斷結果",
  "suggestion": "如果 failed，給出修復建議；ok 時留空字串"
}

【判斷規則】
- "ok"：步驟成功，exit code 0，輸出符合預期（若有）
- "warning"：步驟完成但有非致命問題（如 deprecation warning、部分資料遺失），建議人工確認
- "failed"：步驟失敗，需要介入（exit code 非 0 且 stderr 有真實錯誤、Exception、缺少必要輸出檔案等）

注意：Python DeprecationWarning、UserWarning 不代表失敗；只有真正的 Exception / Error / 致命問題才判為 failed。

【Skill / Outlook 步驟的特殊規則】
若 stdout 出現 `[Skill 完成]` 或 `[Outlook 完成]` 標記、且 exit code 為 0，代表 agent
已完成試錯並成功產出結果。**就算 stdout 含早期 iteration 的 Traceback，也不該判 failed**
（那是 agent 試錯過程的正常產物，並非最終狀態）。
此情境只看：(1) 完成標記是否存在、(2) 輸出檔案是否存在且內容符合預期。

【舊檔冒充新檔的攔截】
若【檔案狀態】欄出現「⚠️ 警告：檔案 mtime ... 早於本步驟開始時間」、或目錄內所有檔案 mtime 早於本步驟開始時間，
代表 agent **沒有真正產出新檔**、輸出的是先前 run 的殘留檔。**此情境必判 failed**、
即使 stdout 有 [Skill 完成] 標記也一樣（agent 可能用虛構工具結果騙過 done）。

【可疑歸零結果的攔截】
若本步驟的工作性質是「抽取 / 篩選 / 解析 / 爬取 / 整理」這類「從輸入資料產出記錄」的任務
（看命令與步驟名稱判斷），而**輸出是空的或近乎空的** —
例如:0 筆記錄、空檔案、JSON 內容是空陣列 `[]`、stdout 出現「0 筆」「filtered 0」「抽出 0」
「找到 0」「共 0」之類字樣 —
這**通常不是真實結果、而是解析 / 篩選 / 選擇器邏輯出錯**（CSS selector 沒對上、切塊規律抓錯、
正規表達式沒命中等）。上游既然有資料餵進來、下游卻一筆都產不出來，幾乎可以斷定是 bug。
**除非任務本身明確說「可能為空也算正常」、否則此情境判 `failed`**（不要判 ok），
suggestion 寫明「輸出歸零、疑似解析 / 篩選邏輯失敗，請檢查抽取規則是否符合實際資料格式」。"""

    try:
        llm = _get_llm(role=llm_role)

        # 構建 message content（支援圖片 vision）
        if file_content["image_b64"]:
            content = [
                {"type": "text", "text": prompt_text},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{file_content['image_mime']};base64,{file_content['image_b64']}"
                    },
                },
            ]
        else:
            content = prompt_text

        from llm_factory import invoke_with_streaming
        raw = (await invoke_with_streaming(
            llm,
            [
                SystemMessage(content="你是一個精確的 pipeline 驗證器，只回傳 JSON 格式。"),
                HumanMessage(content=content),
            ],
            label=f"validator:{step_name}",
            timeout=300.0,
            logger=logger,
        )).strip()
        # 去除 markdown code block（如果有）
        if "```" in raw:
            parts = raw.split("```")
            raw = parts[1].strip()
            if raw.startswith("json"):
                raw = raw[4:].strip()

        data = json.loads(raw)
        result = ValidationResult(
            status=data.get("status", "failed"),
            reason=data.get("reason", ""),
            suggestion=data.get("suggestion", ""),
        )
        logger.info(f"[{step_name}] 驗證：{result.status} — {result.reason}")
        return result

    except Exception as e:
        # 429 / RESOURCE_EXHAUSTED：不退 fallback，直接回 rate_limited 給 runner，
        # 避免下一條 fallback 路徑又叫一次 LLM 再燒一次配額
        if _is_rate_limit_error(e):
            logger.error(f"[{step_name}] LLM 配額/速率受限（429）— 不退 fallback 避免燒光配額：{str(e)[:200]}")
            return ValidationResult(
                status="rate_limited",
                reason=f"LLM provider 配額用盡或速率受限（429）：{str(e)[:300]}",
                suggestion="等配額重置（通常每分鐘 / 每天）或在 Settings 切換 provider（Groq / OpenAI / Anthropic / Ollama 本地）",
            )
        logger.error(f"[{step_name}] LLM 驗證失敗：{e}，退回 exit code 判斷")
        # Fallback：純 exit code 判斷
        if exit_code == 0:
            return ValidationResult(
                status="ok",
                reason=f"Exit code 0（LLM 驗證服務暫時不可用：{e}）",
                suggestion="",
            )
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exit_code}（LLM 驗證服務暫時不可用：{e}）",
            suggestion="請檢查 log 檔取得詳細錯誤訊息",
        )


def _resolve_user_path(path: str) -> Path:
    """統一處理使用者可能給的三種路徑：
    - 絕對路徑 → 直接用
    - `~/xxx` → 展開到使用者家目錄
    - 相對路徑 → 以**專案根目錄**為基準（非 backend cwd），跟 runner 邏輯一致
    """
    p = Path(path).expanduser()
    if not p.is_absolute():
        _PROJ_ROOT = Path(__file__).parent.parent.parent.absolute()
        p = _PROJ_ROOT / p
    return p


def _check_output_file(path: Optional[str], step_start_time: Optional[str] = None) -> str:
    """取得輸出檔案或目錄的基本資訊。
    若 step_start_time(ISO) 有給,額外檢查 mtime 是否早於 step 開始時間 → 抓「舊檔冒充新檔」的假成功。
    """
    if not path:
        return "無需檢查"
    p = _resolve_user_path(path)
    if not p.exists():
        return "❌ 路徑不存在"
    # 解析 step_start_time → epoch seconds(供 mtime 比較)
    _start_ts = None
    if step_start_time:
        try:
            from datetime import datetime as _dt
            _start_ts = _dt.fromisoformat(step_start_time).timestamp()
        except Exception:
            _start_ts = None
    if p.is_dir():
        files = list(p.iterdir())
        if not files:
            return "⚠ 目錄存在但為空"
        total = sum(f.stat().st_size for f in files if f.is_file())
        _stale_note = ""
        if _start_ts is not None:
            newest = max((f.stat().st_mtime for f in files if f.is_file()), default=0)
            if newest > 0 and newest < _start_ts:
                _stale_note = "  ⚠️ **警告：目錄內所有檔案的 mtime 都早於本步驟開始時間，可能是先前 run 的舊產物、本次未實際更新**"
        return f"✅ 目錄存在，共 {len(files)} 個檔案，總大小：{total:,} bytes{_stale_note}"
    size = p.stat().st_size
    if size == 0:
        return "⚠ 檔案存在但為空（0 bytes）"
    _stale_note = ""
    if _start_ts is not None:
        mtime = p.stat().st_mtime
        if mtime < _start_ts:
            from datetime import datetime as _dt
            _mtime_iso = _dt.fromtimestamp(mtime).isoformat(timespec='seconds')
            _stale_note = (
                f"  ⚠️ **警告：檔案 mtime={_mtime_iso} 早於本步驟開始時間 {step_start_time[:19]}、"
                f"可能是先前 run 的舊產物、本次未實際更新 — 應判 failed**"
            )
    return f"✅ 檔案存在，大小：{size:,} bytes{_stale_note}"


# ── Skill 模式：ReAct Agent 驗證 ──────────────────────────────────────────────

# 危險命令黑名單（防止 LLM 生成危險操作）
_DANGEROUS_COMMANDS = {'rm', 'rmdir', 'del', 'format', 'mkfs', 'dd', 'kill', 'shutdown', 'reboot'}


def _run_python_sync(code: str) -> str:
    """在 subprocess 中執行 Python 程式碼，回傳 stdout + stderr。"""
    # 截斷混入程式碼中的 <tool> 標籤
    tool_tag_pos = code.find('<tool>')
    if tool_tag_pos > 0:
        code = code[:tool_tag_pos].rstrip()
    # 注入 done/view_image/read_file 的 no-op stub，避免 LLM 把工具名當 Python 函式呼叫而崩潰
    preamble = (
        "# -*- coding: utf-8 -*-\n"
        "import warnings\n"
        "warnings.filterwarnings('ignore')\n"
        "def done(*args, **kwargs):\n"
        "    print('[info] done() is a tool, not a Python function - ignored')\n"
        "def view_image(*args, **kwargs):\n"
        "    print('[info] view_image() is a tool, not a Python function - ignored')\n"
        "def read_file(*args, **kwargs):\n"
        "    print('[info] read_file() is a tool, not a Python function - ignored')\n"
    )
    code = preamble + code
    # wsl_docker 模式：路由到沙盒；沙盒不可用 / host 模式時 fallback 到下面 subprocess
    try:
        from pipeline.executor import _try_sandbox_exec
        sandbox_out = _try_sandbox_exec("run_python", code, None, "", None)
        if sandbox_out is not None:
            return sandbox_out
    except Exception:
        pass  # 沙盒模組壞了也 fallback 到 host subprocess
    try:
        # UTF-8 寫檔（見 executor.py 同樣 fix 的註解）
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
            f.write(code)
            tmp_path = f.name
        from pipeline.executor import _SKILL_PYTHON, _clean_env
        result = subprocess.run(
            [_SKILL_PYTHON, tmp_path],
            capture_output=True, text=True,
            timeout=SKILL_TOOL_TIMEOUT,
            env=_clean_env(),  # 套用 PYTHONIOENCODING=utf-8 防中文 print 爆炸
        )
        Path(tmp_path).unlink(missing_ok=True)
        output = ""
        if result.stdout:
            output += result.stdout
        if result.stderr:
            tag = "stderr" if result.returncode != 0 else "warnings"
            output += f"\n[{tag}]\n{result.stderr}"
        if result.returncode != 0:
            output += f"\n[exit code: {result.returncode}]"
        elif not result.stdout:
            output += "\n[執行成功，程式無 stdout 輸出]"
        return output.strip() or "(無輸出)"
    except subprocess.TimeoutExpired:
        Path(tmp_path).unlink(missing_ok=True)
        return f"[錯誤] Python 執行超時（>{SKILL_TOOL_TIMEOUT}秒）"
    except Exception as e:
        return f"[錯誤] Python 執行失敗：{e}"


def _run_shell_sync(cmd: str) -> str:
    """執行 shell 命令，回傳輸出。會過濾危險命令。
    wsl_docker 模式下透過 executor._try_sandbox_exec 路由到容器；
    否則走 host subprocess（原行為）。
    """
    first_word = cmd.strip().split()[0] if cmd.strip() else ""
    if first_word in _DANGEROUS_COMMANDS:
        return f"[拒絕] 命令 '{first_word}' 被安全策略封鎖"
    # 先試沙盒（如果 settings.skill_sandbox_mode='wsl_docker'）
    try:
        from pipeline.executor import _try_sandbox_exec
        sandbox_out = _try_sandbox_exec("run_shell", cmd, None, "", None)
        if sandbox_out is not None:
            return sandbox_out
    except Exception:
        pass  # 沙盒模組問題 → 繼續走 host fallback
    # 統一 python interpreter（與 executor._skill_run_shell 一致）
    from pipeline.executor import _rewrite_python_cmd
    cmd = _rewrite_python_cmd(cmd)
    try:
        result = subprocess.run(
            cmd, shell=True,
            capture_output=True, text=True,
            timeout=SKILL_TOOL_TIMEOUT,
        )
        output = ""
        if result.stdout:
            output += result.stdout
        if result.stderr:
            output += f"\n[stderr]\n{result.stderr}"
        return output.strip()[:5000] or "(無輸出)"
    except subprocess.TimeoutExpired:
        return f"[錯誤] 命令執行超時（>{SKILL_TOOL_TIMEOUT}秒）"
    except Exception as e:
        return f"[錯誤] 命令執行失敗：{e}"


def _read_file_sync(path: str, max_lines: int = 100) -> str:
    """讀取檔案內容（最多 max_lines 行）。"""
    try:
        # 清理 LLM 常見的錯誤格式：read_file("path"), 引號, 空白
        cleaned = path.strip()
        import re as _re
        m = _re.match(r'read_file\(["\']?(.+?)["\']?\)\s*$', cleaned)
        if m:
            cleaned = m.group(1)
        cleaned = cleaned.strip().strip('"').strip("'")
        # 沙盒路徑 → Windows 路徑（同 _view_image_sync 同份補丁）
        m_wsl = _re.match(r"^/mnt/([a-z])/(.*)$", cleaned)
        if m_wsl:
            cleaned = f"{m_wsl.group(1).upper()}:\\{m_wsl.group(2).replace('/', chr(92))}"
        p = Path(cleaned).expanduser()
        if not p.exists():
            return f"[錯誤] 檔案不存在：{path}（解析後：{p}）"
        if p.is_dir():
            files = sorted(p.iterdir())[:30]
            listing = "\n".join(f"  {'📁' if f.is_dir() else '📄'} {f.name} ({f.stat().st_size:,} B)" for f in files)
            return f"目錄內容：\n{listing}"
        # 偵測二進制檔案，避免汙染 LLM context
        binary_exts = {'.xlsx', '.xls', '.docx', '.pptx', '.pdf', '.png', '.jpg', '.jpeg', '.gif', '.zip', '.gz', '.tar', '.pkl', '.npy', '.parquet'}
        if p.suffix.lower() in binary_exts:
            size = p.stat().st_size
            return (f"[提示] {p.name} 是二進制檔案（{size:,} bytes），無法用 read_file 讀取。\n"
                    f"請改用 run_python 搭配適當的套件讀取：\n"
                    f"- .xlsx/.xls → pandas.read_excel() 或 openpyxl\n"
                    f"- .docx → python-docx\n"
                    f"- .png/.jpg → PIL 或 view_image 工具\n"
                    f"- .pdf → PyPDF2")
        if p.stat().st_size > 10 * 1024 * 1024:
            return f"[警告] 檔案過大（{p.stat().st_size:,} bytes），只讀前 {max_lines} 行"
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= max_lines:
                    lines.append(f"... (已截斷，共超過 {max_lines} 行)")
                    break
                lines.append(line.rstrip())
        return "\n".join(lines) or "(空檔案)"
    except Exception as e:
        return f"[錯誤] 讀取失敗：{e}"


def _sanitize_code(code: str) -> str:
    """清除混入程式碼中的 LLM 解釋文字（非 Python/Shell 語法的行）。"""
    lines = code.split('\n')
    code_pattern = re.compile(
        r'^(\s*(import |from |def |class |if |for |while |with |try:|except |'
        r'return |print|#|[a-zA-Z_]\w*\s*[=(]|plt\.|df\.|pd\.|np\.|sns\.|'
        r'\[|{|}|\]|\)|"|\'|$))'
    )
    start_idx = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        if code_pattern.match(stripped):
            start_idx = i
            break
    result = []
    for line in lines[start_idx:]:
        stripped = line.strip()
        if not stripped:
            result.append(line)
            continue
        first_char = stripped[0]
        if ord(first_char) > 0x2E00 and not stripped.startswith('#') and not stripped.startswith(("'", '"')):
            continue
        result.append(line)
    return '\n'.join(result).strip()


def _parse_tool_calls(text: str) -> list[dict]:
    """
    從 LLM 回覆中解析工具呼叫。
    支援：標準 <input> 標籤、code block 包裹、無標籤直接跟內容。
    關鍵：確保 run_python 只提取程式碼，不混入 LLM 解釋文字。
    """
    import re
    calls = []

    # Step 1：標準 <tool>...</tool> <input>...</input>
    pattern_std = re.compile(r'<tool>(.*?)</tool>\s*<input>(.*?)</input>', re.DOTALL)
    for m in pattern_std.finditer(text):
        calls.append({"tool": m.group(1).strip(), "input": m.group(2).strip()})
    if calls:
        return calls

    # Step 2：找所有 code blocks，再找離 <tool> 最近的那個
    code_blocks = list(re.finditer(r'```(?:python|json|bash|sh)?\s*\n(.*?)```', text, re.DOTALL))
    tool_tags = list(re.finditer(r'<tool>(.*?)</tool>', text))

    for tag in tool_tags:
        tool_name = tag.group(1).strip()
        tag_start = tag.start()
        tag_end = tag.end()

        # 先找 tag 之後最近的 code block
        best_block = None
        for block in code_blocks:
            if block.start() >= tag_end:
                best_block = block
                break

        # 如果 tag 之後沒有 code block，往前找最近的（LLM 先放 code 再放 tag）
        if not best_block:
            for block in reversed(code_blocks):
                if block.end() <= tag_start:
                    best_block = block
                    break

        if best_block:
            content = best_block.group(1).strip()
            if tool_name in ('run_python', 'run_shell'):
                content = _sanitize_code(content)
            if content and len(content) > 2:
                calls.append({"tool": tool_name, "input": content})
                return calls

    # Step 3：done 工具 — 找 JSON
    done_match = re.search(r'<tool>done</tool>', text)
    if done_match:
        after_done = text[done_match.end():]
        json_match = re.search(r'\{.*?\}', after_done, re.DOTALL)
        if json_match:
            return [{"tool": "done", "input": json_match.group(0).strip()}]

    # Step 4：沒有 <tool> 標籤但有 code block
    if not tool_tags and code_blocks:
        content = code_blocks[-1].group(1).strip()
        if content.startswith('{') and ('success' in content or 'status' in content):
            return [{"tool": "done", "input": content}]

    # Step 5：fallback
    cleaned = re.sub(r'```(?:python|json|bash|sh)?\s*\n?', '', text)
    cleaned = cleaned.replace('```', '')
    pattern_raw = re.compile(r'<tool>(.*?)</tool>\s*(.+?)(?=<tool>|$)', re.DOTALL)
    for m in pattern_raw.finditer(cleaned):
        tool_name = m.group(1).strip()
        content = m.group(2).strip()
        if tool_name in ('run_python', 'run_shell'):
            content = _sanitize_code(content)
        if content and len(content) > 2:
            calls.append({"tool": tool_name, "input": content})
            break
    return calls


IMAGE_EXTS_SKILL = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                    '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}


def _view_image_sync(path: str) -> dict:
    """
    讀取圖片並回傳 base64 資料。
    回傳 {"text": 描述, "image_b64": str|None, "image_mime": str|None}
    """
    try:
        cleaned = path.strip().strip('"').strip("'")
        # 沙盒路徑 → Windows 路徑（V3 view_image bug：LLM 跑沙盒給 /mnt/c/... 結果讀不到）
        import re as _re
        m = _re.match(r"^/mnt/([a-z])/(.*)$", cleaned)
        if m:
            cleaned = f"{m.group(1).upper()}:\\{m.group(2).replace('/', chr(92))}"
        p = Path(cleaned).expanduser()
        if not p.exists():
            return {"text": f"[錯誤] 圖片不存在：{path}（解析後：{p}）", "image_b64": None, "image_mime": None}
        ext = p.suffix.lower()
        if ext not in IMAGE_EXTS_SKILL:
            return {"text": f"[錯誤] 不支援的圖片格式：{ext}，支援 {list(IMAGE_EXTS_SKILL.keys())}", "image_b64": None, "image_mime": None}
        data = p.read_bytes()
        if len(data) > 20 * 1024 * 1024:
            return {"text": f"[錯誤] 圖片過大（{len(data):,} bytes，上限 20MB）", "image_b64": None, "image_mime": None}
        b64 = base64.b64encode(data).decode()
        mime = IMAGE_EXTS_SKILL[ext]
        return {"text": f"圖片 {p.name}（{len(data):,} bytes），已載入供視覺分析", "image_b64": b64, "image_mime": mime}
    except Exception as e:
        return {"text": f"[錯誤] 圖片讀取失敗：{e}", "image_b64": None, "image_mime": None}


def _execute_tool(tool_name: str, tool_input: str) -> str:
    """執行單一工具呼叫（非圖片工具）。"""
    if tool_name == "run_python":
        return _run_python_sync(tool_input)
    elif tool_name == "run_shell":
        return _run_shell_sync(tool_input)
    elif tool_name == "read_file":
        return _read_file_sync(tool_input.strip())
    elif tool_name == "done":
        return "__DONE__"
    elif tool_name == "view_image":
        return "__VIEW_IMAGE__"  # 特殊標記，在 agent loop 中處理
    else:
        return f"[錯誤] 未知工具：{tool_name}"


async def validate_step_with_skill(
    step_name: str,
    command: str,
    exit_code: int,
    stdout: str,
    stderr: str,
    output_path: Optional[str],
    output_expect: Optional[str],
    logger: logging.Logger,
    llm_role: str = "primary",
    step_start_time: Optional[str] = None,
) -> ValidationResult:
    """
    Skill 模式驗證：LLM 作為 ReAct agent，可主動執行程式碼來驗證步驟結果。

    工具：
    - run_python(code): 執行 Python 程式碼
    - run_shell(cmd): 執行 Shell 命令
    - read_file(path): 讀取檔案內容
    - done(json): 結束驗證並回傳結果
    """
    stdout_tail = stdout[-1500:] if len(stdout) > 1500 else stdout
    stderr_tail = stderr[-500:] if len(stderr) > 500 else stderr

    system_prompt = """你是一個 pipeline 步驟的 Skill 驗證 agent。
你可以主動執行程式碼來驗證步驟的輸出是否正確，而不只是被動地閱讀文字。

你有以下工具可用（一律透過原生 function calling 呼叫，不要寫任何文字協議格式）：

1. run_python(code) — 執行 Python 程式碼
   例：讀 csv、print 行數/欄位、檢查內容是否符合預期。

2. run_shell(command) — 執行系統命令
   注意：盡量用 run_python 代替 run_shell，因為 Python 是跨平台的。

3. read_file(path) — 讀取檔案內容。

4. view_image(path) — 查看圖片（視覺分析，支援 png/jpg/gif/webp）
   系統會將圖片顯示給你，你可以用視覺判斷圖片內容是否正確。
   適用場景：驗證圖表是否有標題、座標軸、資料是否合理、圖片是否正常渲染等。

5. done(status, reason, suggestion) — 結束驗證，回傳最終判定
   status 只能是 "ok"、"warning"、"failed" 三者之一；reason / suggestion 用中文。

## ⛔ Tool 呼叫協議（最高優先級）
本對話**使用原生 function calling API**、上述工具已透過 function_declarations 註冊。
你**必須**透過 API 的 function_call 機制呼叫工具、**禁止**寫 `<tool>name</tool>` / `<input>...</input>`
這類文字格式 — 那種純文字 orchestrator 不會解析、會被視為沒呼叫任何工具。直接 emit 結構化 tool_calls 即可。

【可用 Python 套件】
標準庫：csv, json, random, os, pathlib, re, math, datetime, io, collections
資料處理：pandas, numpy, openpyxl, xlrd, tabulate
文件處理：python-docx (docx), python-pptx (pptx), PyPDF2, reportlab, jinja2
網頁/爬蟲：requests, beautifulsoup4 (bs4), lxml
圖表繪製：matplotlib, seaborn, plotly
圖片處理：Pillow (PIL)
其他：faker, pyyaml, chardet

【matplotlib 繪圖注意事項】
- 使用 matplotlib.pyplot 時，務必在最前面加 `import matplotlib; matplotlib.use('Agg')` 以避免 GUI 問題
- boxplot 的 `labels` 參數已在新版棄用，請改用 `tick_labels`
- 繪製分組箱形圖時，需要先將資料按分組欄位 pivot/reshape，再分別傳入各組資料
- 中文顯示：macOS 使用 'PingFang HK'；Windows 使用 'Microsoft JhengHei' 或 'SimHei'
- 繪圖完成後務必呼叫 `plt.savefig(路徑, dpi=150, bbox_inches='tight')` 並 `plt.close()`

【重要規則】
- **路徑處理：一律使用 `pathlib.Path` 或 `os.path.join` 組合路徑，不要用字串拼接 `/`**
- **只使用上方列出的已安裝套件，不要安裝新套件**
- **絕對不要執行 sudo、pip install、apt 等安裝命令**
- 根據「預期輸出描述」主動驗證，不要只看 exit code
- **如果輸出路徑是圖片檔（.png/.jpg 等），一定要使用 view_image 工具查看圖片內容再做判斷**
- 可以多次呼叫工具，逐步分析
- 每次只呼叫一個工具
- 最後一定要呼叫 done 工具回傳結論
- status 只能是 "ok"、"warning"、"failed" 三者之一
- reason 和 suggestion 用中文

【Skill / Outlook agent 試錯歷史】
若 stdout 出現 `[Skill 完成]` 或 `[Outlook 完成]` 標記、且 exit code 為 0，代表 agent 已試錯到成功。
**stdout 裡的早期 Traceback 是試錯過程，並非最終狀態 — 不要因為這些 Traceback 就判 failed**。
此情境請主動 read_file / run_python 驗證輸出檔案是否符合預期，看「現在的檔案」而不是「歷史錯誤」。"""

    user_prompt = f"""請驗證以下 pipeline 步驟的執行結果：

【步驟資訊】
名稱：{step_name}
命令：{command}
Exit Code：{exit_code}
預期輸出描述：{output_expect or "無特定要求"}
輸出路徑：{output_path or "無"}

【stdout（最後部分）】
```
{stdout_tail or "（無輸出）"}
```

【stderr（最後部分）】
```
{stderr_tail or "（無輸出）"}
```

請使用工具主動驗證輸出是否符合預期。開始吧。"""

    try:
        from langchain_core.tools import tool as _lc_tool
        from langchain_core.messages import AIMessage, ToolMessage

        # ── 5 個驗證工具(native function calling)───────────────────────
        # 這些 wrapper 只負責「對 LLM 暴露乾淨 schema」、實際執行仍委派既有的
        # _execute_tool / _view_image_sync(host 端、行為不變)。view_image / done
        # 不在這裡真執行 — agent loop 看 tool_call.name 自己處理(多模態 / 回判定)。
        @_lc_tool
        def run_python(code: str) -> str:
            """執行 Python 程式碼來驗證輸出(讀檔、檢查行數/欄位/內容)。

            Args:
                code: 要執行的 Python 程式碼(完整可獨立執行)
            Returns:
                stdout + stderr(如有)
            """
            return _execute_tool("run_python", code)

        @_lc_tool
        def run_shell(command: str) -> str:
            """執行系統 shell 命令。盡量用 run_python 代替(跨平台)。

            Args:
                command: shell 命令
            Returns:
                stdout + stderr
            """
            return _execute_tool("run_shell", command)

        @_lc_tool
        def read_file(path: str) -> str:
            """讀取檔案內容供判斷。

            Args:
                path: 檔案路徑
            Returns:
                檔案內容(截斷時會說明)
            """
            return _execute_tool("read_file", path)

        @_lc_tool
        def view_image(path: str) -> str:
            """查看圖片(視覺分析、png/jpg/gif/webp)。輸出是圖片檔時必用此工具看內容再判斷。

            Args:
                path: 圖片檔路徑
            Returns:
                圖片會以多模態訊息顯示給你
            """
            return "__VIEW_IMAGE__"  # 不在此真執行、loop 看 tool name 自行注入多模態

        @_lc_tool
        def done(status: str, reason: str = "", suggestion: str = "") -> str:
            """結束驗證、回傳最終判定。

            Args:
                status: 驗證結論、只能是 "ok" / "warning" / "failed"
                reason: 判定理由(中文)
                suggestion: 修復建議(中文、failed 時才有意義)
            Returns:
                結束標記
            """
            return "__DONE__"

        validator_tools = [run_python, run_shell, read_file, view_image, done]

        llm = _get_llm(role=llm_role)
        try:
            llm_with_tools = llm.bind_tools(validator_tools)
        except Exception as _be:
            logger.error(f"[{step_name}] Skill 驗證 bind_tools 失敗:{_be}、退回一般驗證")
            raise  # 由最外層 except 接住 → fallback 到 validate_step

        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt),
        ]

        _empty_streak = 0  # 連續空回應計數:gemma 在 native-FC config 下若 tool_calls + content 皆空 → 早退一般驗證、別空轉到逾時
        for iteration in range(SKILL_MAX_ITERATIONS):
            logger.info(f"[{step_name}] Skill agent 迭代 {iteration + 1}/{SKILL_MAX_ITERATIONS} [NATIVE]")

            # 冷卻機制
            if iteration > 0 and iteration % SKILL_COOLDOWN_EVERY == 0:
                logger.info(f"[{step_name}] ⏸ 達到 {SKILL_COOLDOWN_EVERY} 次呼叫，冷卻 {SKILL_COOLDOWN_SECONDS} 秒...")
                await asyncio.sleep(SKILL_COOLDOWN_SECONDS)

            if iteration > 0:
                await asyncio.sleep(SKILL_REQUEST_INTERVAL)

            response: AIMessage = await asyncio.wait_for(
                llm_with_tools.ainvoke(messages), timeout=180.0
            )
            tool_calls = list(getattr(response, "tool_calls", []) or [])
            content_str = response.content if isinstance(response.content, str) else ""
            logger.info(
                f"[{step_name}] Agent 回覆(content {len(content_str)} 字, tool_calls={len(tool_calls)})"
            )
            if content_str:
                _vp = content_str if len(content_str) <= 2000 else content_str[:2000] + f"...[截斷、共 {len(content_str)} 字]"
                logger.debug(f"[{step_name}] Agent content：\n{_vp}")

            # AIMessage 進歷史(含 tool_calls、維持 tool_call/tool_result 配對)
            messages.append(response)

            if not tool_calls:
                # tool_calls 與 content 皆空 → 模型沒在配合(常見:gemma native-FC config 不吐工具也不吐字)。
                # 連 2 輪空 → 直接 raise、由 except 退回一般驗證,別空轉 15 輪 + 最後 180s 逾時。
                if len(content_str.strip()) < 5:
                    _empty_streak += 1
                    if _empty_streak >= 2:
                        raise RuntimeError(
                            f"skill 驗證 agent 連續 {_empty_streak} 輪空回應(tool_calls + content 皆空、疑似模型不相容)、提早退回一般驗證"
                        )
                else:
                    _empty_streak = 0
                # 有 content 但沒呼叫工具 → 提示繼續用 native function calling
                messages.append(HumanMessage(
                    content="請使用工具(原生 function calling)來驗證，或呼叫 done 工具回傳最終結論。不要用文字描述工具呼叫。"
                ))
                continue

            _empty_streak = 0

            # 掃出 done(若有)、其餘工具先執行;done 在最後處理
            done_call: Optional[dict] = None
            regular_calls = []
            for tc in tool_calls:
                if tc.get("name") == "done":
                    done_call = tc  # 多個 done 取最後一個
                else:
                    regular_calls.append(tc)

            for tc in regular_calls:
                tool_name = tc.get("name", "")
                tc_id = tc.get("id") or ""
                tc_args = tc.get("args", {}) or {}
                logger.info(f"[{step_name}] 執行工具 {tool_name} [NATIVE]")

                # view_image 特殊處理：注入多模態 HumanMessage
                if tool_name == "view_image":
                    _img_path = str(tc_args.get("path", "") or "")
                    img_data = await asyncio.get_event_loop().run_in_executor(
                        None, _view_image_sync, _img_path
                    )
                    logger.debug(f"[{step_name}] view_image：{img_data['text']}")
                    # ToolMessage 先回(維持 tool_call 配對)、圖片本體用後續 HumanMessage 多模態帶
                    messages.append(ToolMessage(
                        content=f"[view_image] {img_data['text']}", tool_call_id=tc_id
                    ))
                    if img_data["image_b64"]:
                        messages.append(HumanMessage(content=[
                            {"type": "text", "text": "請仔細觀察以下圖片內容，判斷是否符合預期。"},
                            {"type": "image_url", "image_url": {
                                "url": f"data:{img_data['image_mime']};base64,{img_data['image_b64']}"
                            }},
                        ]))
                    continue

                # 其餘工具：取出對應參數 → 呼叫既有 _execute_tool(host 端、行為不變)
                if tool_name == "run_python":
                    _ti = str(tc_args.get("code", "") or "")
                elif tool_name == "run_shell":
                    _ti = str(tc_args.get("command", "") or "")
                elif tool_name == "read_file":
                    _ti = str(tc_args.get("path", "") or "")
                else:
                    _ti = ""

                tool_result = await asyncio.get_event_loop().run_in_executor(
                    None, _execute_tool, tool_name, _ti
                )
                _vt = tool_result if len(tool_result) <= 3000 else tool_result[:3000] + f"...[已截斷，完整長度 {len(tool_result)} 字]"
                logger.debug(f"[{step_name}] 工具結果：\n{_vt}")

                messages.append(ToolMessage(content=tool_result, tool_call_id=tc_id))

            # done 工具 → 直接讀 args 拿判定(免 JSON parse)
            if done_call is not None:
                _da = done_call.get("args", {}) or {}
                result = ValidationResult(
                    status=_da.get("status", "failed") or "failed",
                    reason=_da.get("reason", "") or "",
                    suggestion=_da.get("suggestion", "") or "",
                )
                logger.info(f"[{step_name}] Skill 驗證完成：{result.status} — {result.reason}")
                return result

        # 超過最大迭代次數
        logger.warning(f"[{step_name}] Skill agent 達到最大迭代次數")
        return ValidationResult(
            status="warning",
            reason=f"Skill agent 在 {SKILL_MAX_ITERATIONS} 次迭代內未完成驗證",
            suggestion="建議手動檢查輸出結果",
        )

    except Exception as e:
        # 429 / RESOURCE_EXHAUSTED：直接回 rate_limited，不退 validate_step（會再 429 一次）
        if _is_rate_limit_error(e):
            logger.error(f"[{step_name}] Skill 驗證 LLM 配額/速率受限（429）— 不退一般驗證避免燒光配額：{str(e)[:200]}")
            return ValidationResult(
                status="rate_limited",
                reason=f"LLM provider 配額用盡或速率受限（429）：{str(e)[:300]}",
                suggestion="等配額重置或在 Settings 切換 provider（Groq / OpenAI / Anthropic / Ollama 本地）",
            )
        logger.error(f"[{step_name}] Skill 驗證失敗：{e}，退回一般驗證")
        # Fallback to standard validation
        return await validate_step(
            step_name=step_name,
            command=command,
            exit_code=exit_code,
            stdout=stdout,
            stderr=stderr,
            output_path=output_path,
            output_expect=output_expect,
            logger=logger,
            step_start_time=step_start_time,
        )
