"""
【工具 D】電商銷售管理報表產生器
製作者：電商營運部 (E-commerce Ops) / 商務長辦公室

用途：把 Stage 3 的分析包裝成精美 Excel 管理報表，
      含封面、KPI 摘要、各維度明細，並嵌入 matplotlib 圖表
      （類別圓餅、Top 商品長條、每日趨勢折線、星期分布）。

輸入：~/ai_output/ecommerce/sales_summary.xlsx
輸出：~/ai_output/ecommerce/Sales_Report.xlsx
"""
import sys
import io as _io_e2e
try:
    sys.stdout = _io_e2e.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass

import os
import sys
import tempfile
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from dotenv import load_dotenv
from pathlib import Path

load_dotenv(Path(__file__).parent.parent.parent / "backend" / ".env")


def get_paths():
    env_run_dir = os.getenv("PIPELINE_OUTPUT_DIR")
    if env_run_dir:
        return (
            os.path.join(env_run_dir, "sales_summary.xlsx"),
            os.path.join(env_run_dir, "Sales_Report.xlsx"),
        )
    base_path = os.getenv("OUTPUT_BASE_PATH")
    if base_path:
        if not os.path.isabs(base_path):
            base_path = os.path.join(Path(__file__).parent.parent.parent, base_path)
        return (
            os.path.join(base_path, "ecommerce", "sales_summary.xlsx"),
            os.path.join(base_path, "ecommerce", "Sales_Report.xlsx"),
        )
    project_root = Path(__file__).parent.parent.parent
    return (
        os.path.join(project_root, "ai_output", "ecommerce", "sales_summary.xlsx"),
        os.path.join(project_root, "ai_output", "ecommerce", "Sales_Report.xlsx"),
    )


INPUT, OUTPUT = get_paths()

if not os.path.exists(INPUT):
    print(f"[ERROR] 找不到輸入檔案：{INPUT}")
    print("        請先執行 stage3_analyze_sales.py")
    sys.exit(1)

plt.rcParams["font.sans-serif"] = ["Noto Sans CJK TC", "Microsoft JhengHei", "PingFang TC", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

kpi_df    = pd.read_excel(INPUT, sheet_name="KPI 總覽")
cat_df    = pd.read_excel(INPUT, sheet_name="類別別營收")
top_df    = pd.read_excel(INPUT, sheet_name="Top 10 商品")
region_df = pd.read_excel(INPUT, sheet_name="地區別")
pay_df    = pd.read_excel(INPUT, sheet_name="付款方式")
seg_df    = pd.read_excel(INPUT, sheet_name="時段別")
dow_df    = pd.read_excel(INPUT, sheet_name="星期別")
daily_df  = pd.read_excel(INPUT, sheet_name="每日趨勢")


def kpi(name):
    row = kpi_df[kpi_df["指標"] == name]
    return row["值"].values[0] if len(row) else None


total_rev = kpi("總營收 (NT$)")
total_orders = kpi("總訂單數")
aov = kpi("客單價 (NT$)")

# ── 圖表 ────────────────────────────────────────────────────────────
tmp_dir = tempfile.mkdtemp(prefix="ec_charts_")
chart_paths = {}

# 1. 類別圓餅
fig, ax = plt.subplots(figsize=(7, 5))
colors_pie = ["#2E75B6", "#70AD47", "#ED7D31", "#FFC000", "#7030A0", "#C0504D"]
ax.pie(cat_df["營收 (NT$)"],
       labels=[f"{c} {p}%" for c, p in zip(cat_df["類別"], cat_df["占比 (%)"])],
       colors=colors_pie[:len(cat_df)],
       startangle=90, wedgeprops=dict(linewidth=2, edgecolor="white"))
ax.set_title("各類別營收占比", fontsize=14, fontweight="bold")
plt.tight_layout()
p = os.path.join(tmp_dir, "category_pie.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["category"] = p

# 2. Top 10 商品水平 bar
fig, ax = plt.subplots(figsize=(9, 5))
ax.barh(top_df["商品"][::-1], top_df["營收 (NT$)"][::-1], color="#2E75B6")
ax.set_xlabel("營收 (NT$)")
ax.set_title("Top 10 暢銷商品", fontsize=14, fontweight="bold")
for i, v in enumerate(top_df["營收 (NT$)"][::-1]):
    ax.text(v, i, f" {v:,.0f}", va="center", fontsize=8)
plt.tight_layout()
p = os.path.join(tmp_dir, "top10.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["top10"] = p

# 3. 每日營收 + 毛利雙線
fig, ax = plt.subplots(figsize=(10, 5))
dates = pd.to_datetime(daily_df["Order_Date"]).dt.strftime("%m/%d")
ax.plot(dates, daily_df["營收 (NT$)"], marker="o", color="#2E75B6", label="營收", linewidth=2)
ax.plot(dates, daily_df["毛利 (NT$)"], marker="s", color="#70AD47", label="毛利", linewidth=2)
ax.set_xlabel("日期")
ax.set_ylabel("NT$")
ax.set_title("每日營收 / 毛利趨勢", fontsize=14, fontweight="bold")
ax.legend(loc="upper left")
ax.grid(True, alpha=0.3)
# X 軸標籤每 3 天顯示一次
for i, lab in enumerate(ax.get_xticklabels()):
    if i % 3 != 0:
        lab.set_visible(False)
plt.xticks(rotation=45)
plt.tight_layout()
p = os.path.join(tmp_dir, "daily_trend.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["daily"] = p

# 4. 星期別長條
fig, ax = plt.subplots(figsize=(8, 4))
dow_zh_map = {
    "Monday": "週一", "Tuesday": "週二", "Wednesday": "週三",
    "Thursday": "週四", "Friday": "週五", "Saturday": "週六", "Sunday": "週日",
}
labels = [dow_zh_map.get(d, d) for d in dow_df["星期"]]
weekend_mask = [d in ("週六", "週日") for d in labels]
colors_dow = ["#ED7D31" if w else "#2E75B6" for w in weekend_mask]
ax.bar(labels, dow_df["營收 (NT$)"], color=colors_dow)
ax.set_ylabel("營收 (NT$)")
ax.set_title("星期別營收（週末為橘色）", fontsize=13, fontweight="bold")
for i, v in enumerate(dow_df["營收 (NT$)"]):
    ax.text(i, v, f"{v:,.0f}", ha="center", va="bottom", fontsize=8)
plt.tight_layout()
p = os.path.join(tmp_dir, "dow.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["dow"] = p

# ── 樣式 ────────────────────────────────────────────────────────────
NAVY = PatternFill("solid", fgColor="1F3864")
LTBLUE = PatternFill("solid", fgColor="BDD7EE")

TITLE_FONT = Font(bold=True, color="FFFFFF", size=18)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(bold=True, size=12)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

# ── 寫入 + 美化 ────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    pd.DataFrame([[" "]]).to_excel(writer, sheet_name="封面", index=False, header=False)
    kpi_df.to_excel(writer, sheet_name="KPI 摘要", index=False)
    cat_df.to_excel(writer, sheet_name="類別別", index=False)
    top_df.to_excel(writer, sheet_name="Top 10 商品", index=False)
    region_df.to_excel(writer, sheet_name="地區別", index=False)
    pay_df.to_excel(writer, sheet_name="付款方式", index=False)
    seg_df.to_excel(writer, sheet_name="時段別", index=False)
    dow_df.to_excel(writer, sheet_name="星期別", index=False)
    daily_df.to_excel(writer, sheet_name="每日趨勢", index=False)

wb = load_workbook(OUTPUT)


def make_cover(ws):
    ws["A1"] = "電商銷售管理報表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = NAVY
    ws["A1"].alignment = CENTER
    ws.merge_cells("A1:F3")
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 25

    ws["A5"] = "報告產出時間"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A6"] = "資料區間"
    ws["B6"] = str(kpi("資料區間"))
    ws["A7"] = "總營收 (NT$)"
    ws["B7"] = f"{int(total_rev):,}"
    ws["A8"] = "總訂單數"
    ws["B8"] = f"{int(total_orders):,}"
    ws["A9"] = "客單價 (NT$)"
    ws["B9"] = f"{int(aov):,}"
    ws["A10"] = "毛利率"
    ws["B10"] = f"{kpi('毛利率 (%)'):.2f} %"

    for r in range(5, 11):
        ws.cell(r, 1).font = SUB_FONT
        ws.cell(r, 1).fill = LTBLUE
        ws.cell(r, 2).font = Font(size=12)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35


make_cover(wb["封面"])


def style_header_row(ws):
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = NAVY
        cell.alignment = CENTER
        cell.border = BORDER
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, max_len + 4))


for sn in ["KPI 摘要", "類別別", "Top 10 商品", "地區別", "付款方式", "時段別", "星期別", "每日趨勢"]:
    style_header_row(wb[sn])


# 千分號數字格式
def format_money(ws, col_names=("營收 (NT$)", "毛利 (NT$)", "占比 (%)")):
    headers = [c.value for c in ws[1]]
    for col_name in col_names:
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, col_idx)
            if "%" in col_name:
                cell.number_format = "0.00"
            else:
                cell.number_format = "#,##0"


for sn in ["類別別", "Top 10 商品", "地區別", "付款方式", "時段別", "星期別", "每日趨勢"]:
    format_money(wb[sn])


# ── 嵌入圖 ─────────────────────────────────────────────────────────
def insert_chart(ws, png_path, anchor_cell, w=640, h=360):
    img = XLImage(png_path)
    img.width = w
    img.height = h
    ws.add_image(img, anchor_cell)


insert_chart(wb["類別別"],     chart_paths["category"], "I2")
insert_chart(wb["Top 10 商品"], chart_paths["top10"],    "I2", w=720)
insert_chart(wb["每日趨勢"],   chart_paths["daily"],    "F2", w=800)
insert_chart(wb["星期別"],     chart_paths["dow"],      "F2")

wb.active = wb.sheetnames.index("封面")
wb.save(OUTPUT)

print(f"[Stage 4] ✅ 銷售管理報表產生完成")
print(f"  總營收 NT$ {int(total_rev):,} / 訂單 {int(total_orders):,} / AOV NT$ {int(aov):,}")
print(f"  含 9 個工作表 + 4 張嵌入圖表")
print(f"  輸出：{OUTPUT}")
