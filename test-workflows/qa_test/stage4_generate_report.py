"""
【工具 D】產線品質測試管理報表產生器
製作者：產線品保部門 (QA Team) / 廠長辦公室

用途：把 Stage 3 的分析結果包裝成精美 Excel 管理報表，
      含封面、KPI 摘要、各維度明細，並嵌入 matplotlib 圖表（PNG 嵌入）。

輸入：~/ai_output/qa_test/quality_summary.xlsx
輸出：~/ai_output/qa_test/QA_quality_report.xlsx
"""
import sys
import io as _io_e2e
try:
    sys.stdout = _io_e2e.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass

import os
import sys
import tempfile
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")  # 無 X server 環境
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from dotenv import load_dotenv
from pathlib import Path

load_dotenv(Path(__file__).parent.parent.parent / "backend" / ".env")


def get_paths():
    env_run_dir = os.getenv("PIPELINE_OUTPUT_DIR")
    if env_run_dir:
        return (
            os.path.join(env_run_dir, "quality_summary.xlsx"),
            os.path.join(env_run_dir, "QA_quality_report.xlsx"),
        )
    base_path = os.getenv("OUTPUT_BASE_PATH")
    if base_path:
        if not os.path.isabs(base_path):
            base_path = os.path.join(Path(__file__).parent.parent.parent, base_path)
        return (
            os.path.join(base_path, "qa_test", "quality_summary.xlsx"),
            os.path.join(base_path, "qa_test", "QA_quality_report.xlsx"),
        )
    project_root = Path(__file__).parent.parent.parent
    return (
        os.path.join(project_root, "ai_output", "qa_test", "quality_summary.xlsx"),
        os.path.join(project_root, "ai_output", "qa_test", "QA_quality_report.xlsx"),
    )


INPUT, OUTPUT = get_paths()

if not os.path.exists(INPUT):
    print(f"[ERROR] 找不到輸入檔案：{INPUT}")
    print("        請先執行 stage3_analyze_quality.py")
    sys.exit(1)

# ── 中文字型（matplotlib）──────────────────────────────────────────
# sandbox 容器有裝 Noto Sans CJK；host 走系統字型 fallback
plt.rcParams["font.sans-serif"] = ["Noto Sans CJK TC", "Microsoft JhengHei", "PingFang TC", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# ── 讀取分析結果 ───────────────────────────────────────────────────
kpi_df    = pd.read_excel(INPUT, sheet_name="KPI 總覽")
batch_df  = pd.read_excel(INPUT, sheet_name="批次別良率")
item_df   = pd.read_excel(INPUT, sheet_name="測項不良率排行")
line_df   = pd.read_excel(INPUT, sheet_name="產線別良率")
op_df     = pd.read_excel(INPUT, sheet_name="操作員別良率")
weekly_df = pd.read_excel(INPUT, sheet_name="週 trend")


def kpi(name):
    row = kpi_df[kpi_df["指標"] == name]
    return row["值"].values[0] if len(row) else 0


fpy = kpi("項目良率 (FPY %)")
unit_yield = kpi("整機良率 (Unit %)")

# ── 產出圖表 PNG（暫存）─────────────────────────────────────────────
tmp_dir = tempfile.mkdtemp(prefix="qa_charts_")
chart_paths = {}

# 1. 批次別良率長條圖
fig, ax = plt.subplots(figsize=(8, 4))
ax.bar(batch_df["Batch"], batch_df["良率 (%)"], color="#2E75B6")
ax.set_ylim(0, 105)
ax.set_ylabel("良率 (%)")
ax.set_title("各批次良率")
for i, v in enumerate(batch_df["良率 (%)"]):
    ax.text(i, v + 1, f"{v}%", ha="center", fontsize=9)
plt.tight_layout()
p = os.path.join(tmp_dir, "batch_yield.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["batch"] = p

# 2. 測項不良率水平 bar
fig, ax = plt.subplots(figsize=(8, 4))
items = item_df["Test_Item"]
fails = item_df["不良率 (%)"]
colors = ["#C0504D" if r >= 5 else "#ED7D31" if r >= 2 else "#70AD47" for r in fails]
ax.barh(items, fails, color=colors)
ax.set_xlabel("不良率 (%)")
ax.set_title("各測項不良率排行")
for i, v in enumerate(fails):
    ax.text(v + 0.1, i, f"{v}%", va="center", fontsize=9)
plt.tight_layout()
p = os.path.join(tmp_dir, "item_fail.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["item"] = p

# 3. 週 trend 折線
fig, ax = plt.subplots(figsize=(8, 4))
ax.plot(weekly_df["Week"], weekly_df["良率 (%)"], marker="o", linewidth=2, color="#2E75B6")
ax.set_ylim(min(80, weekly_df["良率 (%)"].min() - 5), 105)
ax.set_xlabel("Week")
ax.set_ylabel("良率 (%)")
ax.set_title("各週良率趨勢")
ax.grid(True, alpha=0.3)
for x, y in zip(weekly_df["Week"], weekly_df["良率 (%)"]):
    ax.text(x, y + 0.5, f"{y}%", ha="center", fontsize=8)
plt.tight_layout()
p = os.path.join(tmp_dir, "weekly_trend.png")
plt.savefig(p, dpi=150, bbox_inches="tight")
plt.close()
chart_paths["weekly"] = p

# ── 樣式 ────────────────────────────────────────────────────────────
NAVY = PatternFill("solid", fgColor="1F3864")
BLUE = PatternFill("solid", fgColor="2E75B6")
LTBLUE = PatternFill("solid", fgColor="BDD7EE")
GREEN = PatternFill("solid", fgColor="70AD47")
RED = PatternFill("solid", fgColor="C0504D")
GREY_LT = PatternFill("solid", fgColor="F2F2F2")

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=18)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(bold=True, size=12)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# ── 寫入 Excel + 套樣式 ─────────────────────────────────────────────
# 先用 pandas 寫各 sheet（內容），再用 openpyxl 美化
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    pd.DataFrame([[" "]]).to_excel(writer, sheet_name="封面", index=False, header=False)
    kpi_df.to_excel(writer, sheet_name="KPI 摘要", index=False)
    batch_df.to_excel(writer, sheet_name="各批次良率", index=False)
    item_df.to_excel(writer, sheet_name="測項不良率", index=False)
    line_df.to_excel(writer, sheet_name="產線別", index=False)
    op_df.to_excel(writer, sheet_name="操作員別", index=False)
    weekly_df.to_excel(writer, sheet_name="週 trend", index=False)

wb = load_workbook(OUTPUT)


# ── 封面頁 ──────────────────────────────────────────────────────────
def make_cover(ws):
    ws["A1"] = "產線品質測試管理報表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = NAVY
    ws["A1"].alignment = CENTER
    ws.merge_cells("A1:F3")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25

    ws["A5"] = "報告產出時間"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A6"] = "資料區間"
    ws["B6"] = "2024 Q4 — 全部批次"
    ws["A7"] = "整體項目良率 (FPY)"
    ws["B7"] = f"{fpy} %"
    ws["A8"] = "整機良率 (Unit Yield)"
    ws["B8"] = f"{unit_yield} %"

    for r in range(5, 9):
        ws.cell(r, 1).font = SUB_FONT
        ws.cell(r, 1).fill = LTBLUE
        ws.cell(r, 2).font = Font(size=12)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30


make_cover(wb["封面"])


# ── 一般工作表的標題列樣式 ─────────────────────────────────────────
def style_header_row(ws):
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = NAVY
        cell.alignment = CENTER
        cell.border = BORDER
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(r, col_idx).value)) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(35, max_len + 4))


for sn in ["KPI 摘要", "各批次良率", "測項不良率", "產線別", "操作員別", "週 trend"]:
    style_header_row(wb[sn])


# ── 良率欄位 condition formatting（綠/橘/紅）──────────────────────
def color_yield_cells(ws, col_name="良率 (%)"):
    headers = [c.value for c in ws[1]]
    if col_name not in headers:
        return
    col_idx = headers.index(col_name) + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, col_idx)
        v = cell.value
        try:
            v = float(v)
        except Exception:
            continue
        if v >= 95:
            cell.fill = GREEN
            cell.font = Font(bold=True, color="FFFFFF")
        elif v >= 85:
            cell.fill = PatternFill("solid", fgColor="FFC000")
            cell.font = Font(bold=True)
        else:
            cell.fill = RED
            cell.font = Font(bold=True, color="FFFFFF")


for sn in ["各批次良率", "測項不良率", "產線別", "操作員別", "週 trend"]:
    color_yield_cells(wb[sn])

# 測項不良率還要把「不良率」欄反向上色（高=紅）
def color_fail_cells(ws):
    headers = [c.value for c in ws[1]]
    if "不良率 (%)" not in headers:
        return
    col_idx = headers.index("不良率 (%)") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, col_idx)
        try:
            v = float(cell.value)
        except Exception:
            continue
        if v >= 5:
            cell.fill = RED
            cell.font = Font(bold=True, color="FFFFFF")
        elif v >= 2:
            cell.fill = PatternFill("solid", fgColor="FFC000")
            cell.font = Font(bold=True)
        else:
            cell.fill = GREEN
            cell.font = Font(bold=True, color="FFFFFF")


color_fail_cells(wb["測項不良率"])

# ── 嵌入圖表到對應工作表 ────────────────────────────────────────────
def insert_chart(ws, png_path, anchor_cell):
    img = XLImage(png_path)
    # 微調尺寸（Excel 一格大約 64x18 px）
    img.width = 640
    img.height = 320
    ws.add_image(img, anchor_cell)


# 圖貼在每個工作表的 K2 / 對應位置
insert_chart(wb["各批次良率"], chart_paths["batch"], "I2")
insert_chart(wb["測項不良率"], chart_paths["item"], "I2")
insert_chart(wb["週 trend"], chart_paths["weekly"], "I2")

# ── 把封面頁設成預設開啟 ────────────────────────────────────────────
wb.active = wb.sheetnames.index("封面")

wb.save(OUTPUT)

print(f"[Stage 4] ✅ 管理報表產生完成")
print(f"  項目良率：{fpy}% / 整機良率：{unit_yield}%")
print(f"  含 7 個工作表 + 3 張嵌入圖表")
print(f"  輸出：{OUTPUT}")
